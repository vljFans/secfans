from django.shortcuts import render
from django.core.files.storage import default_storage
# Create your views here.
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.response import Response
from . import models
from django.contrib.messages import get_messages
from django.contrib.auth import authenticate, login, logout
from django.core import serializers
from django.http import JsonResponse
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.hashers import make_password, check_password
from datetime import datetime, timedelta,timezone,date
from django.utils.timezone import now
from openpyxl import Workbook, load_workbook
from django.db import transaction
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from pathlib import Path
from decimal import Decimal
import os
import math
import environ
import csv
from fpdf import FPDF
from django.db.models import Avg, Count, Min, Sum , Case, When, DecimalField, Q, F, IntegerField, Max, Func, Subquery
from django.db.models.functions import Substr, Cast, StrIndex, Length
from fractions import Fraction
import pandas as pd
from django.contrib.auth.models import Permission
import re
from django.http import HttpResponse

import traceback


env = environ.Env()
environ.Env.read_env()
format = lambda x: f'{x.normalize():f}'
current_year = datetime.today().strftime('%Y')

class CustomPaginator:
    def __init__(self, items, per_page):
        self.items = items
        self.per_page = per_page

    def get_page(self, page_number):
        page_number = int(page_number)  # Convert to integer
        start = (page_number - 1) * self.per_page
        end = start + self.per_page
        page_items = self.items[start:end]
        return page_items

    def get_total_pages(self):
        return math.ceil(len(self.items) / self.per_page)


# def ai_digit_5():
#     return str((models.Store_Transaction.objects.annotate(
#         num_part=Cast(Substr('transaction_number', 13, 5), IntegerField())
#     ).aggregate(max_value=Max('num_part'))['max_value'] or 0) ).zfill(5)

def ai_digit_5():
    return str((models.Store_Transaction.objects.filter(status=1, deleted=0).annotate(
        num_part=Cast(Substr('transaction_number', Length('transaction_number') - 4,5), IntegerField())
    ).aggregate(max_value=Max('num_part'))['max_value'] or 0) + 1).zfill(5)

def handle_empty_cell(x):
    if isinstance(x, str):
        if x.strip().isdigit():
            return Decimal(x)
        else :
            return 0
    if isinstance(x, float) and math.isnan(x):
        return 0
    return Decimal(x)


def set_user_permissions_in_session(user, request):
    # Fetch the role associated with the user
    role = user.role

    # If the user has a role, fetch the permissions associated with that role
    if role:
        permissions = models.Role_Permission.objects.filter(role=role, status=1, deleted=0, permitted=1)
        
        # Create a list of dictionaries to mimic what you're checking in `get_session_permission`
        role_permissions = [{'permission__codename': perm.permission.codename} for perm in permissions]

        # Store this list in the session
        request.session['role_permissions'] = role_permissions
    else:
        # If no role is assigned, the user has no permissions
        request.session['role_permissions'] = []


def user_log_details_add(user,task_name):
    try:
        with transaction.atomic():
            user_log_details = models.User_Log_Details()
            user_log_details.user_id = user
            user_log_details.task_name = task_name
            user_log_details.save()
       
    except Exception as e:
        #print(f'Something went wrong: {e}')
        transaction.rollback()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def loginUser(request):
    context = {}
    try:
        user = models.User.objects.get(pk=request.user.id)
        if user is not None:
            login(request, user)

            # Set user permissions in session
            set_user_permissions_in_session(user, request)

            context.update({'status': 200, 'message': "", 'user_id': user.pk})
        else:
            context.update({'status': 501, 'message': "User Not Found."})
    except models.User.DoesNotExist:
        context.update({'status': 501, 'message': "User Not Found."})
    
    return Response(context)



# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def loginUser(request):
#     context = {}
#     user = models.User.objects.get(pk=request.user.id)
#     if user is not None:
#         login(request, user)
#         context.update({'status': 200, 'message': ""})
#     else:
#         context.update({'status': 501, 'message': "User Not Found."})
#     return Response(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logoutUser(request):
    context = {}
    logout(request)
    try:
        del request.session
    except:
        pass
    try:
        storage = get_messages(request)
        for message in storage:
            message = ''
        storage.used = False
    except:
        pass
    context.update({'status': 200, 'message': "Logged Out Successfully."})
    return Response(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getUserDetails(request):
    context = {}
    context.update({'status': 200, 'message': "User Details Fetched Successfully.",
                   'userDetails': serializers.serialize('json', [request.user])})
    return JsonResponse(context, safe=False)


@api_view(['GET'])
def getContentTypes(request):
    context = {}
    page_items = ContentType.objects.prefetch_related('permission_set').filter(app_label='api').exclude(
        model__in=['user', 'role', 'role_permission', 'country', 'state', 'city', 'customer_type', 'kyc_type', 'child_uom', 'bill_of_material_detail', 'purchase_order_detail', 'transaction_type', 'store_transaction_detail','job_order_detail_sent'])
    context.update(
        {'status': 200, 'message': "Content Types Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['GET'])
def getCustomerTypes(request):
    context = {}
    page_items = models.Customer_Type.objects.all()
    context.update(
        {'status': 200, 'message': "Customer Types Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['GET'])
def getKycTypes(request):
    context = {}
    page_items = models.KYC_Type.objects.all()
    context.update(
        {'status': 200, 'message': "KYC Types Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['GET'])
def getCountries(request):
    context = {}
    page_items = models.Country.objects.all()
    context.update(
        {'status': 200, 'message': "Countries Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['POST'])
def getCountryStates(request):
    context = {}
    country_id = request.POST['country_id']
    page_items = models.State.objects.filter(country_id=country_id)
    context.update(
        {'status': 200, 'message': "States Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['POST'])
def getStateCities(request):
    context = {}
    state_id = request.POST['state_id']
    page_items = models.City.objects.filter(state_id=state_id)
    context.update(
        {'status': 200, 'message': "Cities Fetched Successfully", 'page_items': serializers.serialize('json', page_items)})
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def roleList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        role = list(models.Role.objects.filter(pk=id)[:1].values('pk', 'name'))
        context.update({
            'status': 200,
            'message': "Role Fetched Successfully.",
            'page_items': role,
        })
    else:
        if keyword is not None and keyword != "":
            roles = list(models.Role.objects.filter(
                name__icontains=keyword, status=1, deleted=0).values('pk', 'name'))
        else:
            roles = list(models.Role.objects.filter(
                status=1, deleted=0).values('pk', 'name'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Roles Fetched Successfully.",
                'page_items': roles,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(roles, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Roles Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def roleAdd(request):
    context = {}
    userId = request.COOKIES.get('userId', None)
    if not request.POST['name']:
        context.update({
            'status': 502,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Role.objects.filter(
        name__iexact=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 503,
            'message': "Role with this name already exists."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            role = models.Role()
            role.name = request.POST['name']
            role.save()
            role_permission_details = []
            for index, elem in enumerate(request.POST.getlist('permission')):
                role_permission_details.append(models.Role_Permission(
                    permission_id=elem, role_id=role.id, permitted=1))
            models.Role_Permission.objects.bulk_create(role_permission_details)
            user_log_details_add(userId,'Add Role')

        transaction.commit()
        context.update({
            'status': 200,
            'message': "Role Created Successfully."
        })
    except Exception:
        context.update({
            'status': 504,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def roleEdit(request):
    context = {}
    userId = request.COOKIES.get('userId', None)

    if not request.POST['name']:
        context.update({
            'status': 505,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Role.objects.filter(
        name__iexact=request.POST['name']).exclude(id=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 506,
            'message': "Role with this name already exists."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            role = models.Role.objects.prefetch_related(
                'role_permission_set').get(pk=request.POST['id'])
            role.name = request.POST['name']
            role.updated_at = datetime.now()
            role.save()
            role.role_permission_set.all().delete()
            role_permission_details = []
            for index, elem in enumerate(request.POST.getlist('permission')):
                role_permission_details.append(models.Role_Permission(
                    permission_id=elem, role_id=role.id, permitted=1))
            models.Role_Permission.objects.bulk_create(role_permission_details)
            user_log_details_add(userId,'Edit Role ')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Role Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 507,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def roleDelete(request):
    context = {}
    userId = request.COOKIES.get('userId', None)
    

    role = models.Role.objects.prefetch_related(
        'role_permission_set', 'user_set').get(pk=request.POST['id'])
    if len(role.user_set.all()) > 0:
        context.update({
            'status': 507,
            'message': "Can not delete as " + str(len(role.user_set.all())) + " user(s) exist with this role."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            role.delete()
            user_log_details_add(userId,'delete Role ')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Role Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 508,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def userList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        user = list(models.User.objects.filter(pk=id)[:1].values(
            'pk', 'name', 'email', 'phone', 'role__name'))
        context.update({
            'status': 200,
            'message': "User Fetched Successfully.",
            'page_items': user,
        })
    else:
        if keyword is not None and keyword != "":
            users = list(models.User.objects.filter(
                Q(name__icontains=keyword) | Q(email__icontains=keyword) | Q(
                    phone__icontains=keyword) | Q(role__name__icontains=keyword)
            ).filter(status=1, deleted=0).exclude(is_superuser=1).values('pk', 'name', 'email', 'phone', 'role__name'))
        else:
            users = list(models.User.objects.filter(status=1, deleted=0).exclude(
                is_superuser=1).values('pk', 'name', 'email', 'phone', 'role__name'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Users Fetched Successfully.",
                'page_items': users,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(users, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Users Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def userAdd(request):
    context = {}
    userId = request.COOKIES.get('userId', None)

    if not request.POST['name'] or not request.POST['role_id'] or not request.POST['email'] or not request.POST['phone'] or not request.POST['password'] or not request.POST['confirm_password']:
        context.update({
            'status': 508,
            'message': "Name/Role/Email/Phone/Password/Confirmed Password has not been provided."
        })
        return JsonResponse(context)
    if request.POST['password'] != request.POST['confirm_password']:
        context.update({
            'status': 509,
            'message': "Passwords do not match."
        })
        return JsonResponse(context)
    exist_data = models.User.objects.filter(
        Q(email__iexact=request.POST['email']) | Q(phone__iexact=request.POST['phone'])).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 510,
            'message': "User with this email or phone number already exists."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            user = models.User()
            user.username = request.POST['email']
            user.email = request.POST['email']
            user.phone = request.POST['phone']
            user.name = request.POST['name']
            user.pswd_token = request.POST['password']
            user.password = make_password(request.POST['password'])
            user.role_id = request.POST['role_id']
            user.is_superuser = 0
            user.status = 1
            user.save()
            if 'photo' in request.FILES.keys():
                photo = request.FILES['photo']
                user_name=(user.name).replace(' ', '_')
                user_file_name = user_name + "_sign" + Path(photo.name).suffix
                
                directory_path = settings.MEDIA_ROOT + "/" + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/"
                
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/")
                saved_file = fs.save(user_file_name, photo)
                
                photo_path = settings.MEDIA_URL + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/" + saved_file
                user.user_sign = photo_path
                user.save()
                user_log_details_add(userId,'Add New User ')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "User Created Successfully."
        })
    except Exception:
        context.update({
            'status': 511,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def userEdit(request):
    context = {}
    userId = request.COOKIES.get('userId', None)

    if not request.POST['name'] or not request.POST['role_id'] or not request.POST['email'] or not request.POST['phone'] :
        context.update({
            'status': 512,
            'message': "Name/Role/Email/Phone/Password/Confirmed Password has not been provided."
        })
        return JsonResponse(context)
    if request.POST['password'] != "" or not request.POST['confirm_password'] != "":
        if request.POST['password'] != request.POST['confirm_password']:
            context.update({
                'status': 513,
                'message': "Passwords do not match."
            })
            return JsonResponse(context)
    exist_data = models.User.objects.filter(Q(email__iexact=request.POST['email']) | Q(
        phone__iexact=request.POST['phone'])).exclude(id=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 514,
            'message': "User with this email or phone number already exists."
        })
        return JsonResponse(context)
    
    try:
        with transaction.atomic():
            user = models.User.objects.get(pk=request.POST['id'])
            user.username = request.POST['email']
            user.email = request.POST['email']
            user.phone = request.POST['phone']
            user.name = request.POST['name']
            if request.POST['password'] != "":
                user.pswd_token = request.POST['password']
                user.password = make_password(request.POST['password'])
            user.role_id = request.POST['role_id']
            user.updated_at = datetime.now()
            user.save()
            if 'photo' in request.FILES.keys():
                # Remove old photo if it exists
                if user.user_sign:
                    old_photo_path = user.user_sign.replace(settings.MEDIA_URL, settings.MEDIA_ROOT)
                    if os.path.isfile(old_photo_path):
                        os.remove(old_photo_path)

                photo = request.FILES['photo']
                user_name=(user.name).replace(' ', '_')
                user_file_name = user_name + "_sign" + Path(photo.name).suffix
                
                directory_path = settings.MEDIA_ROOT + "/" + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/"
                
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/")
                saved_file = fs.save(user_file_name, photo)
                
                photo_path = settings.MEDIA_URL + env("USER_MEDIA_PROFILE").replace(
                    "${USER}", str(user.pk) + "~~" + user_name) + "/photo/" + saved_file
                user.user_sign = photo_path
                user.save()
                user_log_details_add(userId,'Edit User ')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "User Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 515,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def userDelete(request):
    context = {}
    userId = request.COOKIES.get('userId', None)

    user = models.User.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            user.delete()
            user_log_details_add(userId,'delete User ')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "User Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 516,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def userLogDetailsList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)

    if id is not None and id != "":
        userLogDetailsHead = list(models.User_Log_Details.objects.filter(pk=id).values(
            'pk', 'user_id', 'user__name', 'task_name', 'time_stamp'))
        context.update({
            'status': 200,
            'message': "User Log Details Fetched Successfully.",
            'page_items': userLogDetailsHead,  # Ensure this is userLogDetailsHead, not 'user'
        })
    else:
        if keyword is not None and keyword != "":
            userLogDetailsHead = list(models.User_Log_Details.objects.filter(
                Q(user__name__icontains=keyword) | Q(task_name__icontains=keyword) | Q(
                    time_stamp__icontains=keyword)
            ).filter(status=1, deleted=0).values('pk', 'user_id', 'user__name', 'task_name', 'time_stamp'))
        else:
            userLogDetailsHead = list(models.User_Log_Details.objects.filter(
                status=1, deleted=0).values('pk', 'user_id', 'user__name', 'task_name', 'time_stamp'))

        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Users Log Fetched Successfully.",
                'page_items': userLogDetailsHead,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(userLogDetailsHead, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Users Log Fetched Successfully.",
            'page_items': list(page_items),  # Convert page_items to a list if it isn't one already
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': button_to_show,
        })
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def vendorList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        vendor = list(models.Vendor.objects.filter(pk=id)[:1].values('pk', 'name', 'address', 'country__pk', 'state__pk', 'city__pk',
                      'country__name', 'state__name', 'city__name', 'pin', 'gst_no', 'contact_no', 'contact_name', 'contact_email'))
        if len(vendor) > 0:
            purchase_order_count = models.Purchase_Order.objects.filter(
                vendor_id=vendor[0]['pk']).count()
            vendor[0]['next_purchase_order_number'] = env("PURCHASE_ORDER_NUMBER_SEQ").replace("${VENDOR_SHORT}", ''.join(word[0] for word in vendor[0]['name'].split())).replace(
                "${AI_DIGIT_3}", str(purchase_order_count + 1).zfill(3)).replace("${FINANCE_YEAR}", datetime.today().strftime('%y') + "-" + (datetime(datetime.today().year + 1, 1, 1).strftime('%y')))
        context.update({
            'status': 200,
            'message': "Vendor Fetched Successfully.",
            'page_items': vendor,
        })
    else:
        if keyword is not None and keyword != "":
            vendors = list(models.Vendor.objects.filter(Q(name__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(contact_email__icontains=keyword) | Q(contact_no__icontains=keyword) | Q(pin__icontains=keyword) | Q(city__name__icontains=keyword)).filter(status=1, deleted=0).values(
                'pk', 'name', 'address', 'country__pk', 'state__pk', 'city__pk', 'country__name', 'state__name',
                'city__name', 'pin', 'gst_no', 'contact_no', 'contact_name', 'contact_email'))
        else:
            vendors = list(
                models.Vendor.objects.filter(status=1, deleted=0).values('pk', 'name', 'address', 'country__pk', 'state__pk', 'city__pk', 'country__name', 'state__name', 'city__name', 'pin', 'gst_no', 'contact_no', 'contact_name', 'contact_email'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Vendors Fetched Successfully.",
                'page_items': vendors,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(vendors, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Vendors Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def vendorAdd(request):
    context = {}
    
    if not request.POST['name'] or not request.POST['contact_name']  or not request.POST['gst_no'] or not request.POST['pin'] or not request.POST['address'] or not request.POST['country_id'] or not request.POST['state_id'] or not request.POST['city_id']:
        context.update({
            'status': 517,
            'message': "Name/Contact Name/GST Number/Pin/Address/Country/State/City has not been provided."
        })
        return JsonResponse(context)
    if request.POST.get('contact_email',None) or request.POST.get('contact_no',None):
        exist_data = models.Vendor.objects.filter(Q(contact_email__iexact=request.POST['contact_email']) | Q(
            contact_no__iexact=request.POST['contact_no'])).filter(deleted=0)
        if len(exist_data) > 0:
            context.update({
                'status': 518,
                'message': "Vendor with this email or phone number already exists."
            })
            return JsonResponse(context)
    try:
        with transaction.atomic():
            
            vendor = models.Vendor()
            vendor.name = request.POST['name']
            vendor.contact_name = request.POST['contact_name']
            vendor.contact_email = request.POST['contact_email']
            vendor.contact_no = request.POST['contact_no']
            vendor.gst_no = request.POST['gst_no']
            vendor.pin = request.POST['pin']
            vendor.address = request.POST['address']
            vendor.country_id = request.POST['country_id']
            vendor.state_id = request.POST['state_id']
            vendor.city_id = request.POST['city_id']
            vendor.store_present = 1 if int(request.POST['createStore']) == 1 else 0
            vendor.save()
            
            if int(request.POST['createStore']) == 1:
                store = models.Store()
                store.name = request.POST['name']
                store.address = request.POST['address']
                store.country_id = request.POST['country_id']
                store.state_id = request.POST['state_id']
                store.city_id = request.POST['city_id']
                store.pin = request.POST['pin']
                store.contact_name = request.POST['contact_name']
                store.contact_no = request.POST['contact_no']
                store.contact_email = request.POST['contact_email']
                if request.POST.get('manager_name', None):
                    store.manager_name = request.POST['manager_name']
                store.vendor_id = vendor.id
                store.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Add New Vendor')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Vendor Created Successfully."
        })
    except Exception:
        context.update({
            'status': 519,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def vendorEdit(request):
    context = {}
    if not request.POST['name'] or not request.POST['contact_name']  or not request.POST['gst_no'] or not request.POST['pin'] or not request.POST['address'] or not request.POST['country_id'] or not request.POST['state_id'] or not request.POST['city_id']:
        context.update({
            'status': 520,
            'message': "Name/Contact Name/GST Number/Pin/Address/Country/State/City has not been provided."
        })
        return JsonResponse(context)
    # if request.POST.get('contact_email',None) or request.POST.get('contact_no',None):
    #     exist_data = models.Vendor.objects.filter(Q(contact_email__iexact=request.POST['contact_email']) | Q(
    #         contact_no__iexact=request.POST['contact_no'])).filter(deleted=0)
    #     if len(exist_data) > 0:
    #         context.update({
    #             'status': 518,
    #             'message': "Vendor with this email or phone number already exists."
    #         })
    #         return JsonResponse(context)
    try:
        with transaction.atomic():
            
            vendor = models.Vendor.objects.get(pk=request.POST['id'])
            
            vendor.name = request.POST['name']
            vendor.contact_name = request.POST['contact_name']
            vendor.contact_email = request.POST['contact_email']
            vendor.contact_no = request.POST['contact_no']
            vendor.gst_no = request.POST['gst_no']
            vendor.pin = request.POST['pin']
            vendor.address = request.POST['address']
            vendor.country_id = request.POST['country_id']
            vendor.state_id = request.POST['state_id']
            vendor.city_id = request.POST['city_id']
            
            vendor.store_present = 1 if int(request.POST['createStore']) == 1 else 0
           
            vendor.updated_at = datetime.now()
            vendor.save()
           
            if int(request.POST['createStore']) == 1:
                store = models.Store()
                store.name = request.POST['name']
                store.address = request.POST['address']
                store.country_id = request.POST['country_id']
                store.state_id = request.POST['state_id']
                store.city_id = request.POST['city_id']
                store.pin = request.POST['pin']
                store.contact_name = request.POST['contact_name']
                store.contact_no = request.POST['contact_no']
                store.contact_email = request.POST['contact_email']
                if request.POST.get('manager_name', None):
                    store.manager_name = request.POST['manager_name']
                store.vendor_id = vendor.id
                store.save()

            elif int(request.POST['createStore']) == 0 and models.Store.objects.filter(vendor_id = request.POST['id'] ).exists() :
               
                store = models.Store.objects.get(vendor_id = request.POST['id'] )
                store.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Edit Vendor')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Vendor Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 522,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def vendorDelete(request):
    context = {}
    vendor = models.Vendor.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            vendor.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Delete Vendor')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Vendor Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 523,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
def vendorExport(request):
    keyword = request.GET.get('keyword')
    if keyword is not None and keyword != "":
        page_items = models.Vendor.objects.filter(Q(name__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(contact_email__icontains=keyword) | Q(
            contact_no__icontains=keyword) | Q(pin__icontains=keyword) | Q(city__name__icontains=keyword)).filter(status=1, deleted=0)
    else:
        page_items = models.Vendor.objects.filter(status=1, deleted=0)

    directory_path = settings.MEDIA_ROOT + '/reports/'
    path = Path(directory_path)
    path.mkdir(parents=True, exist_ok=True)

    for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
        if not f.endswith(".xlsx"):
            continue
        os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

    # tmpname = str(datetime.now().microsecond) + ".xlsx"
    tmpname = "Vendor" + ".xlsx"
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = "Name"
    ws['B1'] = "Address"
    ws['C1'] = "Contact Name"
    ws['D1'] = "Contact Number"
    ws['E1'] = "Contact Email"
    ws['F1'] = "GST Number"
    ws['G1'] = "Country"
    ws['H1'] = "State"
    ws['I1'] = "City"
    ws['J1'] = "Pin"

    # Rows can also be appended
    for each in page_items:
        ws.append([each.name, each.address, each.contact_name, each.contact_no,
                  each.contact_email, each.gst_no, each.country.name, each.state.name, each.city.name, each.pin])

    # Save the file
    wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
    os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
    return JsonResponse({
        'code': 200,
        'filename': settings.MEDIA_URL + 'reports/' + tmpname,
        'name':  tmpname
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def configUserAdd(request):
    context = {}
    try:
        with transaction.atomic():
           userConfigadd = models.Configuration_User()
           userConfigadd.client_name = request.POST['name']
           userConfigadd.client_address = request.POST['client_address']
           userConfigadd.client_work_address  = request.POST['client_work_address']
           userConfigadd.client_gst = request.POST['gst_no']
           userConfigadd.client_contact = request.POST['contact_no']
           userConfigadd.finacial_year_start = request.POST['finacial_year_start']
           userConfigadd.finacial_year_end = request.POST['finacial_year_end']
           userConfigadd.save()
        transaction.commit()
        context.update({
            'status': 200,
            'message': "client config added successfully"
        })
    except Exception:
        context.update({
            'status': 526.1,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def configUserEdit(request):
    context = {}
    try:
        with transaction.atomic():
           userConfigadd = models.Configuration_User.objects.get(pk=request.POST['pk'])
           userConfigadd.client_name = request.POST['name']
           userConfigadd.client_address = request.POST['client_address']
           userConfigadd.client_work_address  = request.POST['client_work_address']
           userConfigadd.client_gst = request.POST['gst_no']
           userConfigadd.client_contact = request.POST['contact_no']
           userConfigadd.finacial_year_start = request.POST['finacial_year_start']
           userConfigadd.finacial_year_end = request.POST['finacial_year_end']
           userConfigadd.pan_no = request.POST['pan_no']
           userConfigadd.cin_no = request.POST['cin_no']
           userConfigadd.ice_code = request.POST['ice_code']
           userConfigadd.udyam_no = request.POST['udyam_no']
           userConfigadd.mail_id = request.POST['email']
           userConfigadd.website = request.POST['website']
        #    # # # # #print('812')
           if 'photo' in request.FILES.keys():
            photo = request.FILES['photo']
            directory_path = settings.MEDIA_ROOT + "/" + env("CLIENT_MEDIA_COMPANY_LOGO") + "/photo/"
            # # # # # #print(directory_path)
            path = Path(directory_path)
            path.mkdir(parents=True, exist_ok=True)
            fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("CLIENT_MEDIA_COMPANY_LOGO") + "/photo/")
            saved_file = fs.save(photo.name, photo)
            photo_path = settings.MEDIA_URL + "/" + env("CLIENT_MEDIA_COMPANY_LOGO") + "/photo/" + saved_file
            userConfigadd.logo = photo_path
           userConfigadd.updated_at = datetime.now()
           userConfigadd.save()
        transaction.commit()
        context.update({
            'status': 200,
            'message': "client config Update successfully"
        })
    except Exception:
        context.update({
            'status': 526.2,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customerList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        customer = list(models.Customer.objects.filter(pk=id)[:1].values('pk', 'name', 'address', 'landmark', 'country__pk', 'state__pk',
                        'country__name', 'state__name', 'city', 'pin', 'contact_no', 'contact_name', 'contact_email', 'customer_type__name', 'photo', 'kyc_image'))
        context.update({
            'status': 200,
            'message': "Customer Fetched Successfully.",
            'page_items': customer,
        })
    else:
        if keyword is not None and keyword != "":
            customers = list(models.Customer.objects.filter(Q(name__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(contact_email__icontains=keyword) | Q(contact_no__icontains=keyword) | Q(customer_type__name__icontains=keyword) | Q(pin__icontains=keyword) | Q(city__icontains=keyword)).filter(
                status=1, deleted=0).values('pk', 'name', 'address', 'landmark', 'country__pk', 'state__pk', 'country__name', 'state__name', 'city', 'pin', 'contact_no', 'contact_name', 'contact_email', 'customer_type__name', 'photo', 'kyc_image'))
        else:
            customers = list(models.Customer.objects.filter(status=1, deleted=0).values('pk', 'name', 'address', 'landmark', 'country__pk', 'state__pk',
                             'country__name', 'state__name', 'city', 'pin', 'contact_no', 'contact_name', 'contact_email', 'customer_type__name', 'photo', 'kyc_image'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Customers Fetched Successfully.",
                'page_items': customers,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(customers, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Customers Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def customerAdd(request):
    context = {}
    required_fields = ['name', 'contact_name', 'contact_no', 'landmark', 'pin', 'customer_type_id',
                       'kyc_type_id', 'kyc_detail', 'address', 'country_id', 'state_id', 'city']
    for field in required_fields:
        if not request.POST.get(field):
            context.update({
                'status': 524,
                'message': "Name/Contact Name/Contact No/Landmark/Pin/Customer Type/KYC Type/KYC Detail/"
                           "Address/Country/State/City has not been provided."
            })
            return JsonResponse(context)

    contact_email = request.POST.get('contact_email')
    # Ensure contact_email can be empty
    if contact_email == "":
        contact_email = None

    # Set the flag for customer creation
    customer_can_be_created = True

    # Check if a customer with the same contact number or email exists for a different contact name
    if models.Customer.objects.filter(
            Q(Q(contact_no__iexact=request.POST.get('contact_no')) | Q(contact_email__iexact=contact_email)),
            Q(~Q(contact_name__iexact=request.POST.get('contact_name')))
    ).exists():
        customer_can_be_created = False

    if not customer_can_be_created:
        context.update({
            'status': 525,
            'message': "A customer with this email or phone number already exists for a different contact name."
        })
        return JsonResponse(context)

    try:
        with transaction.atomic():
            customer = models.Customer()
            customer.name = request.POST['name']
            customer.contact_name = request.POST['contact_name']
            customer.contact_email = request.POST['contact_email'] if request.POST['contact_email'] != "" else None
            customer.contact_no = request.POST['contact_no']
            customer.contact_no_std = request.POST['contact_no_std']
            customer.landmark = request.POST['landmark']
            customer.pin = request.POST['pin']
            customer.customer_type_id = request.POST['customer_type_id']
            customer.kyc_type_id = request.POST['kyc_type_id']
            customer.kyc_detail = request.POST['kyc_detail']
            customer.date_of_birth = request.POST['date_of_birth'] if request.POST['date_of_birth'] != "" else None
            customer.date_of_anniversary = request.POST['date_of_anniversary']
            customer.weekly_closing_day = ", ".join(request.POST.getlist(
                'weekly_closing_day')) if 'weekly_closing_day' in request.POST.keys() else None
            customer.morning_from_time = request.POST[
                'morning_from_time'] if request.POST['morning_from_time'] != "" else None
            customer.morning_to_time = request.POST['morning_to_time'] if request.POST['morning_to_time'] != "" else None
            customer.evening_from_time = request.POST[
                'evening_from_time'] if request.POST['evening_from_time'] != "" else None
            customer.evening_to_time = request.POST['evening_to_time'] if request.POST['evening_to_time'] != "" else None
            customer.address = request.POST['address']
            customer.country_id = request.POST['country_id']
            customer.state_id = request.POST['state_id']
            customer.city = request.POST['city']
            customer.save()
            if 'photo' in request.FILES.keys():
                photo = request.FILES['photo']
                customer_name=(customer.name).replace(' ', '_')
                custom_file_name = customer_name + "_photo" + Path(photo.name).suffix
                
                directory_path = settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/"
                
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/")
                saved_file = fs.save(custom_file_name, photo)
                
                photo_path = settings.MEDIA_URL + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/" + saved_file
                customer.photo = photo_path
                customer.save()
            if 'kyc_image' in request.FILES.keys():
                kyc_image = request.FILES['kyc_image']
                customer_name=(customer.name).replace(' ', '_')
                custom_file_name = customer_name + "_kyc_image" + Path(kyc_image.name).suffix
                
                directory_path = settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/"
                
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/")
                saved_file = fs.save(custom_file_name, kyc_image)
                
                kyc_image_path = settings.MEDIA_URL + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/" + saved_file
                customer.kyc_image = kyc_image_path
                customer.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Add New Customer')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Customer Created Successfully."
        })
    except Exception:
        context.update({
            'status': 526,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def customerEdit(request):
    context = {}
    required_fields = ['name', 'contact_name', 'contact_no', 'landmark', 'pin', 'customer_type_id',
                       'kyc_type_id', 'kyc_detail', 'address', 'country_id', 'state_id', 'city']
    for field in required_fields:
        if not request.POST.get(field):
            context.update({
                'status': 524,
                'message': "Name/Contact Name/Contact No/Landmark/Pin/Customer Type/KYC Type/KYC Detail/"
                           "Address/Country/State/City has not been provided."
            })
            return JsonResponse(context)

    contact_email = request.POST.get('contact_email')
    # Ensure contact_email can be empty
    if contact_email == "":
        contact_email = None

    # Set the flag for customer creation
    customer_can_be_edited = True

    # Check if a customer with the same contact number or email exists for a different contact name
    if models.Customer.objects.filter(
            Q(Q(contact_no__iexact=request.POST.get('contact_no')) | Q(contact_email__iexact=contact_email)),
            Q(~Q(contact_name__iexact=request.POST.get('contact_name')))
    ).exists():
        customer_can_be_edited = False

    if not customer_can_be_edited:
        context.update({
            'status': 525,
            'message': "A customer with this email or phone number already exists for a different contact name."
        })
        return JsonResponse(context)

    try:
        with transaction.atomic():
            customer = models.Customer.objects.get(pk=request.POST['id'])
            customer.name = request.POST['name']
            customer.contact_name = request.POST['contact_name']
            customer.contact_email = request.POST['contact_email']
            customer.contact_no = request.POST['contact_no']
            customer.contact_no_std = request.POST['contact_no_std']
            customer.landmark = request.POST['landmark']
            customer.pin = request.POST['pin']
            customer.customer_type_id = request.POST['customer_type_id']
            customer.kyc_type_id = request.POST['kyc_type_id']
            customer.kyc_detail = request.POST['kyc_detail']
            customer.date_of_birth = request.POST['date_of_birth'] if request.POST['date_of_birth'] != "" else None
            customer.date_of_anniversary = request.POST['date_of_anniversary']
            customer.weekly_closing_day = ", ".join(request.POST.getlist(
                'weekly_closing_day')) if 'weekly_closing_day' in request.POST.keys() else None
            customer.morning_from_time = request.POST[
                'morning_from_time'] if request.POST['date_of_birth'] != "" else None
            customer.morning_to_time = request.POST['morning_to_time'] if request.POST['date_of_birth'] != "" else None
            customer.evening_from_time = request.POST[
                'evening_from_time'] if request.POST['date_of_birth'] != "" else None
            customer.evening_to_time = request.POST['evening_to_time'] if request.POST['date_of_birth'] != "" else None
            customer.address = request.POST['address']
            customer.country_id = request.POST['country_id']
            customer.state_id = request.POST['state_id']
            customer.city = request.POST['city']
            customer.updated_at = datetime.now()
            customer.save()

            if 'photo' in request.FILES.keys():
                # Remove old photo if it exists
                if customer.photo:
                    old_photo_path = customer.photo.replace(settings.MEDIA_URL, settings.MEDIA_ROOT)
                    if os.path.isfile(old_photo_path):
                        os.remove(old_photo_path)
    
                photo = request.FILES['photo']
                customer_name=(customer.name).replace(' ', '_')
                custom_file_name = customer_name + "_kyc_image" + Path(photo.name).suffix
                directory_path = settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/"
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/")
                saved_file = fs.save(custom_file_name, photo)
                photo_path = settings.MEDIA_URL + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/photo/" + saved_file
                customer.photo = photo_path
                customer.save()
            if 'kyc_image' in request.FILES.keys():
                # Remove old kyc_image if it exists
                if customer.kyc_image:
                    old_kyc_image_path = customer.kyc_image.replace(settings.MEDIA_URL, settings.MEDIA_ROOT)
                    if os.path.isfile(old_kyc_image_path):
                        os.remove(old_kyc_image_path)
                kyc_image = request.FILES['kyc_image']
                customer_name=(customer.name).replace(' ', '_')
                custom_file_name = customer_name + "_kyc_image" + Path(kyc_image.name).suffix
                directory_path = settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/"
                path = Path(directory_path)
                path.mkdir(parents=True, exist_ok=True)
                fs = FileSystemStorage(location=settings.MEDIA_ROOT + "/" + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/")
                saved_file = fs.save(custom_file_name, kyc_image)
                kyc_image_path = settings.MEDIA_URL + env("CUSTOMER_MEDIA_PROFILE").replace(
                    "${CUSTOMER}", str(customer.pk) + "~~" + customer_name) + "/kyc/" + saved_file
                customer.kyc_image = kyc_image_path
                customer.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Edit Customer')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Customer Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 529,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def customerDelete(request):
    context = {}
    customer = models.Customer.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            customer.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Customer Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Customer Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 530,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
def customerExport(request):
    keyword = request.GET.get('keyword')
    if keyword is not None and keyword != "":
        page_items = models.Customer.objects.filter(Q(name__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(email__icontains=keyword) | Q(
            phone__icontains=keyword) | Q(customer_type__name__icontains=keyword) | Q(pin__icontains=keyword) | Q(city__icontains=keyword)).filter(status=1, deleted=0)
    else:
        page_items = models.Customer.objects.filter(status=1, deleted=0)

    directory_path = settings.MEDIA_ROOT + '/reports/'
    path = Path(directory_path)
    path.mkdir(parents=True, exist_ok=True)

    for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
        if not f.endswith(".xlsx"):
            continue
        os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

    # tmpname = str(datetime.now().microsecond) + ".xlsx"
    tmpname = "Customer" + ".xlsx"
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = "Name"
    ws['B1'] = "Address"
    ws['C1'] = "Contact Name"
    ws['D1'] = "Contact Number"
    ws['E1'] = "Contact Email"
    ws['F1'] = "Customer Type"
    ws['G1'] = "Country"
    ws['H1'] = "State"
    ws['I1'] = "City"
    ws['J1'] = "Pin"

    # Rows can also be appended
    for each in page_items:
        ws.append([each.name, each.address, each.contact_name, each.contact_no, each.contact_email,
                  each.customer_type.name, each.country.name, each.state.name, each.city.name, each.pin])

    # Save the file
    wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
    os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
    return JsonResponse({
        'code': 200,
        'filename': settings.MEDIA_URL + 'reports/' + tmpname,
        'name':  tmpname
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def uomList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        uom = list(models.Uom.objects.filter(pk=id)[:1].values('pk', 'name'))
        context.update({
            'status': 200,
            'message': "UOM Fetched Successfully.",
            'page_items': uom,
        })
    else:
        if keyword is not None and keyword != "":
            uoms = list(models.Uom.objects.filter(
                name__icontains=keyword, status=1, deleted=0).values('pk', 'name'))
        else:
            uoms = list(models.Uom.objects.filter(
                status=1, deleted=0).values('pk', 'name'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Uoms Fetched Successfully.",
                'page_items': uoms,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(uoms, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "UOMs Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uomAdd(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 531,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Uom.objects.filter(
        name=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 532,
            'message': "Uom with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            uom = models.Uom()
            uom.name = request.POST['name']
            uom.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'UOM Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "UOM Created Successfully."
        })
    except Exception:
        context.update({
            'status': 533,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uomEdit(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 534,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Uom.objects.filter(
        name=request.POST['name']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 535,
            'message': "Uom with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            uom = models.Uom.objects.get(pk=request.POST['id'])
            uom.name = request.POST['name']
            uom.updated_at = datetime.now()
            uom.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Uom Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "UOM Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 536,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uomDelete(request):
    context = {}
    uom = models.Uom.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            uom.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Uom Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "UOM Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 537,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def childUomList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        childUom = list(models.Child_Uom.objects.filter(pk=id)[:1].values(
            'pk', 'name', 'uom__name', 'conversion_rate'))
        context.update({
            'status': 200,
            'message': "UOM Fetched Successfully.",
            'page_items': childUom,
        })
    else:
        if keyword is not None and keyword != "":
            childUoms = list(models.Child_Uom.objects.filter(name__icontains=keyword, status=1, deleted=0).values(
                'pk', 'name', 'uom__name', 'conversion_rate'))
        else:
            childUoms = list(models.Child_Uom.objects.filter(status=1, deleted=0).values(
                'pk', 'name', 'uom__name', 'conversion_rate'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Child Uoms Fetched Successfully.",
                'page_items': childUoms,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(childUoms, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "UOMs Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def childUomAdd(request):
    context = {}
    if not request.POST['name'] or not request.POST['uom_id'] or not request.POST['conversion_rate']:
        context.update({
            'status': 538,
            'message': "Name/Uom/Conversion Rate has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Child_Uom.objects.filter(
        name=request.POST['name'], uom_id=request.POST['uom_id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 539,
            'message': "Child Uom with this name and Uom already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            childUom = models.Child_Uom()
            childUom.name = request.POST['name']
            childUom.uom_id = request.POST['uom_id']
            childUom.conversion_rate = request.POST['conversion_rate']
            childUom.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Child UOM Add')
        transaction.commit()

        context.update({
            'status': 200,
            'message': "Child UOM Created Successfully."
        })
    except Exception:
        context.update({
            'status': 540,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def childUomEdit(request):
    context = {}
    if not request.POST['name'] or not request.POST['uom_id'] or not request.POST['conversion_rate']:
        context.update({
            'status': 541,
            'message': "Name/Uom/Conversion Rate has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Child_Uom.objects.filter(
        name=request.POST['name'], uom_id=request.POST['uom_id']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 542,
            'message': "Child Uom with this name and Uom already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            childUom = models.Child_Uom.objects.get(pk=request.POST['id'])
            childUom.name = request.POST['name']
            childUom.uom_id = request.POST['uom_id']
            childUom.conversion_rate = request.POST['conversion_rate']
            childUom.updated_at = datetime.now()
            childUom.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Child UOM Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Child UOM Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 543,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def childUomDelete(request):
    context = {}
    childUom = models.Child_Uom.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            childUom.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Child UOM Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Child UOM Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 544,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def itemCategoryList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        itemCategory = list(models.Item_Category.objects.get(
            pk=id).values('pk', 'name'))
        context.update({
            'status': 200,
            'message': "Item Category Fetched Successfully.",
            'page_items': itemCategory,
        })
    else:
        if keyword is not None and keyword != "":
            itemCategories = list(models.Item_Category.objects.filter(
                name__icontains=keyword, status=1, deleted=0).values('pk', 'name'))
        else:
            itemCategories = list(models.Item_Category.objects.filter(
                status=1, deleted=0).values('pk', 'name'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Item Categories Fetched Successfully.",
                'page_items': itemCategories,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(itemCategories, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Item Categories Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemCategoryAdd(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 545,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Category.objects.filter(
        name=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 546,
            'message': "Item Category with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemCategory = models.Item_Category()
            itemCategory.name = request.POST['name']
            itemCategory.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Catagory Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Category Created Successfully."
        })
    except Exception:
        context.update({
            'status': 547,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemCategoryEdit(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 548,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Category.objects.filter(
        name=request.POST['name']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 549,
            'message': "Item Category with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemCategory = models.Item_Category.objects.get(
                pk=request.POST['id'])
            itemCategory.name = request.POST['name']
            itemCategory.updated_at = datetime.now()
            itemCategory.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Category Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Category Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 550,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemCategoryDelete(request):
    context = {}
    itemCategory = models.Item_Category.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            itemCategory.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Category Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Category Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 551,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def itemTypeList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    itemCatId = request.GET.get('itemCatId', None)
    if id is not None and id != "":
        itemType = list(models.Item_Type.objects.filter(pk=id)[:1].values(
            'pk', 'name', 'item_category__name', 'gst_percentage'))
        context.update({
            'status': 200,
            'message': "Item Type Fetched Successfully.",
            'page_items': itemType,
        })
    elif itemCatId is not None and itemCatId != "":
        itemType = list(models.Item_Type.objects.filter(item_category_id=itemCatId).values(
            'pk', 'name', 'item_category__name', 'gst_percentage'))
        # # # # # #print(itemType)
        context.update({
            'status': 200,
            'message': "Item Type Fetched Successfully.",
            'page_items': itemType,
        })
    else:
        if keyword is not None and keyword != "":
            itemTypes = list(models.Item_Type.objects.filter(
                Q(name__icontains=keyword) | Q(item_category__name__icontains=keyword) 
            ).filter(status=1, deleted=0).values(
                'pk', 'name', 'item_category__name', 'gst_percentage'))
        else:
            itemTypes = list(models.Item_Type.objects.filter(status=1, deleted=0).values(
                'pk', 'name', 'item_category__name', 'gst_percentage'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Item Types Fetched Successfully.",
                'page_items': itemTypes,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(itemTypes, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Item Types Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    # # # # # #print(context,'\n')
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemTypeAdd(request):
    context = {}
    if not request.POST['name'] or not request.POST['item_category_id']  or not request.POST['gst_percentage']:
        context.update({
            'status': 552,
            'message': "Name/Item Category/HSN Code/GST Percentage has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Type.objects.filter(
        name=request.POST['name'], item_category_id=request.POST['item_category_id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 553,
            'message': "Item Type with this name and item category already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemType = models.Item_Type()
            itemType.name = request.POST['name']
            itemType.item_category_id = request.POST['item_category_id']
            itemType.gst_percentage = request.POST['gst_percentage']
            itemType.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Type Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Type Created Successfully."
        })
    except Exception:
        context.update({
            'status': 554,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemTypeEdit(request):
    context = {}
    if not request.POST['name'] or not request.POST['item_category_id']  or not request.POST['gst_percentage']:
        context.update({
            'status': 555,
            'message': "Name/Item Category/HSN Code/GST Percentage has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Type.objects.filter(
        name=request.POST['name'], item_category_id=request.POST['item_category_id']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 556,
            'message': "Item Type with this name and item category already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemType = models.Item_Type.objects.get(pk=request.POST['id'])
            itemType.name = request.POST['name']
            itemType.item_category_id = request.POST['item_category_id']
            itemType.gst_percentage = request.POST['gst_percentage']
            itemType.updated_at = datetime.now()
            itemType.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Type Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Type Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 557,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemTypeDelete(request):
    context = {}
    itemType = models.Item_Type.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            itemType.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Type Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Type Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 558,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def itemColorList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        itemColor = list(models.Item_Color.objects.get(
            pk=id).values('pk', 'name', 'color_code'))
        context.update({
            'status': 200,
            'message': "Item Color Fetched Successfully.",
            'page_items': itemColor,
        })
    else:
        if keyword is not None and keyword != "":
            itemColors = list(models.Item_Color.objects.filter(
                name__icontains=keyword, status=1, deleted=0).values('pk', 'name', 'color_code'))
        else:
            itemColors = list(models.Item_Color.objects.filter(
                status=1, deleted=0).values('pk', 'name', 'color_code'))
        if find_all is not None and int(find_all) == 1:
            # ok 
            context.update({
                'status': 200,
                'message': "Item Colors Fetched Successfully.",
                'page_items': itemColors,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(itemColors, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Item Colors Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemColorAdd(request):
    context = {}
    if not request.POST['name'] or not request.POST['color_code']:
        context.update({
            'status': 559,
            'message': "Name/Color Code has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Color.objects.filter(
        name=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 560,
            'message': "Item Color with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemColor = models.Item_Color()
            itemColor.name = request.POST['name']
            itemColor.color_code = request.POST['color_code']
            itemColor.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Color Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Color Created Successfully."
        })
    except Exception:
        context.update({
            'status': 561,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemColorEdit(request):
    context = {}
    if not request.POST['name'] or not request.POST['color_code']:
        context.update({
            'status': 562,
            'message': "Name/Color Code has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item_Color.objects.filter(
        name=request.POST['name']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 563,
            'message': "Item Color with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            itemColor = models.Item_Color.objects.get(pk=request.POST['id'])
            itemColor.name = request.POST['name']
            itemColor.color_code = request.POST['color_code']
            itemColor.updated_at = datetime.now()
            itemColor.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Color Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Color Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 564,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemColorDelete(request):
    context = {}
    itemColor = models.Item_Color.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            itemColor.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Color Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Color Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 565,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def gstList(request):
    context = {}
    id = request.GET.get('id', None)
    keyword = request.GET.get('keyword', None)
    find_all = request.GET.get('find_all', None)
    if id is not None and id != "":
        gst = list(models.Gst.objects.filter(pk=id)[:1].values(
            'pk', 'gst_value', 'igst_value', 'cgst_value', 'sgst_value','name'))
        context.update({
            'status': 200,
            'message': "Item Fetched Successfully.",
            'page_items': gst,
        })
        return JsonResponse(context)
    else:
        if keyword is not None and keyword != "":
            gsts=models.Gst.objects.filter(status=1, deleted=0).filter(Q(name__icontains=keyword))
        else:
            gsts=models.Gst.objects.filter(status=1, deleted=0)

        gsts = list(gsts.values(
                'pk', 'gst_value', 'igst_value', 'cgst_value', 'sgst_value','name'))

        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Gst Fetched Successfully.",
                'page_items': gsts,
            })
            return JsonResponse(context)
        
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(gsts, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Gst Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def gstAdd(request):
    context ={}
    # #print(request.POST)
    try:
        with transaction.atomic():
            gstHeader = models.Gst()
            gstHeader.name = request.POST['gst']
            gstHeader.gst_value = Decimal(request.POST['gst'])
            gstHeader.igst_value = Decimal(request.POST['gst'])
            gstHeader.cgst_value = Decimal(request.POST['gst'])/ Decimal('2')
            gstHeader.sgst_value =  Decimal(request.POST['gst'])/Decimal('2')
            gstHeader.save()
        transaction.commit()    
        userId = request.COOKIES.get('userId', None)
        user_log_details_add(userId,'new gst updated')
        context.update({
            'status': 200,
            'message': 'gst updated'
        })
    except Exception:
        context.update({
            'status': 500,
            'message': 'something went wrong'
        })
        transaction.rollback()

    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def itemList(request):
    context = {}
    # # # # # #print(request.GET)
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    item_type_id = request.GET.get('item_type_id', None)
    item_category_id = request.GET.get('item_category_id', None)
    item_category = request.GET.get('item_category', None)
    if id is not None and id != "":
        item = list(models.Item.objects.filter(pk=id)[:1].values(
            'pk', 'name', 'item_type__name', 'item_type__item_category__name', 'item_type__gst_percentage', 'uom__name', 'price','hsn_code'))
        context.update({
            'status': 200,
            'message': "Item Fetched Successfully.",
            'page_items': item,
        })
        return JsonResponse(context)
    else:
        items=models.Item.objects.filter(status=1, deleted=0)

        if item_category_id is not None and item_category_id!="":
            items=items.filter(item_type__item_category_id=item_category_id)
        if item_type_id is not None and item_type_id!="":
            items=items.filter(item_type_id=item_type_id)
        if keyword is not None and keyword != "":
            items = items.filter(name__icontains=keyword)
        if item_category is not None and item_category!="":
            if item_category=="not-finish":
                items = items.filter(Q(item_type__item_category__name__icontains="raw") | Q(item_type__item_category__name__icontains="semi"))
            if item_category=="semi":
                items = items.filter(item_type__item_category__name__icontains="semi")
            if item_category=="finish":
                items = items.filter(item_type__item_category__name__icontains="finish")

        items = list(items.values(
                'pk', 'name', 'item_type__name', 'item_type__item_category__name', 'item_type__gst_percentage',
                'uom__name', 'price', 'hsn_code'))

        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Items Fetched Successfully.",
                'page_items': items,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(items, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Items Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemAdd(request):
    context = {}
    if not request.POST['name'] or not request.POST['uom_id'] or not request.POST['item_type_id']:
        context.update({
            'status': 566,
            'message': "Name/UOM/Item Type/Price has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item.objects.filter(
        name__iexact=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 567,
            'message': "Item with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            item = models.Item()
            item.name = request.POST['name']
            item.uom_id = request.POST['uom_id']
            item.item_type_id = request.POST['item_type_id']
            if request.POST['price']:
                item.price = request.POST['price']
            item.hsn_code = request.POST['hsn_code']
            item.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'New Item Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Created Successfully."
        })
    except Exception:
        context.update({
            'status': 568,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemEdit(request):
    context = {}
    if not request.POST['name'] or not request.POST['uom_id'] or not request.POST['item_type_id'] or not request.POST['price']:
        context.update({
            'status': 569,
            'message': "Name/UOM/Item Type/Price has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Item.objects.filter(name__iexact=request.POST['name']).exclude(
        pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 570,
            'message': "Item with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            item = models.Item.objects.get(pk=request.POST['id'])
            item.name = request.POST['name']
            item.item_type_id = request.POST['item_type_id']
            item.uom_id = request.POST['uom_id']
            item.price = request.POST['price']
            item.hsn_code = request.POST['hsn_code']
            item.updated_at = datetime.now()
            item.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 571,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def itemDelete(request):
    context = {}
    item = models.Item.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            item.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Item Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Item Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 572,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
def itemExport(request):
    keyword = request.GET.get('keyword')
    file_type=request.GET.get('file_type')
    if keyword is not None and keyword != "":
        page_items = models.Item.objects.filter(Q(name__icontains=keyword) | Q(
            item_type__name__icontains=keyword) | Q(uom__name__icontains=keyword)).filter(status=1, deleted=0)
    else:
        page_items = models.Item.objects.filter(status=1, deleted=0)
    if file_type=="xlsx":
        directory_path = settings.MEDIA_ROOT + '/reports/'
        path = Path(directory_path)
        path.mkdir(parents=True, exist_ok=True)

        for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
            if not f.endswith(".xlsx"):
                continue
            os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

        # tmpname = str(datetime.now().microsecond) + ".xlsx"
        tmpname = "Item" + ".xlsx"
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = "Name"
        ws['B1'] = "Item Type"
        ws['C1'] = "Item Category"
        ws['D1'] = "UOM"
        ws['E1'] = "Price"
        ws['F1'] = "HSN Code"

        # Rows can also be appended
        for each in page_items:
            ws.append([each.name, each.item_type.name,
                      each.item_type.item_category.name, each.uom.name, each.price, each.hsn_code])

        # Save the file
        wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
        os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
        return JsonResponse({
            'code': 200,
            'filename': settings.MEDIA_URL + 'reports/' + tmpname,
            'name':  tmpname
        })

    elif file_type == "csv":
        directory_path = settings.MEDIA_ROOT + '/reports/'
        path = Path(directory_path)
        path.mkdir(parents=True, exist_ok=True)

        # Clean up any existing CSV files in the directory
        for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
            if not f.endswith(".csv"):
                continue
            os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

        tmpname = "Item" + ".csv"

        with open(os.path.join(directory_path, tmpname), 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Name", "Item Type", "Item Category", "UOM", "Price", "HSN Code"])
            for item in page_items:
                writer.writerow(
                    [item.name, item.item_type.name, item.item_type.item_category.name, item.uom.name, item.price, item.hsn_code])

        os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
        return JsonResponse({
            'code': 200,
            'filename': settings.MEDIA_URL + 'reports/' + tmpname,
            'name': tmpname
        })

    elif file_type == "pdf":
        # Create a new PDF document with smaller margins
        pdf = FPDF(orientation='P', unit='mm', format='A4')  # Adjust unit and format if needed
        pdf.set_left_margin(5)
        pdf.set_top_margin(5)

        pdf.add_page()
        pdf.set_font("Arial", size=9)

        # Add a header row with bold text
        pdf.set_font("Arial", size=9, style='B')  # Set font style to bold
        pdf.cell(9, 10, txt="S.No.", border=1, align='C')  # Add S.No. column
        pdf.cell(55, 10, txt="Name", border=1, align='C')
        pdf.cell(38, 10, txt="Item Type", border=1, align='C')
        pdf.cell(38, 10, txt="Item Category", border=1, align='C')
        pdf.cell(18, 10, txt="UOM", border=1, align='C')
        pdf.cell(20, 10, txt="Price", border=1, align='C')
        pdf.cell(20, 10, txt="HSN Code", border=1, align='C')
        pdf.set_font("Arial", size=9)  # Reset font style to normal
        pdf.ln(10)  # Move to the next line

        # Cell height (adjust as needed)
        cell_height = 10

        # Add data rows
        counter = 1  # Counter for serial numbers
        for item in page_items:
            pdf.cell(9, 10, txt=str(counter), border=1, align='C')  # Add S.No. for each row
            pdf.cell(55, 10, txt=item.name, border=1)
            pdf.cell(38, 10, txt=item.item_type.name, border=1)
            pdf.cell(38, 10, txt=item.item_type.item_category.name, border=1)
            pdf.cell(18, 10, txt=item.uom.name, border=1)
            # Right align price for each data row
            pdf.cell(20, 10, txt=str(item.price), align='R', border=1)
            pdf.cell(20, 10, txt=item.hsn_code if item.hsn_code else "", border=1)
            pdf.ln(10)
            counter += 1  # Increment counter for next row

        # Save the PDF file
        directory_path = settings.MEDIA_ROOT + '/reports/'
        path = Path(directory_path)
        path.mkdir(parents=True, exist_ok=True)
        tmpname = "Item.pdf"
        pdf.output(os.path.join(directory_path, tmpname))
        os.chmod(os.path.join(directory_path, tmpname), 0o777)

        return JsonResponse({
            'code': 200,
            'filename': settings.MEDIA_URL + 'reports/' + tmpname,
            'name': tmpname
        })


@api_view(['POST'])
def itemImport(request):
    context = {}

    if request.FILES.get('file'):
        
        excel = request.FILES['file']
        # trying to process files without error
        try:
           
            df = pd.read_excel(excel)
            df.columns = [col.strip().lower() for col in df.columns]
            for index, row in df.iterrows():
                # trying to fetch required cells from 
                try:
                    name = row['name']
                    item_type_name = row['item type']
                    item_category_name = row['item category']
                    uom_name = row['uom'] 
                    price=  handle_empty_cell(row['price'])
                    hsn_code = row['hsn code']
                   
                    # Skip the row if any required field is empty
                    if not all([name, item_type_name, item_category_name, uom_name, hsn_code]):
                        continue  # Skip this row and move to the next one
                    if not models.Item.objects.filter(name__iexact=name).exists():
                       
                        if (
                            models.Item_Type.objects.filter(name__iexact=item_type_name).exists()
                            and models.Item_Category.objects.filter(name__iexact=item_category_name).exists()
                            and models.Uom.objects.filter(name__iexact=uom_name).exists()
                        ):
                            try:
                                with transaction.atomic():
                                    obj = models.Item(
                                        name=name, 
                                        item_type=models.Item_Type.objects.get(name__iexact=item_type_name),
                                        uom=models.Uom.objects.get(name__iexact=uom_name),
                                        price=price,
                                        hsn_code=hsn_code
                                    )
                                    # # # # # #print(obj.__dict__) 
                                    obj.save()
                                    
                                    userId = request.COOKIES.get('userId', None)
                                    user_log_details_add(userId,'Item Bulk Import')
                                transaction.commit()
                                context.update({
                                    'status': 200,
                                    'message': "Items Created Successfully."
                                })
                            except Exception:

                                context.update({
                                    'status': 568,
                                    'message': "Items cannot be created something wrong"
                                })
                                transaction.rollback()
                            
                except KeyError as e:
                    # Handle missing columns
                    context.update({
                        'status': 568,
                        'message': "Required column missing"
                    })
                    return JsonResponse(context)
            if not context:
                context.update({
                'status': 568,
                'message': "No new items in the excel"
            })
        except Exception as e:
            context.update({
                'status': 568,
                'message': "Error processing file"
            })
    else:
        context.update({
            'status': 568,
            'message': "File has not been uploaded"
        })
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def storeList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    store_type=request.GET.get('store_type', None)
    keyword = request.GET.get('keyword', None)
    
    if id is not None and id != "":
        store = list(models.Store.objects.filter(logical_grn_store=0).filter(pk=id)[:1].values(
            'pk', 'name', 'address', 'contact_name', 'contact_no', 'contact_email', 'manager_name', 'pin', 'city__name', 'state__name', 'country__name','vendor_id'))
        context.update({
            'status': 200,
            'message': "Store Fetched Successfully.",
            'page_items': store,
        })
   
    else:
        if keyword is not None and keyword != "":
            stores = list(models.Store.objects.filter(logical_grn_store=0).filter(Q(name__icontains=keyword) | Q(address__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(contact_no__icontains=keyword) | Q(contact_email__icontains=keyword) | Q(manager_name__icontains=keyword) | Q(pin__icontains=keyword)).filter(
                status=1, deleted=0).values('pk', 'name', 'address', 'contact_name', 'contact_no', 'contact_email', 'manager_name', 'pin', 'city__name', 'state__name', 'country__name','vendor_id'))
        else:
            stores = list(models.Store.objects.filter(logical_grn_store=0).filter(status=1, deleted=0).values('pk', 'name', 'address', 'contact_name',
                          'contact_no', 'contact_email', 'manager_name', 'pin', 'city__name', 'state__name', 'country__name','vendor_id'))


        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Stores Fetched Successfully.",
                'page_items': stores,
            })
            return JsonResponse(context)

        if store_type is not None :
            if store_type=="InHouse":
                stores = list(models.Store.objects.filter(logical_grn_store=0).filter(status=1, deleted=0,vendor_id=None).values('pk', 'name', 'address', 'contact_name',
                              'contact_no', 'contact_email', 'manager_name', 'pin', 'city__name', 'state__name', 'country__name','vendor_id'))
                context.update({
                    'status': 200,
                    'message': "Stores Fetched Successfully.",
                    'page_items': stores,
                })
                return JsonResponse(context)

            if store_type=="Vendor":
                stores = list(models.Store.objects.filter(logical_grn_store=0).filter(status=1, deleted=0,vendor_id__isnull=False).values('pk', 'name', 'address', 'contact_name',
                              'contact_no', 'contact_email', 'manager_name', 'pin', 'vendor_id', 'city__name', 'state__name', 'country__name'))
                context.update({
                    'status': 200,
                    'message': "Stores Fetched Successfully.",
                    'page_items': stores,
                })
                return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(stores, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Stores Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeAdd(request):
    context = {}
    # if not request.POST['name'] or not request.POST['address'] or not request.POST['contact_name'] or not request.POST['contact_no'] or not request.POST['contact_email'] or not request.POST['pin'] or not request.POST['city_id'] or not request.POST['state_id'] or not request.POST['country_id']:
    #     context.update({
    #         'status': 566,
    #         'message': "Name/Address/Contact Name/Contact Number/Contact Email/Pin/City/State/Country has not been provided."
    #     })
    #     return JsonResponse(context)
    exist_data = models.Store.objects.filter(
        name__iexact=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 567,
            'message': "Store with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            store = models.Store()
            store.name = request.POST['name']
            store.address = request.POST['address']
            store.country_id = request.POST['country_id']
            store.state_id = request.POST['state_id']
            store.city_id = request.POST['city_id']
            store.pin = request.POST['pin']
            store.contact_name = request.POST['contact_name']
            store.contact_no = request.POST['contact_no']
            store.contact_email = request.POST['contact_email']
            store.manager_name = request.POST['manager_name']
            store.vendor_id = request.POST.get('vendor_id', None)
            store.store_type = request.POST.get('store_type2', 'r')
            store.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'New Store Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Created Successfully."
        })
    except Exception:
        context.update({
            'status': 568,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeEdit(request):
    context = {}
    # if not request.POST['name'] or not request.POST['address'] or not request.POST['contact_name'] or not request.POST['contact_no'] or not request.POST['contact_email'] or not request.POST['pin'] or not request.POST['city_id'] or not request.POST['state_id'] or not request.POST['country_id']:
    #     context.update({
    #         'status': 566,
    #         'message': "Name/Address/Contact Name/Contact Number/Contact Email/Pin/City/State/Country has not been provided."
    #     })
    #     return JsonResponse(context)
    exist_data = models.Store.objects.filter(name__iexact=request.POST['name']).exclude(
        pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 570,
            'message': "Store with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            store = models.Store.objects.get(pk=request.POST['id'])
            store.name = request.POST['name']
            store.address = request.POST['address']
            store.country_id = request.POST['country_id']
            store.state_id = request.POST['state_id']
            store.city_id = request.POST['city_id']
            store.pin = request.POST['pin']
            # # # # #print(2727)
            store.contact_name = request.POST['contact_name']
            store.contact_no = request.POST['contact_no']
            store.contact_email = request.POST['contact_email']
            store.manager_name = request.POST['manager_name']
            if  request.POST['store_type1'] == 1:
                store.vendor_id = request.POST['vendor_id']
            # # # # #print(2734)
            store.store_type = request.POST.get('store_type2', 'r')
            store.updated_at = datetime.now()
            store.save()

            userId = request.COOKIES.get('userId', None)           
            user_log_details_add(userId,'Store Edit')
     
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 571,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeDelete(request):
    context = {}
    store = models.Store.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            store.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Store Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 572,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
def storeExport(request):
    keyword = request.GET.get('keyword')
    if keyword is not None and keyword != "":
        page_items = models.Store.objects.filter(Q(name__icontains=keyword) | Q(address__icontains=keyword) | Q(contact_name__icontains=keyword) | Q(
            contact_no__icontains=keyword) | Q(contact_email__icontains=keyword) | Q(manager_name__icontains=keyword) | Q(pin__icontains=keyword)).filter(status=1, deleted=0)
    else:
        page_items = models.Store.objects.filter(status=1, deleted=0)

    directory_path = settings.MEDIA_ROOT + '/reports/'
    path = Path(directory_path)
    path.mkdir(parents=True, exist_ok=True)

    for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
        if not f.endswith(".xlsx"):
            continue
        os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

    # tmpname = str(datetime.now().microsecond) + ".xlsx"
    tmpname = "Store" + ".xlsx"
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = "Name"
    ws['B1'] = "Address"
    ws['C1'] = "Contact Name"
    ws['D1'] = "Contact Number"
    ws['E1'] = "Contact Email"
    ws['F1'] = "Manager"
    ws['G1'] = "Pin"
    ws['H1'] = "City"
    ws['I1'] = "State"
    ws['J1'] = "Country"

    # Rows can also be appended
    for each in page_items:
        ws.append([each.name, each.address, each.contact_name, each.contact_no, each.contact_email,
                  each.manager_name, each.pin, each.city.name, each.state.name, each.country.name])

    # Save the file
    wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
    os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
    return JsonResponse({
        'code': 200,
        'filename': settings.MEDIA_URL + 'reports/' + tmpname,
        'name':  tmpname
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def billOfMaterialMasterList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        billOfMaterialMaster = list(models.Bill_Of_Material_Master.objects.filter(
            pk=id)[:1].values('pk', 'name', 'item__name', 'item__uom__name'))
        context.update({
            'status': 200,
            'message': "Bill Of Material Master Fetched Successfully.",
            'page_items': billOfMaterialMaster,
        })
        return JsonResponse(context)
    if keyword is not None and keyword != "":
        # #print(2937)
        billOfMaterialMaster = models.Bill_Of_Material_Master.objects.filter(
            Q(item__name__icontains=keyword) | Q(item__uom__name__icontains=keyword)).filter(status=1, deleted=0)
    else:
        billOfMaterialMaster = models.Bill_Of_Material_Master.objects.filter(
            status=1, deleted=0)
    billOfMaterialMaster = list(billOfMaterialMaster.values(
            'pk', 'name', 'item__name', 'item__uom__name'))
    if find_all is not None and int(find_all) == 1:
        context.update({
            'status': 200,
            'message': "Bill Of Materials Master Fetched Successfully.",
            'page_items': billOfMaterialMaster,
        })
        return JsonResponse(context)
    # #print(billOfMaterialMaster)
    per_page = int(env("PER_PAGE_DATA"))
    button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
    current_page = request.GET.get('current_page', 1)

    paginator = CustomPaginator(billOfMaterialMaster, per_page)
    page_items = paginator.get_page(current_page)
    total_pages = paginator.get_total_pages()

    context.update({
        'status': 200,
        'message': "BOM Levels Master Fetched Successfully.",
        'page_items': page_items,
        'total_pages': total_pages,
        'per_page': per_page,
        'current_page': int(current_page),
        'button_to_show': int(button_to_show),
    })
    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def billOfMaterialList(request):
    context = {}
    id = request.GET.get('id', None)
    bomMasterId = request.GET.get('bomMasterId', None)
    find_all = request.GET.get('find_all', None)
    level = request.GET.get('level', None)
    keyword = request.GET.get('keyword', None)
    item_id = request.GET.get('item_id', None)
    type_bom =  request.GET.get('type_bom', None)
    if id is not None and id != "":
        billOfMaterial = list(models.Bill_Of_Material.objects.filter(
            pk=id)[:1].values('pk', 'bom_item__name', 'uom__name', 'quantity', 'price','bom_master_id','bom_type'))
        context.update({
            'status': 200,
            'message': "Bill Of Material Fetched Successfully.",
            'page_items': billOfMaterial,
        })
   
    if item_id is not None and item_id != "":
        billOfMaterial = list(models.Bill_Of_Material.objects.filter(
            bom_item_id=item_id).values('pk', 'bom_master_id','bom_type'))
        context.update({
            'status': 200,
            'message': "Bill Of Material Fetched Successfully.",
            'page_items': billOfMaterial,
        })
    else:
        if bomMasterId is not None and bomMasterId !="":
            billOfMaterials = models.Bill_Of_Material.objects.filter(bom_master_id = bomMasterId)
        else :
            billOfMaterials = models.Bill_Of_Material.objects.filter(status=1, deleted=0)

        if keyword is not None and keyword != "":
            billOfMaterials = billOfMaterials.filter(
                Q(bom_item__name__icontains=keyword) | Q(uom__name__icontains=keyword) | Q(price__icontains=keyword)| Q(bom_type__icontains=keyword)).filter(status=1, deleted=0)
        else:
            billOfMaterials = billOfMaterials.filter(
                status=1, deleted=0)
        if type_bom is not None and type_bom !="":
            billOfMaterials =billOfMaterials.filter(bom_type=type_bom)

        if level is not None:
            billOfMaterials = billOfMaterials.filter(level__lte=level)
        billOfMaterials = list(billOfMaterials.values(
            'pk', 'bom_item__name', 'uom__name', 'quantity', 'price','bom_master_id','bom_type'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Bill Of Materials Fetched Successfully.",
                'page_items': billOfMaterials,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(billOfMaterials, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "BOM Levels Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def billOfMaterialAdd(request):
    context = {}
    # #print(2988)
    if not request.POST['bom_item_id'] or not request.POST['uom_id'] or not request.POST['total_amount'] or not request.POST['level']:
        context.update({
            'status': 573,
            'message': "BOM Item/UOM/Total Amount/Level has not been provided."
        })
        return JsonResponse(context)
    billOfMaterialHeaderTypeExist = models.Bill_Of_Material.objects.filter(bom_item_id = request.POST['bom_item_id'],bom_type = request.POST['type_bom']).exists()
    if billOfMaterialHeaderTypeExist :
        context.update({
        'status': 574,
        'message': "Bill Of Material with this item as BOM and type already exists.",
        })
        return JsonResponse(context)
    # exist_data = models.Bill_Of_Material.objects.filter(
    #     bom_item_id=request.POST['bom_item_id'], level=request.POST['level']).filter(deleted=0)
    # if len(exist_data) > 0:
    #     context.update({
    #         'status': 574,
    #         'message': "Bill Of Material with this item as BOM and level already exists.",
    #     })
    #     return JsonResponse(context)
    # #print(3003)
    try:
        with transaction.atomic():
            # #print(3005)
            billOfMaterialMasterHeaderExist = models.Bill_Of_Material_Master.objects.filter(item_id = request.POST['bom_item_id']).exists()
            
            if not billOfMaterialMasterHeaderExist:
                billOfMaterialHeaderMaster =   models.Bill_Of_Material_Master()
                billOfMaterialHeaderMaster.item_id = request.POST['bom_item_id']
                billOfMaterialHeaderMaster.save()
            else:
                # #print(request.POST['bom_item_id'])
                billOfMaterialHeaderMaster = models.Bill_Of_Material_Master.objects.filter(item_id = request.POST['bom_item_id']).first()
                # billOfMaterialHeaderMaster = models.Bill_Of_Material_Master.objects.get(item_id = int(request.POST['bom_item_id']))
            billOfMaterialHeader = models.Bill_Of_Material()
            billOfMaterialHeader.bom_item_id = request.POST['bom_item_id']
            billOfMaterialHeader.uom_id = request.POST['uom_id']
            billOfMaterialHeader.quantity = 1
            billOfMaterialHeader.bom_master_id = billOfMaterialHeaderMaster.id
            billOfMaterialHeader.bom_type = request.POST['type_bom']
            # #print(3029)
            billOfMaterialHeader.price = request.POST['total_amount']
            billOfMaterialHeader.level = request.POST['level']
            billOfMaterialHeader.save()
            if len(models.Bill_Of_Material_Detail.objects.filter(bom_level_id=billOfMaterialHeader.id)) == 0:
                billOfMaterialHeader.is_final = 1
                billOfMaterialHeader.save()
            bom_item=models.Item.objects.get(pk=billOfMaterialHeader.bom_item_id)
            bom_item.price=billOfMaterialHeader.price
            bom_item.save()
            bill_of_material_details = []
            for index, elem in enumerate(request.POST.getlist('bom_level_id')):
                billOfMaterialDetail = models.Bill_Of_Material.objects.get(
                    pk=elem)
                billOfMaterialDetail.is_final = 0
                billOfMaterialDetail.save()
                bill_of_material_details.append(models.Bill_Of_Material_Detail(bill_of_material_header_id=billOfMaterialHeader.id,
                                                bom_level_id=elem, quantity=request.POST.getlist('bom_level_quantity')[index], price=request.POST.getlist('bom_level_price')[index]))
            for index, elem in enumerate(request.POST.getlist('item_id')):
                bill_of_material_details.append(models.Bill_Of_Material_Detail(
                    bill_of_material_header_id=billOfMaterialHeader.id,
                    item_id=elem,
                    quantity=request.POST.getlist('item_quantity')[index],
                    price=request.POST.getlist('item_price')[index])
                )
            models.Bill_Of_Material_Detail.objects.bulk_create(bill_of_material_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'BOM add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Bill Of Material Created Successfully."
        })
    except Exception:
        context.update({
            'status': 575,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def billOfMaterialEdit(request):
    context = {}
    if not request.POST['bom_item_id'] or not request.POST['uom_id'] or not request.POST['total_amount'] or not request.POST['level']:
        context.update({
            'status': 576,
            'message': "BOM Item/UOM/Total Amount/Level has not been provided."
        })
        return JsonResponse(context)
    # exist_data = models.Bill_Of_Material.objects.filter(
    #     bom_item_id=request.POST['bom_item_id'], level=request.POST['level']).exclude(pk=request.POST['id']).filter(deleted=0)
    # if len(exist_data) > 0:
    #     context.update({
    #         'status': 577,
    #         'message': "Bill Of Material with this item as bom already exists.",
    #     })
    #     return JsonResponse(context)
    try:
        with transaction.atomic():
            billOfMaterialHeader = models.Bill_Of_Material.objects.prefetch_related(
                'bill_of_material_detail_set').get(pk=request.POST['id'])
            billOfMaterialHeader.bom_item_id = request.POST['bom_item_id']
            billOfMaterialHeader.uom_id = request.POST['uom_id']
            billOfMaterialHeader.quantity = 1
            billOfMaterialHeader.price = request.POST['total_amount']
            billOfMaterialHeader.level = request.POST['level']
            billOfMaterialHeader.updated_at = datetime.now()
            billOfMaterialHeader.save()
            bom_item = models.Item.objects.get(pk=billOfMaterialHeader.bom_item_id)
            bom_item.price = billOfMaterialHeader.price
            bom_item.save()

            for bom_detail in models.Bill_Of_Material_Detail.objects.filter(bom_level_id=request.POST['id']):
                bom_detail.price=float(bom_detail.quantity) * float(request.POST['total_amount'])
                bom_detail.save()

            if len(models.Bill_Of_Material_Detail.objects.filter(bom_level_id=billOfMaterialHeader.id)) == 0:
                billOfMaterialHeader.is_final = 1
                billOfMaterialHeader.save()
            billOfMaterialHeader.bill_of_material_detail_set.all().delete()
            bill_of_material_details = []
            for index, elem in enumerate(request.POST.getlist('bom_level_id')):
                billOfMaterialDetail = models.Bill_Of_Material.objects.get(pk=elem)
                billOfMaterialDetail.is_final = 0
                billOfMaterialDetail.save()
                bill_of_material_details.append(models.Bill_Of_Material_Detail(bill_of_material_header_id=billOfMaterialHeader.id,
                                                bom_level_id=elem, quantity=request.POST.getlist('bom_level_quantity')[index], price=request.POST.getlist('bom_level_price')[index]))
            for index, elem in enumerate(request.POST.getlist('item_id')):
                bill_of_material_details.append(models.Bill_Of_Material_Detail(bill_of_material_header_id=billOfMaterialHeader.id,
                                                item_id=elem, quantity=request.POST.getlist('item_quantity')[index], price=request.POST.getlist('item_price')[index]))
            models.Bill_Of_Material_Detail.objects.bulk_create(
                bill_of_material_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'BOM Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Bill Of Material Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 578,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def billOfMaterialDelete(request):
    context = {}
    bomLevel = models.Bill_Of_Material.objects.get(
        pk=request.POST['id'])
    try:
        with transaction.atomic():
            bomLevel.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'BOM Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Bill Of Material Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 579,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


def getStructureOfBOM(bom_id):
    billOfMaterial = list(models.Bill_Of_Material.objects.filter(pk=bom_id)[:1].values('pk', 'bom_item_id' , 'bom_item__name', 'uom__name', 'quantity', 'price'))[0]
    childBOMS = models.Bill_Of_Material_Detail.objects.filter(
        status=1, deleted=0, bill_of_material_header_id=bom_id).order_by('bom_level_id')
    id_lists = []
    for each in childBOMS:
        id_lists.append(each.id)
    structure = []
    for id in id_lists:
        childDetail = models.Bill_Of_Material_Detail.objects.get(pk=id)
        each_child_structure = {}
        each_child_structure['pk'] = id
        each_child_structure['quantity'] = childDetail.quantity
        each_child_structure['price'] = childDetail.price
        if childDetail.item_id is not None:
            each_child_structure['item'] = list(models.Item.objects.filter(pk=childDetail.item_id)[:1].values(
                'pk', 'name', 'item_type__name', 'item_type__item_category__name', 'uom__name', 'price'))[0]
        if childDetail.bom_level_id is not None:
            each_child_structure['bom'] = getStructureOfBOM(
                childDetail.bom_level_id)
        structure.append(each_child_structure)
    # #print(billOfMaterial)
    billOfMaterial['structure'] = structure
    return billOfMaterial


def getBillOfMaterialStructure(request):
    context = {}
    bom_id = request.GET.get('bom_id', None)
    if bom_id is not None and bom_id != "":
        billOfMaterial = getStructureOfBOM(bom_id)
        context.update({
            'status': 200,
            'message': "BOM Structure fetched successfully",
            'page_items': billOfMaterial,
        })
    else:
        context.update({
            'status': 580,
            'message': "Please Provide valid BOM id."
        })
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def jobOrderBomDetails(request):
    context = {}
    id = request.GET.get('id', None)
    try:
        details = models.Bill_Of_Material_Detail.objects.filter(bill_of_material_header_id=id)
        result = []
        for detail in details:
            # Case 1: item_id is not null
            if detail.item_id:
                result.append({
                    "item_id" : detail.item_id,
                    "uom_name" : detail.item.uom.name,  # Fetch the uom_id directly from the item
                    "quantity" : detail.quantity
                })
            # Case 2: bom_level_id is not null
            elif detail.bom_level_id:
                bom_item = detail.bom_level.bom_item  # Navigate through bom_level to bom_item
                if bom_item:
                    result.append({
                        "item_id" : bom_item.id,
                        "uom_name" : bom_item.uom.name,  # Fetch the uom_id from the bom_item
                        "quantity" : detail.quantity
                    })
        # #print(result)
        context.update({
            'status': 200,
            'message': "Bill Of Material- items Fetched Successfully.",
            'page_items': result,
        })
    except Exception:
        context.update({
            'status': 568,
            'message': "Something Went Wrong. Please Try Again."
        })
    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def purchaseOrderList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    vendor_id = request.GET.get('vendor_id', None)
    delivery_status = request.GET.get('delivery_status', None)
    if id is not None and id != "":
        purchaseOrder = list(models.Purchase_Order.objects.filter(pk=id)[:1].values('pk', 'order_number', 'order_date', 'quotation_number', 'quotation_date', 'reference_number', 'business_terms', 'discount_type',
                             'discount_value', 'discounted_value', 'excise_duty_percentage', 'insurance', 'octroi', 'freight', 'packing', 'payment_terms', 'delivery_schedule', 'delivery_at', 'notes', 'total_amount', 'vendor__name'))
        context.update({
            'status': 200,
            'message': "Purchase Order Fetched Successfully.",
            'page_items': purchaseOrder,
        })
    else:
        if vendor_id is not None and vendor_id != "":
            purchaseOrders = models.Purchase_Order.objects.filter(
                vendor_id=vendor_id).filter(status=1, deleted=0)
        else:
            purchaseOrders = models.Purchase_Order.objects.filter(
                status=1, deleted=0)
        if delivery_status is not None and delivery_status != "":
            purchaseOrders = purchaseOrders.filter(
                delivery_status__in=delivery_status.split(","))
        if keyword is not None and keyword != "":
            purchaseOrders = purchaseOrders.filter(Q(vendor__name__icontains=keyword) | Q(order_number__icontains=keyword) | Q(
                order_date__icontains=keyword) | Q(total_amount__icontains=keyword)).filter(status=1, deleted=0)
        purchaseOrders = list(purchaseOrders.values('pk', 'order_number', 'order_date', 'quotation_number', 'quotation_date', 'reference_number', 'business_terms', 'discount_type', 'discount_value',
                              'discounted_value', 'excise_duty_percentage', 'insurance', 'octroi', 'freight', 'packing', 'payment_terms', 'delivery_schedule', 'delivery_at', 'notes', 'total_amount', 'vendor__name'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Purchase Orders Fetched Successfully.",
                'page_items': purchaseOrders,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(purchaseOrders, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Purchase Orders Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseOrderAdd(request):
    context = {}
    if not request.POST['vendor_id'] or not request.POST['order_number'] or not request.POST['order_date'] or not request.POST['quotation_number'] or not request.POST['quotation_date'] or not request.POST['total_amount']:
        context.update({
            'status': 581,
            'message': "Vendor/Order Number/Order Date/Quotation Number/Quotation Date/Total Amount has not been provided."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            purchaseOrderHeader = models.Purchase_Order()
            purchaseOrderHeader.vendor_id = request.POST['vendor_id']
            purchaseOrderHeader.order_number = request.POST['order_number']
            purchaseOrderHeader.order_date = request.POST['order_date']
            purchaseOrderHeader.quotation_number = request.POST['quotation_number']
            purchaseOrderHeader.quotation_date = request.POST['quotation_date']
            purchaseOrderHeader.reference_number = request.POST['reference_number']
            purchaseOrderHeader.business_terms = request.POST['business_terms']
            purchaseOrderHeader.discount_type = request.POST['discount_type']
            purchaseOrderHeader.discount_value = request.POST[
                'discount_value'] if request.POST['discount_value'] != "" else 0
            purchaseOrderHeader.discounted_value = request.POST[
                'discounted_value'] if request.POST['discounted_value'] != "" else 0
            purchaseOrderHeader.excise_duty_percentage = request.POST[
                'excise_duty_percentage'] if request.POST['excise_duty_percentage'] != "" else 0
            purchaseOrderHeader.insurance = request.POST['insurance'] if request.POST['insurance'] != "" else 0
            purchaseOrderHeader.octroi = request.POST['octroi'] if request.POST['octroi'] != "" else 0
            purchaseOrderHeader.freight = request.POST['freight'] if request.POST['freight'] != "" else 0
            purchaseOrderHeader.packing = request.POST['packing']
            purchaseOrderHeader.payment_terms = request.POST['payment_terms']
            purchaseOrderHeader.delivery_schedule = request.POST['delivery_schedule']
            purchaseOrderHeader.delivery_at = request.POST['delivery_at']
            purchaseOrderHeader.notes = request.POST['notes']
            purchaseOrderHeader.total_amount = request.POST['total_amount']
            purchaseOrderHeader.creator_id = request.POST['user_id']
            purchaseOrderHeader.save()

            purchase_order_details = []
            for index, elem in enumerate(request.POST.getlist('item_id')):
                purchase_order_details.append(
                    models.Purchase_Order_Detail(
                        purchase_order_header_id=purchaseOrderHeader.id,
                        item_id=elem,
                        quantity=request.POST.getlist('item_quantity')[index],
                        rate=request.POST.getlist('rate')[index],
                        amount=request.POST.getlist('item_price')[index],
                        gst_percentage=request.POST.getlist(
                            'gst_percentage')[index],
                        amount_with_gst=request.POST.getlist(
                            'amount_with_gst')[index]
                    )
                )
            models.Purchase_Order_Detail.objects.bulk_create(
                purchase_order_details)
            user_log_details_add(request.POST['user_id'],'add purchase order')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Purchase Order Created Successfully."
        })
    except Exception:
        context.update({
            'status': 582,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseOrderEdit(request):
    context = {}
    if not request.POST['vendor_id'] or not request.POST['order_number'] or not request.POST['order_date'] or not request.POST['quotation_number'] or not request.POST['quotation_date'] or not request.POST['total_amount']:
        context.update({
            'status': 583,
            'message': "Vendor/Order Number/Order Date/Quotation Number/Quotation Date/Total Amount has not been provided."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related(
                'purchase_order_detail_set').get(pk=request.POST['id'])
            purchaseOrderHeader.vendor_id = request.POST['vendor_id']
            purchaseOrderHeader.order_number = request.POST['order_number']
            purchaseOrderHeader.order_date = request.POST['order_date']
            purchaseOrderHeader.quotation_number = request.POST['quotation_number']
            purchaseOrderHeader.quotation_date = request.POST['quotation_date']
            purchaseOrderHeader.reference_number = request.POST['reference_number']
            purchaseOrderHeader.business_terms = request.POST['business_terms']
            purchaseOrderHeader.discount_type = request.POST['discount_type']
            purchaseOrderHeader.discount_value = request.POST[
                'discount_value'] if request.POST['discount_value'] != "" else 0
            purchaseOrderHeader.discounted_value = request.POST[
                'discounted_value'] if request.POST['discounted_value'] != "" else 0
            purchaseOrderHeader.excise_duty_percentage = request.POST[
                'excise_duty_percentage'] if request.POST['excise_duty_percentage'] != "" else 0
            purchaseOrderHeader.insurance = request.POST['insurance'] if request.POST['insurance'] != "" else 0
            purchaseOrderHeader.octroi = request.POST['octroi'] if request.POST['octroi'] != "" else 0
            purchaseOrderHeader.freight = request.POST['freight'] if request.POST['freight'] != "" else 0
            purchaseOrderHeader.packing = request.POST['packing']
            purchaseOrderHeader.payment_terms = request.POST['payment_terms']
            purchaseOrderHeader.delivery_schedule = request.POST['delivery_schedule']
            purchaseOrderHeader.delivery_at = request.POST['delivery_at']
            purchaseOrderHeader.notes = request.POST['notes']
            purchaseOrderHeader.total_amount = request.POST['total_amount']
            purchaseOrderHeader.creator_id = request.POST['user_id']
            purchaseOrderHeader.updated_at = datetime.now()
            purchaseOrderHeader.save()
            purchaseOrderHeader.purchase_order_detail_set.all().delete()

            purchase_order_details = []
            for index, elem in enumerate(request.POST.getlist('item_id')):
                purchase_order_details.append(
                    models.Purchase_Order_Detail(
                        purchase_order_header_id=purchaseOrderHeader.id,
                        item_id=elem,
                        quantity=request.POST.getlist('item_quantity')[index],
                        rate=request.POST.getlist('rate')[index],
                        amount=request.POST.getlist('item_price')[index],
                        gst_percentage=request.POST.getlist(
                            'gst_percentage')[index],
                        amount_with_gst=request.POST.getlist(
                            'amount_with_gst')[index]
                    )
                )
            models.Purchase_Order_Detail.objects.bulk_create(
                purchase_order_details)
            user_log_details_add(request.POST['user_id'],'Edit purchase order')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Purchase Order Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 584,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseOrderDelete(request):
    context = {}
    purchaseOrder = models.Purchase_Order.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            purchaseOrder.delete()
            user_log_details_add(request.POST['creator_id'],'delete purchase order')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Purchase Order Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 585,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def purchaseOrderDetails(request):
    context = {}
    header_id = request.GET.get('header_id', None)
    if header_id is not None and header_id != "":
        header_detail = list(models.Purchase_Order.objects.filter(id=header_id).values('pk', 'order_number', 'order_date', 'quotation_number', 'quotation_date', 'reference_number', 'business_terms', 'discount_type',
                             'discount_value', 'discounted_value', 'excise_duty_percentage', 'insurance', 'octroi', 'freight', 'packing', 'payment_terms', 'delivery_schedule', 'delivery_at', 'notes', 'total_amount', 'delivered_total_amount'))
        orderDetails = list(models.Purchase_Order_Detail.objects.filter(purchase_order_header_id=header_id).values('pk', 'quantity', 'rate', 'amount', 'gst_percentage', 'amount_with_gst', 'delivered_quantity',
                            'delivered_rate', 'delivered_amount', 'delivered_gst_percentage', 'delivered_amount_with_gst', 'item_id', 'item__name', 'purchase_order_header_id', 'purchase_order_header__order_number'))
        context.update({
            'status': 200,
            'message': "Purchase Order Details Fetched Successfully.",
            'header_detail': header_detail,
            'page_items': orderDetails,
        })
    else:
        context.update({
            'status': 588,
            'message': "Please Provide Header Id.",
        })
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactionTypeList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    if id is not None and id != "":
        transactionType = list(models.Transaction_Type.objects.filter(pk=id)[:1].values('pk', 'name','desc'))
        context.update({
            'status': 200,
            'message': "Transaction Type Fetched Successfully.",
            'page_items': transactionType,
        })
    else:
        if keyword is not None and keyword != "":
            transactionTypes = list(models.Transaction_Type.objects.filter(
                name__icontains=keyword, status=1, deleted=0).values('pk', 'name','desc'))
        else:
            transactionTypes = list(models.Transaction_Type.objects.filter(
                status=1, deleted=0).values('pk', 'name','desc'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Transaction Types Fetched Successfully.",
                'page_items': transactionTypes,
            })
            return JsonResponse(context)
        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(transactionTypes, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Transaction Types Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transactionTypeAdd(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 531,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Transaction_Type.objects.filter(
        name=request.POST['name']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 532,
            'message': "Transaction Type with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            transactionType = models.Transaction_Type()
            transactionType.name = request.POST['name']
            transactionType.desc = request.POST['desc']
            transactionType.save()
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Transaction Type Created Successfully."
        })
    except Exception:
        context.update({
            'status': 533,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transactionTypeEdit(request):
    context = {}
    if not request.POST['name']:
        context.update({
            'status': 534,
            'message': "Name has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Transaction_Type.objects.filter(
        name=request.POST['name']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 535,
            'message': "Transaction Type with this name already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            transactionType = models.Transaction_Type.objects.get(pk=request.POST['id'])
            transactionType.name = request.POST['name']
            transactionType.updated_at = datetime.now()
            transactionType.save()
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Transaction Type Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 536,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transactionTypeDelete(request):
    context = {}
    transactionType = models.Transaction_Type.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            transactionType.delete()
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Transaction Type Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 537,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def storeItemList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    storeId = request.GET.get('storeId', None)
    itemTypeId = request.GET.get('itemTypeId', None)
    itemCatId = request.GET.get('itemCatId', None)
    itemId = request.GET.get('itemId', None)
    if id is not None and id != "":
        storeItem = list(models.Store_Item.objects.filter(pk=id)[:1].values(
            'pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty'))
        context.update({
            'status': 200,
            'message': "Store Item Fetched Successfully.",
            'page_items': storeItem,
        })
    elif storeId is not None and storeId != "":
        if itemTypeId is not None and itemTypeId != "" :
            if itemCatId is not None and itemCatId != "" :
                # # # # # #print("2890")
                storeItems = list(models.Store_Item.objects.filter(store_id = storeId ,item__item_type_id = itemTypeId , item__item_type__item_category_id = itemCatId).values(
                    'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price','item__item_type_id'))
                # # # # # #print(storeItems)
                per_page = int(env("PER_PAGE_DATA"))
                button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
                current_page = request.GET.get('current_page', 1)

                paginator = CustomPaginator(storeItems, per_page)
                page_items = paginator.get_page(current_page)
                total_pages = paginator.get_total_pages()

                context.update({
                    'status': 200,
                    'message': "Store Items Fetched Successfully.",
                    'page_items': page_items,
                    'total_pages': total_pages,
                    'per_page': per_page,
                    'current_page': int(current_page),
                    'button_to_show': int(button_to_show),
                })
        elif itemId is not None and itemId != "" :
            storeItem = list(models.Store_Item.objects.filter(store_id = storeId ,item_id = itemId)[:1].values(
                    'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price','item__item_type_id'))
            context.update({
                'status': 200,
                'message': "Store Item Fetched Successfully.",
                'page_items': storeItem,
            })
        
        else:
            storeItem = list(models.Store_Item.objects.filter(store_id = storeId ).values(
            'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price'))
            context.update({
                'status': 200,
                'message': "Store Fetched Successfully.",
                'page_items': storeItem,
            })
        return JsonResponse(context)
    else:
        if keyword is not None and keyword != "":
            storeItems = list(
                models.Store_Item.objects.filter(
                    Q(store__name__icontains=keyword) | Q(item__name__icontains=keyword) | Q(
                        opening_qty__icontains=keyword) | Q(on_hand_qty__icontains=keyword) | Q(closing_qty__icontains=keyword)
                ).filter(
                    status=1, deleted=0).values('pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty')
            )
        else:
            storeItems = list(models.Store_Item.objects.filter(status=1, deleted=0).values(
                'pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Store Items Fetched Successfully.",
                'page_items': storeItems,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(storeItems, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Store Items Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def storeItemTrackingList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    storeId = request.GET.get('storeId', None)
    itemTypeId = request.GET.get('itemTypeId', None)
    itemCatId = request.GET.get('itemCatId', None)
    itemId = request.GET.get('itemId', None)
    if id is not None and id != "":
        storeItem = list(models.Store_Item_Current.objects.filter(pk=id)[:1].values(
            'pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty'))
        context.update({
            'status': 200,
            'message': "Store Item tracking Fetched Successfully.",
            'page_items': storeItem,
        })
    elif storeId is not None and storeId != "":
        if itemTypeId is not None and itemTypeId != "" :
            if itemCatId is not None and itemCatId != "" :
                # # # # # #print("2890")
                storeItems = list(models.Store_Item_Current.objects.filter(store_id = storeId ,item__item_type_id = itemTypeId , item__item_type__item_category_id = itemCatId,status=1, deleted=0).values(
                    'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price','item__item_type_id','transaction_date','store_transaction_id','store_transaction__transaction_number').order_by('transaction_date', 'created_at', 'item_id'))
                # # # # # #print(storeItems)
                per_page = int(env("PER_PAGE_DATA"))
                button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
                current_page = request.GET.get('current_page', 1)

                paginator = CustomPaginator(storeItems, per_page)
                page_items = paginator.get_page(current_page)
                total_pages = paginator.get_total_pages()

                context.update({
                    'status': 200,
                    'message': "Store Item tracking Fetched Successfully.",
                    'page_items': page_items,
                    'total_pages': total_pages,
                    'per_page': per_page,
                    'current_page': int(current_page),
                    'button_to_show': int(button_to_show),
                })
        elif itemId is not None and itemId != "" :
            storeItem = list(models.Store_Item_Current.objects.filter(store_id = storeId ,item_id = itemId,status=1, deleted=0)[:1].values(
                    'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price','item__item_type_id','transaction_date','store_transaction_id','store_transaction__transaction_number','quantity_Transfer').order_by('transaction_date', 'created_at', 'item_id'))
            context.update({
                'status': 200,
                'message': "Store Item tracking Fetched Successfully.",
                'page_items': storeItem,
            })
        
        else:
            storeItem = list(models.Store_Item_Current.objects.filter(store_id = storeId ).values(
            'pk', 'store__name','item_id','item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','item__price','transaction_date','store_transaction_id','store_transaction__transaction_number','quantity_Transfer').order_by('transaction_date', 'created_at', 'item_id'))
            context.update({
                'status': 200,
                'message': "Store Item tracking Fetched Successfully.",
                'page_items': storeItem,
            })
        return JsonResponse(context)
    else:
        if keyword is not None and keyword != "":
            storeItems = list(
                models.Store_Item_Current.objects.filter(
                    Q(store__name__icontains=keyword) | Q(item__name__icontains=keyword) | Q(
                        opening_qty__icontains=keyword) | Q(on_hand_qty__icontains=keyword) | Q(closing_qty__icontains=keyword)|Q(transaction_date__icontains=keyword)
                ).filter(
                    status=1, deleted=0).values('pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','transaction_date','store_transaction_id','store_transaction__transaction_number','quantity_Transfer').order_by('transaction_date', 'created_at', 'item_id')
            )
        else:
            storeItems = list(models.Store_Item_Current.objects.filter(status=1, deleted=0).values(
                'pk', 'store__name', 'item__name', 'opening_qty', 'on_hand_qty', 'closing_qty','transaction_date','store_transaction_id','store_transaction__transaction_number','quantity_Transfer').order_by('transaction_date', 'created_at', 'item_id'))
        if find_all is not None and int(find_all) == 1:
            context.update({
                'status': 200,
                'message': "Store Item tracking Fetched Successfully.",
                'page_items': storeItems,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(storeItems, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Store Item tracking Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeItemAdd(request):
    context = {}
   
    if not request.POST['store_id'] or not request.POST['item_id'] or not request.POST['opening_qty']:
        context.update({
            'status': 586,
            'message': "Store/Item/Opening Quantity has not been provided."
        })
        return JsonResponse(context)
    exist_data = models.Store_Item.objects.filter(
        store=request.POST['store_id'], item=request.POST['item_id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 587,
            'message': "Item in this Store already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            storeItem = models.Store_Item()
            storeItem.store_id = request.POST['store_id']
            storeItem.item_id = request.POST['item_id']
            storeItem.opening_qty = Decimal(request.POST['opening_qty'])
            storeItem.on_hand_qty = Decimal(request.POST['opening_qty'])
            storeItem.closing_qty = Decimal(request.POST['opening_qty'])
            storeItem.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Store Item Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Item Created Successfully."
        })
    except Exception:
        context.update({
            'status': 588,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeItemEdit(request):
    context = {}
    if not request.POST['store_id'] or not request.POST['item_id'] or not request.POST['opening_qty']:
        context.update({
            'status': 589,
            'message': "Store/Item/Opening Quantity has not been provided."
        })
        return JsonResponse(context)
        return JsonResponse(context)
    exist_data = models.Store_Item.objects.filter(
        store=request.POST['store_id'], item=request.POST['item_id']).exclude(pk=request.POST['id']).filter(deleted=0)
    if len(exist_data) > 0:
        context.update({
            'status': 590,
            'message': "Item in this Store already exists.",
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            storeItem = models.Store_Item.objects.get(pk=request.POST['id'])
            storeItem.store_id = request.POST['store_id']
            storeItem.item_id = request.POST['item_id']
            storeItem.opening_qty = Decimal(request.POST['opening_qty'])
            storeItem.on_hand_qty = Decimal(request.POST['opening_qty'])
            storeItem.closing_qty = Decimal(request.POST['opening_qty'])
            storeItem.updated_at = datetime.now()
            storeItem.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Store Item Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Item Updated Successfully."
        })
    except Exception:
        context.update({
            'status': 591,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeItemDelete(request):
    context = {}
    storeItem = models.Store_Item.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            storeItem.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Store Item Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Item Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 592,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def stockTransfer(request):
    context = {}
    try:
        with transaction.atomic():
            from_store_item=models.Store_Item.objects.get(store_id=request.POST["store_id"], item_id=request.POST["from_item_id"])
            from_store_item.on_hand_qty-=Decimal(request.POST["quantity"])
            from_store_item.closing_qty-=Decimal(request.POST["quantity"])
            from_store_item.save()
            # change in storeItemCurrent mout
            # Fetch the last transaction_date less than the given_date if mout
            given_date = request.POST['transaction_date']
            
            
            # Check for the last record on the given_date
            record = models.Store_Item_Current.objects.filter(transaction_date=given_date,store_id=request.POST["store_id"], item_id=request.POST["from_item_id"],status=1, deleted=0 ).last()

            if not record:
                # If no record is found for the given_date, look for the last record before that date
                last_transaction_date = models.Store_Item_Current.objects.filter(
                    transaction_date__lt=given_date ,store_id=request.POST["store_id"], item_id=request.POST["from_item_id"],status=1, deleted=0
                ).aggregate(Max('transaction_date'))['transaction_date__max']

                

                if last_transaction_date:
                    # Fetch the record for the last_transaction_date
                    record = models.Store_Item_Current.objects.filter(
                        transaction_date=last_transaction_date ,store_id=request.POST["store_id"], item_id=request.POST["from_item_id"],status=1, deleted=0
                    ).last()

            # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
            
            
            store_item_instance = models.Store_Item_Current()

            if record:
                # Set values based on the last record found
                store_item_instance.opening_qty = record.closing_qty
                store_item_instance.on_hand_qty = record.closing_qty - Decimal(request.POST["quantity"])
                store_item_instance.closing_qty = record.closing_qty -  Decimal(request.POST["quantity"])
                if(store_item_instance.on_hand_qty<0):
                    raise ValueError(f"out quantity is more than available quantity")
                quant = request.POST["quantity"]
                SoucestoreId = request.POST["from_item_id"]
                destiStore = request.POST["to_item_id"]

                store_item_instance.quantity_Transfer = f"{quant} get transafer from {SoucestoreId} to { destiStore} "
            

                # Set other fields for the new transaction
               
                store_item_instance.transaction_date = given_date
                store_item_instance.item_id = request.POST["from_item_id"]
                store_item_instance.store_id = request.POST["store_id"]

                # Save the instance to the database
                store_item_instance.save()
                
            else:
                message = "canot possible"
                raise ValueError(f"message item not exist in store")
            store_item_curreEdit(request.POST["store_id"], request.POST["from_item_id"],given_date,'mout', request.POST["quantity"]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

            to_store_item = models.Store_Item.objects.get(store_id=request.POST["store_id"], item_id=request.POST["to_item_id"])
            to_store_item.on_hand_qty += Decimal(request.POST["quantity"])
            to_store_item.closing_qty += Decimal(request.POST["quantity"])
            to_store_item.save()

            # change in storeItemCurrent min
            # Fetch the last transaction_date less than the given_date
            given_date = request.POST['transaction_date']
            
            
            # Check for the last record on the given_date
            record = models.Store_Item_Current.objects.filter(transaction_date=given_date,store_id=request.POST["store_id"], item_id=request.POST["to_item_id"],status=1, deleted=0).last()

            if not record:
                # If no record is found for the given_date, look for the last record before that date
                last_transaction_date = models.Store_Item_Current.objects.filter(
                    transaction_date__lt=given_date,store_id=request.POST["store_id"], item_id=request.POST["to_item_id"],status=1, deleted=0
                ).aggregate(Max('transaction_date'))['transaction_date__max']

                

                if last_transaction_date:
                    # Fetch the record for the last_transaction_date
                    record = models.Store_Item_Current.objects.filter(
                        transaction_date=last_transaction_date,store_id=request.POST["store_id"], item_id=request.POST["to_item_id"],status=1, deleted=0
                    ).last()

            # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
            
            
            store_item_instance = models.Store_Item_Current()

            if record:
                # Set values based on the last record found
                store_item_instance.opening_qty = record.closing_qty
                store_item_instance.on_hand_qty = record.closing_qty + Decimal(request.POST["quantity"])
                store_item_instance.closing_qty = record.closing_qty + Decimal(request.POST["quantity"])
                if(store_item_instance.on_hand_qty<0):
                    raise ValueError(f"on_hand_qty quantity is less than 0")
            else:
                # Set values based on the current transaction if no prior record exists
                store_item_instance.opening_qty = Decimal(0.00)
                store_item_instance.on_hand_qty = Decimal(request.POST["quantity"])
                store_item_instance.closing_qty = Decimal(request.POST["quantity"])

            # Set other fields for the new transaction
            itemId = request.POST["from_item_id"]
            storeId = request.POST["to_item_id"]
            quant = request.POST["quantity"]
            store_item_instance.quantity_Transfer = f"{quant} get transafer from {itemId} to { storeId} "
            store_item_instance.transaction_date = given_date
            store_item_instance.item_id = request.POST["to_item_id"]
            store_item_instance.store_id = request.POST["store_id"]

            # Save the instance to the database
            store_item_instance.save()
            
            store_item_curreEdit(request.POST["store_id"],request.POST["to_item_id"],given_date,'min', request.POST["quantity"]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Transfer completed Successfully."
        })
    except Exception as e:
        context.update({
            'status': 588,
            'message': f"Something Went Wrong. Please Try Again.{e}"
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
def storeItemExport(request):
    keyword = request.GET.get('keyword')
    if keyword is not None and keyword != "":
        page_items = models.Store_Item.objects.filter(Q(store__name__icontains=keyword) | Q(
            item__name__icontains=keyword) | Q(opening_qty__icontains=keyword)).filter(status=1, deleted=0)
    else:
        page_items = models.Store_Item.objects.filter(status=1, deleted=0)
    directory_path = settings.MEDIA_ROOT + '/reports/'
    path = Path(directory_path)
    path.mkdir(parents=True, exist_ok=True)

    for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
        if not f.endswith(".xlsx"):
            continue
        os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

    # tmpname = str(datetime.now().microsecond) + ".xlsx"
    tmpname = "Store Item" + ".xlsx"
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = "Store"
    ws['B1'] = "Item"
    ws['C1'] = "Opening quantity"
    ws['D1'] = "On hand quantity"
    ws['E1'] = "Closing quantity"

    # Rows can also be appended
    for each in page_items:
        ws.append([each.store.name, each.item.name, each.opening_qty,
                  each.on_hand_qty, each.closing_qty])

    # Save the file
    wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
    os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
    return JsonResponse({
        'code': 200,
        'filename': settings.MEDIA_URL + 'reports/' + tmpname,
        'name':  tmpname
    })


@api_view(['POST'])
def storeItemImport(request):
    context = {}
    if request.FILES.get('file'):
        
        excel = request.FILES['file']
        # trying to process files without error
        try:
            df = pd.read_excel(excel)
            df.columns = [col.strip().lower() for col in df.columns]
            for index, row in df.iterrows():
                # trying to fetch required cells from 
                try:
                    store_name = row['store']
                    item_name = row['item']
                    opening_quantity = row['opening quantity']
                    on_hand_quantity = row['on hand quantity']
                    closing_quantity= row['closing quantity']
                    # Skip the row if any required field is empty
                    if not all([store_name, item_name, opening_quantity, on_hand_quantity, closing_quantity]):
                        continue  # Skip this row and move to the next one
                    if not models.Store_Item.objects.filter(store__name__iexact=store_name,item__name__iexact=item_name).exists():
                        
                        if (
                            models.Store.objects.filter(name__iexact=store_name).exists() and
                            models.Item.objects.filter(name__iexact=item_name).exists()
                        ):
                            try:
                                with transaction.atomic():
                                    obj = models.Store_Item(
                                        store=models.Store.objects.get(name__iexact=store_name), 
                                        item=models.Item.objects.get(name__iexact=item_name),
                                        opening_qty=Decimal(opening_quantity),
                                        on_hand_qty=Decimal(on_hand_quantity),
                                        closing_qty=Decimal(closing_quantity)
                                    )
                                    obj.save(obj)
                                    userId = request.COOKIES.get('userId', None)
                                    user_log_details_add(userId,'store Item Bulk Import')
                                transaction.commit()
                                context.update({
                                    'status': 200,
                                    'message': "srore Items Created Successfully."
                                })
                            except Exception:
                                
                                context.update({
                                    'status': 568,
                                    'message': "store Items cannot be created something wrong"
                                })
                                transaction.rollback()
                            
                except KeyError as e:
                    # Handle missing columns
                    context.update({
                        'status': 568,
                        'message': "Required column missing"
                    })
                    return JsonResponse(context)
            if not context:
                context.update({
                'status': 568,
                'message': "No new store items in the excel"
            })
        except Exception as e:
            context.update({
                'status': 568,
                'message': "Error processing file"
            })
    else:
        context.update({
            'status': 568,
            'message': "File has not been uploaded"
        })
    return JsonResponse(context)



# @api_view(['GET'])
# def storeItemReportExport(request):
#     # # # # # #print(request.GET)
#     keyword = request.GET.get('keyword')
#     file_type=request.GET.get('file_type')
#     # # # # # #print(keyword)
#     # # # # # #print("3314")
#     # return JsonResponse({})
#     if keyword is not None and keyword != "":
#         # # # # # #print("3316")
#         page_items = models.Store_Item.objects.filter(item__item_type__item_category_id=keyword,status =1 , deleted=0).order_by('store_id','item_id')
        
#     else:
#         page_items = models.Store_Item.objects.filter(status=1 , deleted = 0 ).order_by('store_id','store_id')
#         # page_items = models.Store_Item.objects.raw("SELECT * FROM store_items GROUP BY store_id ,item_id ")    
#     # for p in page_items:
#     #     # # # # #print(p.item_id)
#     # return JsonResponse({})
#     # # # # # #print(page_items)
#     if file_type=="xlsx":
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)

#         for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
#             if not f.endswith(".xlsx"):
#                 continue
#             os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

#         # tmpname = str(datetime.now().microsecond) + ".xlsx"
#         tmpname = "Store Item Report" + ".xlsx"
#         wb = Workbook()

#         # grab the active worksheet
#         ws = wb.active

#         # Data can be assigned directly to cells
#         ws['A1'] = "Store"
#         ws['B1'] = "Item"
#         ws['C1'] = "On hand quantity"
#         ws['D1'] = "Item Catagory"
#         ws['E1'] = "value"

#         # Rows can also be appended
#         for each in page_items:
#             val = float(each.on_hand_qty)* float(each.item.price)
#             ws.append([each.store.name, each.item.name,each.on_hand_qty, 
#             each.item.item_type.item_category.name, val ])

#         # Save the file
#         wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
#         os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name':  tmpname
#         })
#     elif file_type == "csv":
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)

#         # Clean up any existing CSV files in the directory
#         for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
#             if not f.endswith(".csv"):
#                 continue
#             os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

#         tmpname = "Store Item Report" + ".csv"

#         with open(os.path.join(directory_path, tmpname), 'w', newline='') as csvfile:
#             writer = csv.writer(csvfile)
#             writer.writerow(["Store", "Item","On hand quantity", "Item Category", "Value"])
#             for each in page_items:
#                 val = float(each.on_hand_qty)* float(each.item.price)
#                 writer.writerow(
#                     [each.store.name, each.item.name,each.on_hand_qty, 
#             each.item.item_type.item_category.name, val])

#         os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name': tmpname
#         })

#     elif file_type == "pdf":
#         # # # # # #print("3395")
#         # Create a new PDF document with smaller margins
#         pdf = FPDF(orientation='P', unit='mm', format='A4')  # Adjust unit and format if needed
#         pdf.set_left_margin(5)
#         pdf.set_top_margin(5)

#         pdf.add_page()
#         pdf.set_font("Arial", size=9)

#         # Add a header row with bold text
#         pdf.set_font("Arial", size=9, style='B')  # Set font style to bold
#         pdf.cell(9, 10, txt="S.No.", border=1, align='C')  # Add S.No. column
#         pdf.cell(50, 10, txt="Store", border=1, align='C')
#         pdf.cell(50, 10, txt="Item", border=1, align='C')
#         pdf.cell(40, 10, txt="On hand quantity", border=1, align='C')
#         pdf.cell(30, 10, txt="Item Category", border=1, align='C')
#         pdf.cell(20, 10, txt="Value", border=1, align='C')
#         pdf.set_font("Arial", size=9)  # Reset font style to normal
#         pdf.ln(10)  # Move to the next line

#         # Cell height (adjust as needed)
#         cell_height = 10

#         # Add data rows
#         counter = 1  # Counter for serial numbers
#         for each in page_items:
#             val = float(each.on_hand_qty)* float(each.item.price)
#             pdf.cell(9, 10, txt=str(counter), border=1, align='C')  # Add S.No. for each row
#             pdf.cell(50, 10, txt=each.store.name, border=1)
#             pdf.cell(50, 10, txt=each.item.name, border=1)
#             pdf.cell(40, 10, txt=str(each.on_hand_qty), border=1)
#             pdf.cell(30, 10, txt=each.item.item_type.item_category.name, border=1)
#             # Right align price for each data row
#             pdf.cell(20, 10, txt=str(val), align='R', border=1)
#             pdf.ln(10)
#             counter += 1  # Increment counter for next row

#         # Save the PDF file
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)
#         tmpname = "Store_Item_Report.pdf"
#         pdf.output(os.path.join(directory_path, tmpname))
#         os.chmod(os.path.join(directory_path, tmpname), 0o777)
#         # # # # # #print("3439")
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name': tmpname
#         })

def store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity):
    # #print(4221)
    # Fetch the last transaction_date less than the given_date
    given_date = transaction_date

    # Check for the last record on the given_date
    record =  models.Store_Item_Current.objects.filter(transaction_date__gt=given_date,store_id=store_id,item_id=item_id,status=1, deleted=0)

    if record:
        if transact_type == 'min':
            models.Store_Item_Current.objects.filter(
                transaction_date__gt=given_date, 
                store_id=store_id, 
                item_id=item_id,
                status= 1,
                deleted = 0
            ).update(
                opening_qty=F('opening_qty') + Decimal(quantity),
                closing_qty=F('closing_qty') + Decimal(quantity),
                on_hand_qty=F('on_hand_qty') + Decimal(quantity),
                updated_at=now()
            )
        if transact_type == 'mout':
           
            models.Store_Item_Current.objects.filter(
                transaction_date__gt=given_date, 
                store_id=store_id, 
                item_id=item_id,
                status= 1,
                deleted = 0
            ).update(
                opening_qty=F('opening_qty') - Decimal(quantity),
                closing_qty=F('closing_qty') - Decimal(quantity),
                on_hand_qty=F('on_hand_qty') - Decimal(quantity),
                updated_at=now()
            )
  



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def storeTransactionList(request):
    context = {}

    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    vendor_id = request.GET.get('vendor_id', None)
    transaction_type = request.GET.get('transaction_type', None)
    try:
        if id is not None and id != "":
            storeTransaction = list(models.Store_Transaction.objects.filter(status=1, deleted=0).filter(pk=id)[:1].values(
                'pk', 'transaction_number', 'transaction_date', 'total_amount', 'purchase_order_header_id', 'purchase_order_header__order_number', 'vendor__name', 'transaction_type_id', 'transaction_type__name','invoice_challan','job_order_id','job_order__order_number','is_logical_grn_store'))
            context.update({
                'status': 200,
                'message': "Store Transaction Fetched Successfully.",
                'page_items': storeTransaction,
            })
        else:
            if transaction_type:
                storeTransactions = models.Store_Transaction.objects.filter(transaction_type__name=transaction_type).filter(status=1, deleted=0)
            else:
                storeTransactions = models.Store_Transaction.objects.filter(status=1, deleted=0)

            if keyword is not None and keyword != "":
                storeTransactions = list(storeTransactions.filter(Q(vendor__name__icontains=keyword) | Q(transaction_number__icontains=keyword) | Q(
                    transaction_date__icontains=keyword) | Q(total_amount__icontains=keyword) | Q(job_order__manufacturing_material_type__icontains=keyword ) | Q(job_order__order_number__icontains=keyword) | Q(invoice_challan__icontains=keyword)).filter(status=1, deleted=0).values('pk', 'transaction_number', 'transaction_date', 'total_amount',
                                     'purchase_order_header_id', 'purchase_order_header__order_number', 'vendor__name', 'transaction_type_id', 'transaction_type__name', 'invoice_challan','job_order_id','job_order__order_number','is_logical_grn_store'))
                
            else:
                storeTransactions = list(storeTransactions.values('pk', 'transaction_number', 'transaction_date', 'total_amount',
                                     'purchase_order_header_id', 'purchase_order_header__order_number', 'vendor__name', 'transaction_type_id', 'transaction_type__name', 'invoice_challan','job_order_id','job_order__order_number','is_logical_grn_store'))
                       
            if find_all is not None and int(find_all) == 1:
                context.update({
                    'status': 200,
                    'message': "Store Transactions Fetched Successfully.",
                    'page_items': storeTransactions,
                })
                return JsonResponse(context)
            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)
            paginator = CustomPaginator(storeTransactions, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()
            context.update({
                'status': 200,
                'message': "Store Transactions Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })
    except Exception:
        context.update({
            'status': 591.1,
            'message': "internal error",
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeTransactionAdd(request):
    context = {}
    check1 = 0
    test =""
    check2 = 0
    logicalgrnSore = 0
    userId = request.COOKIES.get('userId', None)
    # exit()
    invoice_challan_exist = models.Store_Transaction.objects.filter(invoice_challan=  request.POST['invoice_challan'],status=1, deleted=0 )
    if invoice_challan_exist.exists():
        context.update({
            'status': 586,
            'message': "invoice/challan number already exist",
        })
        return JsonResponse(context)
    if not request.POST['vendor_id'] or not request.POST['transaction_date'] or not request.POST['total_amount']:
        context.update({
            'status': 586,
            'message': "Transaction Type/Vendor/Transaction Date/Total Amount has not been provided."
        })
        return JsonResponse(context)
    message = 'Something Went Wrong. Please Try Again.'

    logical_grn_values = request.POST.getlist('logical_grn')
    if logical_grn_values and ('1' in logical_grn_values) :
        logical_grn_values = request.POST.getlist('logical_grn') 
        storeList = request.POST.getlist('store_id') 
        filtered_store_list = [storeList[i] for i in range(len(storeList)) if logical_grn_values[i] == '1']
        all_same = len(set(filtered_store_list)) == 1
        if all_same :
           logicalgrnSore = 1
           #print(4309)
        else:
            context.update({
                'status': 586,
                'message': "all logical grn vendor store must be same "
            })
            return JsonResponse(context)

    try:

        # #print("3130")
        inspect = request.POST.getlist('inspect')
        # if "1" in inspect:
        #     # # # # # #print("SAswata")
        #     pass
        with transaction.atomic():
            # # # # # #print("3481")

            #-------for job order present with_purchase_job_order ==2 means joborder---------

            #----------------- for vendor transaction ------------
            if (request.POST.get('purchase_job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2):
                
                storeTransactionDetail =[]
                jobOrderHeader = models.Job_Order.objects.get(pk=request.POST['purchase_job_order_header_id'] )
                jobOrderDetails = list(models.Job_Order_Detail.objects.filter(job_order_header_id =request.POST['purchase_job_order_header_id'] ))
                # storeTransactionVHead_count= models.Store_Transaction.objects.all().count()
                storeTransactionVhead= models.Store_Transaction()
                transaction_type = models.Transaction_Type.objects.get(name='MIST')
                storeTransactionVhead.transaction_type = transaction_type
                storeTransactionVhead.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                if request.POST.get('vendor_id',None):
                    storeTransactionVhead.vendor_from_id = request.POST['vendor_id']
                # storeTransactionVhead.transaction_type =models.Transaction_Type.objects.get(name = 'MIST')
                if(int(request.POST['with_purchase_job_order']) == 2):
                    storeTransactionVhead.job_order_id =  request.POST[
                        'purchase_job_order_header_id']
                storeTransactionVhead.transaction_date = request.POST['transaction_date'] 
                storeTransactionVhead.total_amount = request.POST['total_amount']
                storeTransactionVhead.invoice_challan = request.POST['invoice_challan']
                storeTransactionVhead.notes = request.POST['notes']
                storeTransactionVhead.save()
                # #print(4083)
                amount_total = float( request.POST['total_amount'])
                #outgoing material utilised store transaction by vendor
                for index in range(0, len(jobOrderDetails)):
                    
                    incoming_item_quantity = float(request.POST.getlist('item_quantity')[0])
                    incoming_item_id = request.POST.getlist('item_id')[0]
                    if jobOrderDetails[index].direction == 'outgoing' and float(jobOrderDetails[index].quantity_result) != 0:
                        boMHeadDetailsExist = models.Bill_Of_Material_Detail.objects.filter(item_id =jobOrderDetails[index].item_id, bill_of_material_header_id = jobOrderHeader.bom_type_head_id).exists()
                        if boMHeadDetailsExist:
                            bomDetailsFirst = models.Bill_Of_Material_Detail.objects.filter(item_id= jobOrderDetails[index].item_id , bill_of_material_header_id =jobOrderHeader.bom_type_head).first()
                            BomQuantity  = float(bomDetailsFirst.quantity)
                        storeTransactionDetail.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionVhead.id,
                                item_id=jobOrderDetails[index].item_id,
                                store=models.Store.objects.get(vendor_id = request.POST['vendor_id']),
                                quantity=float(jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else (BomQuantity*incoming_item_quantity) ,
                                rate=float(jobOrderDetails[index].item.price),
                                direction = 'outgoing'
                            )    
                        )
                        amount_total += float(jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else (BomQuantity*incoming_item_quantity)
                      
                        # #print(4098)
                        store = models.Store.objects.get(vendor_id=request.POST['vendor_id'])
                        storeItem = models.Store_Item.objects.filter(item_id=jobOrderDetails[index].item_id, store=store).first()
                        storeItem.on_hand_qty -= Decimal(
                                jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else Decimal(BomQuantity*incoming_item_quantity)
                        storeItem.closing_qty -= Decimal(
                            jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else Decimal(BomQuantity*incoming_item_quantity)
                        storeItem.updated_at = datetime.now()
                        storeItem.save()
                        resultant_quantity_result =  models.Job_Order_Detail.objects.filter(item_id = jobOrderDetails[index].item_id,job_order_header_id=request.POST['purchase_job_order_header_id']).first()
                        resultant_quantity_result.quantity_result = 0.0  if not boMHeadDetailsExist else (resultant_quantity_result.quantity_result - Decimal(BomQuantity*incoming_item_quantity))
                        resultant_quantity_result.save()
                        # #print(4423)
                        # change in storeItemCurrent
                        # Fetch the last transaction_date less than the given_date if mout
                        given_date = request.POST['transaction_date']
                        
                        
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=jobOrderDetails[index].item_id, store=store ,status=1, deleted=0).last()
                        #print(f'record: {record},given date:{given_date}')
                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date ,item_id=jobOrderDetails[index].item_id, store_id=store.id,status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']
                            

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date ,item_id=jobOrderDetails[index].item_id, store_id=store.id, status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                       
                        
                        store_item_instance = models.Store_Item_Current()
                        if record:
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty - (Decimal(
                                jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else Decimal(BomQuantity*incoming_item_quantity))
                            store_item_instance.closing_qty = record.closing_qty - (Decimal(
                                jobOrderDetails[index].quantity_result) if not boMHeadDetailsExist else Decimal(BomQuantity*incoming_item_quantity))
                            # Set other fields for the new transaction
                            store_item_instance.store_transaction_id = storeTransactionVhead.id
                            if(store_item_instance.on_hand_qty<0):
                                raise ValueError(f"out quantity is more than available quantity {store.name} {jobOrderDetails[index].item.name}")
                            store_item_instance.transaction_date = given_date
                            store_item_instance.item_id = jobOrderDetails[index].item_id
                            store_item_instance.store_id = store.id
                            
                            # Save the instance to the database
                            store_item_instance.save()
                        else:
                            message = f"canot possible item not present in the store{store.name} "
                            raise ValueError(message)
                        moutQuantity = (jobOrderDetails[index].quantity_result if not boMHeadDetailsExist else Decimal(BomQuantity*incoming_item_quantity)) 
                        store_item_curreEdit(store.id,jobOrderDetails[index].item_id,given_date,'mout',moutQuantity) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                        # #print(resultant_quantity_result.quantity_result , jobOrderDetails[index].item.name)
                if storeTransactionDetail:
                    models.Store_Transaction_Detail.objects.bulk_create(storeTransactionDetail)
                storeTransactionVhead.total_amount = amount_total
                storeTransactionVhead.save()      
            
            storeTransactionDetail =[]
            
            if "1" in inspect:
                # #print("313s4")
                grn_inspection_transaction_count = models.Grn_Inspection_Transaction.objects.all().count()
                grnTransactionheader = models.Grn_Inspection_Transaction()
                grnTransactionheader.vendor_id = request.POST['vendor_id']
                grnTransactionheader.transaction_type = models.Transaction_Type.objects.get(name = 'GRNI')
                grnTransactionheader.invoice_challan = request.POST['invoice_challan']
                grnTransactionheader.transaction_number = env("GRN_TRANSACTION_INSPECTION_SEQ").replace(
                    "${CURRENT_YEAR}", datetime.today().strftime('%Y')).replace("${AI_DIGIT_5}", str(grn_inspection_transaction_count + 1).zfill(5))
                # # # # # #print("3143")
                if (request.POST.get('purchase_job_order_header_id',None) and int(request.POST['with_purchase_job_order']) != 2):
                    grnTransactionheader.purchase_order_header_id = request.POST[
                        'purchase_job_order_header_id']
                # # # # # #print("3147")
                if(int(request.POST['with_purchase_job_order']) == 2):
                    grnTransactionheader.job_order_id =  request.POST[
                        'purchase_job_order_header_id']
                grnTransactionheader.transaction_date = request.POST['transaction_date']
                # # # # # #print("3149")
                grnTransactionheader.total_amount = request.POST['total_amount']
                # # # # # #print("3151")
                grnTransactionheader.notes = request.POST['notes']

                grnTransactionheader.is_logical_grn_store = logicalgrnSore
                # # # # # #print("3153")
                grnTransactionheader.save()
                # # # # # #print("3148")
               
                order_details = []
                total_amounts = 0 
                material_reciept_all = 0
                for index, elem in enumerate(request.POST.getlist('item_id')):
                    # #print('3605')
                    if inspect[index] == "1":
                        check1 +=1
                        # # # # # #print( request.POST.getlist(
                        #             'amount_with_gst')[index])
                       
                        
                        total_amounts = float(request.POST.getlist(
                                    'amount_with_gst')[index])
                        order_details.append(
                            models.Grn_Inspection_Transaction_Detail(
                                grn_inspection_transaction_header_id= grnTransactionheader.id,
                                item_id=elem,
                                store_id=request.POST.getlist('store_id')[index],
                                quantity=request.POST.getlist('item_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('item_price')[index],
                                gst_percentage=request.POST.getlist(
                                    'gst_percentage')[index],
                                amount_with_gst=request.POST.getlist(
                                    'amount_with_gst')[index],
                                logical_grn_store = request.POST.getlist(
                                    'logical_grn')[index]
                                
                            )
                        )
                        #incoming material exported by vendor to comapany ---- for grn
                        if (request.POST.get('purchase_job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2):
                            # # # # # #print('3627')
                            job_order_details= models.Job_Order_Detail.objects.filter(item_id=elem, 
                                                job_order_header_id= request.POST['purchase_job_order_header_id'],direction='incoming').get()
                            # # # # # #print(job_order_details.quantity_result - Decimal(request.POST.getlist('item_quantity')[index]))
                            job_order_details.quantity_result -= Decimal(request.POST.getlist('item_quantity')[index]) 
                            # # # # # #print('3632')
                            job_order_details.updated_at = datetime.now()
                            job_order_details.save()
                            storeTransactionDetail.append(
                                models.Store_Transaction_Detail(
                                    store_transaction_header_id=storeTransactionVhead.id,
                                    item_id=elem,
                                    store=models.Store.objects.get(vendor_id = request.POST['vendor_id']),
                                    quantity=float(request.POST.getlist('item_quantity')[index]),
                                    rate=float(request.POST.getlist('rate')[index]),
                                    direction = 'incoming'
                                )    
                            )
                            store = models.Store.objects.get(vendor_id=request.POST['vendor_id'])
                            storeItem = models.Store_Item.objects.filter(item_id = elem , store = store).first()
                            storeItem.on_hand_qty -= Decimal(request.POST.getlist('item_quantity')[index])
                            storeItem.closing_qty -= Decimal(request.POST.getlist('item_quantity')[index])
                            storeItem.updated_at = datetime.now()
                            storeItem.save()
                            given_date = request.POST['transaction_date']
                        
                        
                            # Check for the last record on the given_date
                            record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id = elem , store_id = store.id,status=1, deleted=0 ).last()

                            if not record:
                                # If no record is found for the given_date, look for the last record before that date
                                last_transaction_date = models.Store_Item_Current.objects.filter(
                                    transaction_date__lt=given_date ,item_id = elem , store_id = store.id,status=1, deleted=0
                                ).aggregate(Max('transaction_date'))['transaction_date__max']

                                

                                if last_transaction_date:
                                    # Fetch the record for the last_transaction_date
                                    record = models.Store_Item_Current.objects.filter(
                                        transaction_date=last_transaction_date ,item_id = elem , store_id = store.id,status=1, deleted=0
                                    ).last()

                            # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                            
                            store_item_instance = models.Store_Item_Current()
                            if record:
                                # Set values based on the last record found
                                store_item_instance.opening_qty = record.closing_qty
                                store_item_instance.on_hand_qty = record.closing_qty - (Decimal(request.POST.getlist('item_quantity')[index]))
                                store_item_instance.closing_qty = record.closing_qty - (Decimal(request.POST.getlist('item_quantity')[index]))
                                # Set other fields for the new transaction
                                store_item_instance.store_transaction_id = storeTransactionVhead.id
                                if(store_item_instance.on_hand_qty<0):
                                    raise ValueError(f"out quantity is more than available quantity {store.name} , {elem}")
                                store_item_instance.transaction_date = given_date
                                store_item_instance.item_id = elem
                                store_item_instance.store_id = store.id
                                
                                # Save the instance to the database
                                store_item_instance.save()
                            else:
                                message = "canot possible item not present in the store "
                                raise ValueError(message)
                            moutQuantity =(Decimal(request.POST.getlist('item_quantity')[index]))
                            store_item_curreEdit(store.id,elem,given_date,'mout',moutQuantity) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                            # #print(resultant_quantity_result.quantity_result , jobOrderDetails[index].item.name)

                            material_reciept_all = 0 if float(job_order_details.quantity_result)>0.00 else 1
                        #     # # # # #print('3634') 

                            #closed job order
                            if (material_reciept_all == 1):
                                job_order = models.Job_Order.objects.filter(pk=request.POST[
                                    'purchase_job_order_header_id']).get() 
                                # # # # # #print(job_order)
                                job_order.material_reciept = 1
                                job_order.job_status = 2
                                current_time = datetime.now(timezone.utc)
                                time_difference = current_time - job_order.created_at
                                total_hours = time_difference.days * 24 + time_difference.seconds / 3600
                                job_order.actual_time_take = str(round(total_hours, 3)) + 'hr'
                                job_order.updated_at = datetime.now()   
                                job_order.save()
                        # # # # # #print("3170")
                if storeTransactionDetail:
                    models.Store_Transaction_Detail.objects.bulk_create(storeTransactionDetail)
                models.Grn_Inspection_Transaction_Detail.objects.bulk_create(order_details)
                grnTransactionheader.total_amount = total_amounts
                grnTransactionheader.save()
                # # # # # #print("3166")
                

            if "0" in inspect:
                # # #print(4214)
                store_transaction_count = models.Store_Transaction.objects.all().count()
                storeTransactionHeader = models.Store_Transaction()
                storeTransactionHeader.vendor_id = request.POST['vendor_id']
                storeTransactionHeader.creator_id = userId
                # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'GRN')
                transaction_type = models.Transaction_Type.objects.get(name='GRN')
                storeTransactionHeader.transaction_type = transaction_type
                storeTransactionHeader.invoice_challan = request.POST['invoice_challan']
                # # # # # #print("3182")
                if (request.POST.get('purchase_job_order_header_id',None) and int(request.POST['with_purchase_job_order']) != 2):
                    storeTransactionHeader.purchase_order_header_id = request.POST[
                        'purchase_job_order_header_id']
                # # # # #print("3186")
                if(int(request.POST['with_purchase_job_order']) == 2):
                    storeTransactionHeader.job_order_id =  request.POST[
                        'purchase_job_order_header_id']
               
                storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                storeTransactionHeader.transaction_date = request.POST['transaction_date']
                storeTransactionHeader.total_amount = request.POST['total_amount']
                storeTransactionHeader.notes = request.POST['notes']
                storeTransactionHeader.is_logical_grn_store = logicalgrnSore
                storeTransactionHeader.save()

                # # #print('3549')
                order_details = []
                total_amounts = 0 
                material_reciept_all = 0
                for index, elem in enumerate(request.POST.getlist('item_id')):
                    if inspect[index] == "0":
                        check2 +=1
                        order_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionHeader.id,
                                item_id=elem,
                                store_id=request.POST.getlist('store_id')[index],
                                quantity=request.POST.getlist('item_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('item_price')[index],
                                gst_percentage=request.POST.getlist(
                                    'gst_percentage')[index],
                                amount_with_gst=request.POST.getlist(
                                    'amount_with_gst')[index],
                                logical_grn_store = request.POST.getlist(
                                    'logical_grn')[index],
                            )
                        )
                        total_amounts += float(request.POST.getlist(
                                    'amount_with_gst')[index])
                        # # # # # #print('3569')
                        storeItem = models.Store_Item.objects.filter(
                            item_id=elem, store_id=request.POST.getlist('store_id')[index]).first()
                        # # #print(storeItem)
                        if storeItem is None:
                            # # #print(4484)
                            storeItem = models.Store_Item()
                            storeItem.opening_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.on_hand_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.closing_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.item_id = elem
                            storeItem.store_id = request.POST.getlist('store_id')[
                                index]
                            storeItem.save()
                            # # #print(4496)
                        else:
                            # # #print(4498)
                            storeItem.on_hand_qty += Decimal(
                                request.POST.getlist('item_quantity')[index])
                            # # #print(4501)
                            storeItem.closing_qty += Decimal(
                                request.POST.getlist('item_quantity')[index])  
                            storeItem.updated_at = datetime.now()
                            # # #print(storeItem.closing_qty)
                            storeItem.save()

                        # change in storeItemCurrent
                        # Fetch the last transaction_date less than the given_date
                        given_date = request.POST['transaction_date']   
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=elem, store_id=request.POST.getlist('store_id')[index],status=1, deleted=0).last()
                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date,item_id=elem, store_id=request.POST.getlist('store_id')[index],status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date,item_id=elem, store_id=request.POST.getlist('store_id')[index], status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                        store_item_instance = models.Store_Item_Current()
                        if record:
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty + Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                            store_item_instance.closing_qty = record.closing_qty + Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                        else:
                            # Set values based on the current transaction if no prior record exists
                            store_item_instance.opening_qty = Decimal(0.00
                            )
                            store_item_instance.on_hand_qty = Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                            store_item_instance.closing_qty = Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                        if(store_item_instance.on_hand_qty<0):
                            raise ValueError(f"onhand  quantity can not be negative {request.POST.getlist('store_id')[index]} {elem}")
                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionHeader.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = elem
                        store_item_instance.store_id = request.POST.getlist('store_id')[index]
                        # Save the instance to the database
                        store_item_instance.save()
                        store_item_curreEdit(request.POST.getlist('store_id')[index],elem,given_date,'min', request.POST.getlist('item_quantity')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

                        # incoming material exported by vendor to company
                        if (request.POST.get('purchase_job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2): #it is a job order
                            # # # # # #print('3641')
                            job_order_details= models.Job_Order_Detail.objects.filter(item_id=elem, 
                                                job_order_header_id= request.POST['purchase_job_order_header_id'],direction='incoming').get()
                            # # # # # #print(request.POST.getlist('item_quantity')[index])
                            job_order_details.quantity_result -=  Decimal(request.POST.getlist('item_quantity')[index]) 
                            # # # # # #print('3646')
                            job_order_details.updated_at = datetime.now()
                            job_order_details.save()
                            material_reciept_all = 0 if float(job_order_details.quantity_result)>0.00 else 1 
                            # # # # #print("3170")
                            storeTransactionDetail.append(
                                models.Store_Transaction_Detail(
                                    store_transaction_header_id=storeTransactionVhead.id,
                                    item_id=elem,
                                    store=models.Store.objects.get(vendor_id = request.POST['vendor_id']),
                                    quantity=float(request.POST.getlist('item_quantity')[index]),
                                    rate=float(request.POST.getlist('rate')[index]),
                                    direction = 'incoming'
                                )    
                            )
                            store = models.Store.objects.get(vendor_id=request.POST['vendor_id'])
                            # # # # #print(store.id, request.POST.getlist('item_id')[index])
                            if models.Store_Item.objects.filter(item_id = request.POST.getlist('item_id')[index] , store_id = store.id).exists():
                                storeItem = models.Store_Item.objects.filter(item_id = elem , store = store).first()
                                # # # # #print( models.Store_Item.objects.filter(item_id = request.POST.getlist('item_id')[index] , store_id = store.id).exists())
                                storeItem.on_hand_qty -= Decimal(
                                request.POST.getlist('item_quantity')[index])
                                storeItem.closing_qty -= Decimal(
                                request.POST.getlist('item_quantity')[index])
                                storeItem.updated_at = datetime.now()
                                storeItem.save()
                            
                            # change in storeItemCurrent mout
                            # Fetch the last transaction_date less than the given_date if mout
                            given_date = request.POST['transaction_date']
                            
                            
                            # Check for the last record on the given_date
                            record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=elem, store_id=store.id,status=1, deleted=0 ).last()

                            if not record:
                                # If no record is found for the given_date, look for the last record before that date
                                last_transaction_date = models.Store_Item_Current.objects.filter(
                                    transaction_date__lt=given_date ,item_id=elem, store=store.id,status=1, deleted=0 
                                ).aggregate(Max('transaction_date'))['transaction_date__max']

                                

                                if last_transaction_date:
                                    # Fetch the record for the last_transaction_date
                                    record = models.Store_Item_Current.objects.filter(
                                        transaction_date=last_transaction_date ,item_id=elem, store_id=store.id,status=1, deleted=0
                                    ).last()

                            # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                            
                            store_item_instance = models.Store_Item_Current()

                            if record:
                                # Set values based on the last record found
                                store_item_instance.opening_qty = record.closing_qty
                                store_item_instance.on_hand_qty = record.closing_qty - Decimal(request.POST.getlist('item_quantity')[index])
                                store_item_instance.closing_qty = record.closing_qty - Decimal(request.POST.getlist('item_quantity')[index])
                            
                                if(store_item_instance.on_hand_qty<0):
                                    raise ValueError(f"out quantity is more than available quantity {store.name} , {elem}")
                                # Set other fields for the new transaction
                                store_item_instance.store_transaction_id = storeTransactionVhead.id
                                store_item_instance.transaction_date = given_date
                                store_item_instance.item_id = elem #item_id=request.POST.getlist('item_id')[index], store=store.id 
                                store_item_instance.store_id = store.id 

                                # Save the instance to the database
                                store_item_instance.save()
                            
                            else:
                                message = "canot possible manufacture item is missing on vendor store how will you deducted"
                                raise ValueError(message)
                            store_item_curreEdit(store.id,elem,given_date,'mout', request.POST.getlist('item_quantity')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                            #closed job order
                            if (material_reciept_all == 1):
                                job_order = models.Job_Order.objects.filter(pk=request.POST[
                                'purchase_job_order_header_id']).get() 
                                    # # # # # #print(job_order)
                                job_order.material_reciept = 1
                                job_order.job_status = 2
                                current_time = datetime.now(timezone.utc)
                                time_difference = current_time - job_order.created_at
                                total_hours = time_difference.days * 24 + time_difference.seconds / 3600
                                job_order.actual_time_take = str(round(total_hours, 3)) + 'hr'
                                job_order.updated_at = datetime.now()   
                                job_order.save()
                # #print(4548)
                if storeTransactionDetail:
                    models.Store_Transaction_Detail.objects.bulk_create(storeTransactionDetail)
                models.Store_Transaction_Detail.objects.bulk_create(order_details)
                storeTransactionHeader.total_amount = total_amounts
                storeTransactionHeader.save()
                # # # # #print(order_details)
                
                # # # # # #print('3589')
                if request.POST['with_purchase_job_order'] != "" and int(request.POST['with_purchase_job_order']) != 0 and int(request.POST['with_purchase_job_order']) != 2 :
                    for index, elem in enumerate(request.POST.getlist('detail_id')):
                        purchaseOrderItem = models.Purchase_Order_Detail.objects.get(
                            pk=elem)
                       
                        purchaseOrderItem.delivered_quantity += Decimal(
                            request.POST.getlist('item_quantity')[index])
                        purchaseOrderItem.delivered_rate = Decimal(
                            request.POST.getlist('rate')[index])
                        purchaseOrderItem.delivered_amount += Decimal(
                            request.POST.getlist('item_price')[index])
                        purchaseOrderItem.delivered_gst_percentage = Decimal(
                            request.POST.getlist('gst_percentage')[index])
                        purchaseOrderItem.delivered_amount_with_gst += Decimal(
                            request.POST.getlist('amount_with_gst')[index])
                        purchaseOrderItem.updated_at = datetime.now()
                        purchaseOrderItem.save()
                    # # # # # #print(request.POST['purchase_job_order_header_id'])
                    purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related(
                        'purchase_order_detail_set').get(pk=request.POST['purchase_job_order_header_id'])
                    
                    flag = True
                   
                    for purchaseOrderDetail in purchaseOrderHeader.purchase_order_detail_set.all():
                        if Decimal(purchaseOrderDetail.quantity) > Decimal(purchaseOrderDetail.delivered_quantity):
                            flag = False
                            break
                    if flag == True:
                        purchaseOrderHeader.delivery_status = 3
                    else:
                        purchaseOrderHeader.delivery_status = 2
                    purchaseOrderHeader.updated_at = datetime.now()
                    purchaseOrderHeader.save()
                # # # # # #print('4348') 
            
            user_log_details_add(userId,'Store Transaction Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Transaction Created Successfully."
        })
    except Exception as e:
        #print(f"error is {e}")
        # tb = traceback.format_exc() 
        # exc_type, exc_value, exc_tb = e.__traceback__.tb_frame.f_globals['__builtins__']["sys"].exc_info()
        # line_number = exc_tb.tb_lineno
        # #print(f"Error occurred on line {line_number}")
        context.update({
            'status': 588,
            'message': message
        })
        transaction.rollback()
    return JsonResponse(context)


def data_revertive_from_transaction(storeTrId, itemId, storeId, quantity, retrivaltype):
    #print('Inside data_revertive_from_transaction')
    try:
        # Subqueries to fetch transaction_date and created_at for the specific transaction
        subquery_transaction_date = models.Store_Item_Current.objects.filter(
            store_transaction_id=storeTrId, status=1, deleted=0
        ).values('transaction_date')[:1]

        subquery_created_at = models.Store_Item_Current.objects.filter(
            store_transaction_id=storeTrId, status=1, deleted=0
        ).values('created_at')[:1]


        # Determine the adjustment factor based on retrieval type
        adjustment = Decimal(quantity) if retrivaltype == 'in' else -(Decimal(quantity))

        # Filter and update matching records in bulk
        updated_count = models.Store_Item_Current.objects.filter(
            item_id=itemId,
            store_id=storeId,
            status=1,
            deleted=0,
        ).filter(
            Q(
                Q(
                    transaction_date=Subquery(subquery_transaction_date),
                    created_at__gte=Subquery(subquery_created_at)
                ) |
                Q(transaction_date__gt=Subquery(subquery_transaction_date))
            )
        ).update(
            closing_qty=F('closing_qty') + adjustment,
            on_hand_qty=F('on_hand_qty') + adjustment,
            opening_qty=F('opening_qty') + adjustment,
            updated_at=datetime.now()
        )

        # Log the result of the update
        if updated_count == 0:
            raise ValueError('No data found for the given criteria.')

    except Exception as e:
        raise ValueError(f"An error occurred: {e}")

def jobOrderStoreTranasctionRetriveVendor(storeTranscationOld):
    #data deducted from vendor store
    invoice_no = storeTranscationOld.invoice_challan
    store_transaction_MisT_head = models.Store_Transaction.objects.filter(invoice_challan=invoice_no ,transaction_type__name = 'MIST',deleted= 0, status=1)
    
    # Check if any records exist
    if store_transaction_MisT_head.exists():
        store_transaction_MisT_head = store_transaction_MisT_head.first()
        store_transaction_MisT_head.status = 0 
        store_transaction_MisT_head.invoice_challan =''
        store_transaction_MisT_head.updated_at = datetime.now()
        store_transaction_MisT_head.save()
        store_transaction_details = models.Store_Transaction_Detail.objects.filter(store_transaction_header_id = store_transaction_MisT_head.id)
        for store_transaction_detail in store_transaction_details:
            data_revertive_from_transaction(store_transaction_MisT_head.id, store_transaction_detail.item_id,store_transaction_detail.store_id,store_transaction_detail.quantity,'in')
            store_items = models.Store_Item.objects.filter(item_id = store_transaction_detail.item_id, store_id = store_transaction_detail.store_id).first() 
            store_items.closing_qty += Decimal(store_transaction_detail.quantity)
            store_items.on_hand_qty += Decimal(store_transaction_detail.quantity)
            store_items.updated_at = datetime.now()
            store_items.save()
            store_item_current = models.Store_Item_Current.objects.filter(store_transaction_id = store_transaction_MisT_head.id,status=1, deleted=0)
           
            if store_item_current.exists():
                store_item_current = store_item_current.first()
                store_item_current.on_hand_qty += Decimal(store_transaction_detail.quantity)
                store_item_current.closing_qty += Decimal(store_transaction_detail.quantity)
                store_item_current.status = 0
                store_item_current.updated_at = datetime.now()
                store_item_current.save()
        
            joborderDet = models.Job_Order_Detail.objects.filter(job_order_header_id = storeTranscationOld.job_order.id, item_id = store_transaction_detail.item_id)
            if joborderDet.exists():
                joborderDet = joborderDet.first()
                joborderDet.quantity_result += Decimal(store_transaction_detail.quantity)
                joborderDet.updated_at = datetime.now()
                joborderDet.save()

    else:
        raise ValueError('No transaction present in vendor store')

def jobOrderStoreTranasctionRetriveInHouse(storeTranscationOld):
    #print('inside jobOrderStoreTranasctionRetriveInHouse')
    storeTransactionDets = models.Store_Transaction_Detail.objects.filter(store_transaction_header_id = storeTranscationOld.id)

    for storeTransactionDet in storeTransactionDets:
        data_revertive_from_transaction(storeTranscationOld.id,
                                        storeTransactionDet.item_id,
                                        storeTransactionDet.store_id,
                                        storeTransactionDet.quantity,'out')
        store_items = models.Store_Item.objects.filter(item_id = storeTransactionDet.item_id, store_id = storeTransactionDet.store_id).first() 
        store_items.closing_qty -= Decimal(storeTransactionDet.quantity)
        store_items.on_hand_qty -= Decimal(storeTransactionDet.quantity)
        store_items.updated_at = datetime.now()
        store_items.save()
        store_item_current = models.Store_Item_Current.objects.filter(store_transaction_id = storeTranscationOld.id,status=1, deleted=0)
        if store_item_current.exists():
            store_item_current = store_item_current.first()
            store_item_current.on_hand_qty -= Decimal(storeTransactionDet.quantity)
            store_item_current.closing_qty -= Decimal(storeTransactionDet.quantity)
            store_item_current.status = 0
            store_item_current.updated_at = datetime.now()
            store_item_current.save()


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeTransactionEdit(request):
    context = {}
    # # #print(request.POST)
    check1 = 0
    test =""
    check2 = 0
    userId = request.COOKIES.get('userId', None)
    inspect = 1 if '1' in request.POST.getlist('itemInspect') else 0

    inspectZero = 1 if '0' in  request.POST.getlist('itemInspect') else 0
    #print(models.Store_Transaction.objects.filter(invoice_challan=request.POST['invoice_challan']))
    invoice_challan_exist = models.Store_Transaction.objects.filter(invoice_challan=request.POST['invoice_challan'],status=1, deleted=0 ,transaction_type__name='GRN').exclude(pk = request.POST['id'])
    if invoice_challan_exist.exists():
        context.update({
            'status': 591.1,
            'message': "invoice/challan number already exist",
        })
        return JsonResponse(context)
    # logical_grn_values = request.POST.getlist('logical_grn')
    # if logical_grn_values and ('1' in logical_grn_values) :
    #     logical_grn_values = request.POST.getlist('logical_grn') 
    #     storeList = request.POST.getlist('store_id') 
    #     filtered_store_list = [storeList[i] for i in range(len(storeList)) if logical_grn_values[i] == '1']
    #     all_same = len(set(filtered_store_list)) == 1
    #     if all_same :
    #        logicalgrnSore = 1
    #        #print(4309)
    #     else:
    #         context.update({
    #             'status': 586,
    #             'message': "all logical grn vendor store must be same "
    #         })
    #         return JsonResponse(context)
         
    # #print(inspect,inspectZero)
    try:
        with transaction.atomic():
            # #print(4645)
            storeTranasctionHeaderOld = models.Store_Transaction.objects.get(pk = request.POST['id'], status=1, deleted=0)
            storeTranasctionDetOld = list(models.Store_Transaction_Detail.objects.filter(store_transaction_header_id = request.POST['id']))

            #setting new transaction number from old transaction number
            transaction_no = storeTranasctionHeaderOld.transaction_number
            new_transaction_no = re.sub(r"TR(\d*)_", lambda m: f"TR{int(m.group(1) or 0) + 1}_", transaction_no)
            # new_transaction_no=str(storeTranasctionHeaderOld.transaction_number+1).zfill(5)
            # closing this old transaction remove from original list
            storeTranasctionHeaderOld.status = 0 
            storeTranasctionHeaderOld.updated_at = datetime.now()
            storeTranasctionHeaderOld.save()
            # #print(4503)
            #retriving orginal data to in house store
            if int(request.POST['with_purchase_job_order']) != 2: #it is a purchase(with/without) 
                #print(5305)
                for transact in storeTranasctionDetOld:
                    storeItemExists = models.Store_Item.objects.filter(store_id=transact.store_id,item_id = transact.item_id).exists()
                    if storeItemExists:
                        storeItem = models.Store_Item.objects.get(store_id=transact.store_id,item_id = transact.item_id)
                        storeItem.on_hand_qty -= transact.quantity
                        storeItem.closing_qty -= transact.quantity
                        storeItem.updated_at = datetime.now()
                        storeItem.save()
                    data_revertive_from_transaction(storeTranasctionHeaderOld.id,transact.item_id,transact.store_id,transact.quantity,'out')
                    storeItemCurrent = models.Store_Item_Current.objects.filter(store_transaction_id = storeTranasctionHeaderOld.id ,transaction_date = storeTranasctionHeaderOld.transaction_date,store_id=transact.store_id,item_id = transact.item_id , status=1, deleted=0)
                    if storeItemCurrent.exists():
                        storeItemCurrent = storeItemCurrent.first()
                        storeItemCurrent.on_hand_qty -= Decimal(transact.quantity)
                        storeItemCurrent.closing_qty -= Decimal(transact.quantity)
                        storeItemCurrent.status = 0
                        storeItemCurrent.updated_at = datetime.now()
                        storeItemCurrent.save()
                    
                    storecurrent = models.Store_Item_Current.objects.filter(store_id=transact.store_id,item_id = transact.item_id , status=1, deleted=0).order_by('transaction_date','created_at')
                    if storecurrent.exists():
                        storecurrent = storecurrent.last()
                        #print(storecurrent.closing_qty,storecurrent.on_hand_qty,storecurrent.transaction_date)
                    
                       

                    if int(request.POST['with_purchase_job_order']) == 1: #with purchase order
                        purchaseOrderDetExist = models.Purchase_Order_Detail.objects.filter(purchase_order_header_id =request.POST['purchase_job_order_header_id'],item_id = transact.item_id).exists()
                        if purchaseOrderDetExist:
                            purchaseOrderDet = models.Purchase_Order_Detail.objects.get(purchase_order_header_id =request.POST['purchase_job_order_header_id'],item_id = transact.item_id)
                            purchaseOrderDet.delivered_quantity -= transact.quantity
                            purchaseOrderDet.delivered_amount -= transact.amount
                            purchaseOrderDet.delivered_amount_with_gst  -= transact.amount_with_gst
                            purchaseOrderDet.updated_at = datetime.now() 
                            purchaseOrderDet.save()
                    
                #status changing of purchase order
                if int(request.POST['with_purchase_job_order']) == 1: #with purchase order
                    purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related(
                                'purchase_order_detail_set').get(pk=request.POST['purchase_job_order_header_id']) 
                    
                    flag = True
                        
                    for purchaseOrderDetail in purchaseOrderHeader.purchase_order_detail_set.all():
                        if Decimal(purchaseOrderDetail.quantity) > Decimal(purchaseOrderDetail.delivered_quantity):
                            flag = False
                            break    
                    if flag == True:
                        purchaseOrderHeader.delivery_status = 3
                    else:
                        purchaseOrderHeader.delivery_status = 2
                        purchaseOrderHeader.updated_at = datetime.now()
                        purchaseOrderHeader.save()
            else:
                
                jobOrderStoreTranasctionRetriveVendor(storeTranasctionHeaderOld)
                jobOrderStoreTranasctionRetriveInHouse(storeTranasctionHeaderOld)
               
                
                jobOrderHeader = models.Job_Order.objects.get(pk = storeTranasctionHeaderOld.job_order.id)
                all_material_recieved =  models.Job_Order_Detail.objects.filter(
                                            job_order_header_id=storeTranasctionHeaderOld.job_order.id,
                                            direction='incoming',
                                            quantity_result=0.00
                                        ).exists()

                # all material not  recieved
                if not all_material_recieved:
                    jobOrderHeader.material_reciept = 0
                    jobOrderHeader.actual_time_take = ''
                    jobOrderHeader.job_status = 1
                    jobOrderHeader.updated_at = datetime.now()
                    jobOrderHeader.save()

            # ---end retruving old data----

            # adding actual data 

            storeTranasctionHeaderOld.invoice_challan = ''
            storeTranasctionHeaderOld.save()
            # #print(4503)

            # if some data are in inspect
            if inspect == 1 : 

                # grn header
                grn_inspection_transaction_count = models.Grn_Inspection_Transaction.objects.all().count()
                grnTransactionheader = models.Grn_Inspection_Transaction()
                grnTransactionheader.vendor_id = request.POST['vendor_id']
                grnTransactionheader.transaction_type = models.Transaction_Type.objects.get(name = 'GRNI')
                grnTransactionheader.invoice_challan = request.POST['invoice_challan']
                grnTransactionheader.old_store_transaction_id = int(request.POST['id'])
                grnTransactionheader.transaction_number = env("GRN_TRANSACTION_INSPECTION_SEQ").replace(
                    "${CURRENT_YEAR}", datetime.today().strftime('%Y')).replace("${AI_DIGIT_5}", str(grn_inspection_transaction_count + 1).zfill(5))
                if (request.POST.get('job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2): # it is job order reciept
                    grnTransactionheader.job_order_id =  request.POST[
                        'job_order_header_id']
                if(request.POST.get('purchase_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 1): # it is purchase order reciept
                    grnTransactionheader.purchase_order_header_id = request.POST['purchase_order_header_id']
                grnTransactionheader.transaction_date = request.POST['transaction_date']
                grnTransactionheader.total_amount = request.POST['total_amount']
                grnTransactionheader.notes = request.POST['notes']
                grnTransactionheader.save()
                
                order_details = []
                total_amounts = 0 
                material_reciept_all = 0
                
                # Grn details
                for index, elem in enumerate(request.POST.getlist('item_id')):
                    # # # # # #print('3605
                    if int(request.POST.getlist('itemInspect')[index]) == 1:
                        check1 +=1
                        total_amounts = float(request.POST.getlist(
                                    'amount_with_gst')[index])
                        order_details.append(
                            models.Grn_Inspection_Transaction_Detail(
                                grn_inspection_transaction_header_id= grnTransactionheader.id,
                                item_id=elem,
                                store_id=request.POST.getlist('store_id')[index],
                                quantity=request.POST.getlist('item_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('item_price')[index],
                                gst_percentage=request.POST.getlist(
                                    'gst_percentage')[index],
                                amount_with_gst=request.POST.getlist(
                                    'amount_with_gst')[index]
                            )
                        )
                models.Grn_Inspection_Transaction_Detail.objects.bulk_create(order_details)
                grnTransactionheader.total_amount = total_amounts
                grnTransactionheader.save()

            # if some data not in inspect
            if inspectZero == 1 :
                
                
                # storeTransaction Header
                storeTransactionHeader = models.Store_Transaction()
                #print(5429)
                if (request.POST.get('vendor_id',None)):
                    storeTransactionHeader.vendor_id = request.POST['vendor_id']
                storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'GRN')
                storeTransactionHeader.invoice_challan = request.POST['invoice_challan']
                storeTransactionHeader.creator_id = userId
                # # # # # #print("3182")
                if(request.POST.get('purchase_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 1): # it is purchase order reciept
                    storeTransactionHeader.purchase_order_header_id = request.POST[
                        'purchase_order_header_id']
                # # # # #print("3186")
                if (request.POST.get('job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2): # it is job order reciept
                    storeTransactionHeader.job_order_id =  request.POST[
                        'job_order_header_id']
                storeTransactionHeader.transaction_number = new_transaction_no
                storeTransactionHeader.transaction_date = request.POST['transaction_date']
                storeTransactionHeader.total_amount = request.POST['total_amount']
                storeTransactionHeader.notes = request.POST['notes']
                storeTransactionHeader.save()
                
                # # # # #print('3549')
                order_details = []
                total_amounts = 0.00 
                material_reciept_all = 0

                # Store transaction details
                for index, elem in enumerate(request.POST.getlist('item_id')):

                    if int(request.POST.getlist('itemInspect')[index]) == 0:
                        order_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionHeader.id,
                                item_id=elem,
                                store_id=request.POST.getlist('store_id')[index],
                                quantity=request.POST.getlist('item_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('item_price')[index],
                                gst_percentage=request.POST.getlist(
                                    'gst_percentage')[index],
                                amount_with_gst=request.POST.getlist(
                                    'amount_with_gst')[index]
                            )
                        )
                        total_amounts += float(request.POST.getlist(
                                    'amount_with_gst')[index])
                        # #print('3569')

                        #adding data to my store
                        storeItem = models.Store_Item.objects.filter(
                            item_id=elem, store_id=request.POST.getlist('store_id')[index]).first()
                        if storeItem is None:
                            storeItem = models.Store_Item()
                            storeItem.opening_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.on_hand_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.closing_qty = Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.item_id = elem
                            storeItem.store_id = request.POST.getlist('store_id')[
                                index]
                            storeItem.save()
                        else:
                            storeItem.on_hand_qty += Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.closing_qty += Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.updated_at = datetime.now()
                            storeItem.save()
                        # change in storeItemCurrent min
                        # Fetch the last transaction_date less than the given_date
                        given_date = request.POST['transaction_date']
                        # #print(5564)
                        
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date, item_id=elem, store_id=request.POST.getlist('store_id')[index], status=1, deleted=0).last()

                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date, item_id=elem, store_id=request.POST.getlist('store_id')[index],status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date, item_id=elem, store_id=request.POST.getlist('store_id')[index],status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                        
                        store_item_instance = models.Store_Item_Current()

                        if record:
                           
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty + Decimal(
                               request.POST.getlist('item_quantity')[index]
                            )
                            store_item_instance.closing_qty = record.closing_qty + Decimal(
                               request.POST.getlist('item_quantity')[index]
                            )
                        else:
                            # Set values based on the current transaction if no prior record exists
                            store_item_instance.opening_qty = Decimal(
                                0.00
                            )
                            store_item_instance.on_hand_qty = Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                            store_item_instance.closing_qty = Decimal(
                                request.POST.getlist('item_quantity')[index]
                            )
                        if(store_item_instance.on_hand_qty<0):
                            raise ValueError(f"onhand quantity can not be negative")
                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionHeader.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = elem
                        store_item_instance.store_id = request.POST.getlist('store_id')[index]

                        # Save the instance to the database
                        store_item_instance.save()
                        
                        store_item_curreEdit(request.POST.getlist('store_id')[index],elem,given_date,'min', request.POST.getlist('item_quantity')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                # #print(5620)
                models.Store_Transaction_Detail.objects.bulk_create(order_details)
                storeTransactionHeader.total_amount = total_amounts
                
                storeTransactionHeader.save()
                # #print(4607)
                if(int(request.POST['with_purchase_job_order']) == 1 and request.POST.get('purchase_order_header_id',None)): #it is a purchase order
                    
                    for index, elem in enumerate(request.POST.getlist('item_id')):
                        if int(request.POST.getlist('itemInspect')[index]) == 0:

                            purchaseOrderItem = models.Purchase_Order_Detail.objects.get(
                                purchase_order_header_id =  request.POST['purchase_order_header_id'], item_id = elem)
                            
                            purchaseOrderItem.delivered_quantity += Decimal(
                                request.POST.getlist('item_quantity')[index])
                            purchaseOrderItem.delivered_amount += Decimal(
                                request.POST.getlist('item_price')[index])
                            purchaseOrderItem.delivered_amount_with_gst += Decimal(
                                request.POST.getlist('amount_with_gst')[index])
                            purchaseOrderItem.updated_at = datetime.now()
                            purchaseOrderItem.save()
                    
                    purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related(
                                'purchase_order_detail_set').get(pk=request.POST['purchase_order_header_id']) 
                    
                    # checking whether closed purchase order
                    flag = True
                        
                    for purchaseOrderDetail in purchaseOrderHeader.purchase_order_detail_set.all():
                        if Decimal(purchaseOrderDetail.quantity) > Decimal(purchaseOrderDetail.delivered_quantity):
                            flag = False
                            break    
                    if flag == True:
                        purchaseOrderHeader.delivery_status = 3
                    else:
                        purchaseOrderHeader.delivery_status = 2
                        purchaseOrderHeader.updated_at = datetime.now()
                        purchaseOrderHeader.save()  

            # job order incoming 
            if (request.POST.get('job_order_header_id',None) and int(request.POST['with_purchase_job_order']) == 2): #it is a job order
                jobOrderHeader = models.Job_Order.objects.get(pk = request.POST['job_order_header_id'] )  
            
                # #print('4648')
                #vendor job order transaction
                # storeTransactionVHead_count= models.Store_Transaction.objects.all().count()
                storeTransactionVhead = models.Store_Transaction()
                transaction_type = models.Transaction_Type.objects.get(name='MIST')
                storeTransactionVhead.transaction_type = transaction_type
                storeTransactionVhead.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                storeTransactionVhead.invoice_challan = request.POST['invoice_challan']
                if request.POST.get('vendor_id',None):
                    storeTransactionVhead.vendor_from_id = request.POST['vendor_id']
                if(int(request.POST['with_purchase_job_order']) == 2):
                    storeTransactionVhead.job_order_id =  request.POST[
                        'job_order_header_id']
                #print(4663)
                storeTransactionVhead.transaction_date = request.POST['transaction_date']
                storeTransactionVhead.total_amount = request.POST['total_amount']
                storeTransactionVhead.notes = request.POST['notes']
                storeTransactionVhead.save()
                models.Store_Transaction_Detail.objects.filter(store_transaction_header_id=storeTransactionVhead.id).delete()
                # # # #print(4666)
                order_details = []
                store = models.Store.objects.get(vendor_id = request.POST['vendor_id'])
                amount_total = Decimal(request.POST['total_amount'])
                #incoming material exported by vendor to company
                for index, elem in enumerate(request.POST.getlist('item_id')):
                    # # # #print(4674)
                    # # # #print(request.POST)
                    if int(request.POST.getlist('itemInspect')[index]) == 0:
                        # # # #print(4675)
                        order_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionVhead.id,
                                item_id=elem,
                                store_id=store.id,
                                quantity=request.POST.getlist('item_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('item_price')[index],
                                gst_percentage=request.POST.getlist(
                                    'gst_percentage')[index],
                                amount_with_gst=request.POST.getlist(
                                    'amount_with_gst')[index],
                                direction = 'incoming'
                                
                            )
                        )
                        # # # #print(4691)
                        #vendor store substraction
                        storeItem = models.Store_Item.objects.filter(
                        item_id=elem, store_id=store.id).first()
                        if storeItem :
                            storeItem.on_hand_qty -= Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.closing_qty -= Decimal(
                                request.POST.getlist('item_quantity')[index])
                            storeItem.updated_at = datetime.now()
                            storeItem.save()
                        # change in storeItemCurrent mout
                        # Fetch the last transaction_date less than the given_date if mout
                        given_date = request.POST['transaction_date']
                        
                        
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=elem, store_id=store.id,status=1, deleted=0).last()

                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date ,item_id=elem, store_id=store.id,status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date ,item_id=elem, store_id=store.id,status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                       
                        
                        store_item_instance = models.Store_Item_Current()

                        if record:
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty - Decimal(request.POST.getlist('item_quantity')[index]) 
                            store_item_instance.closing_qty = record.closing_qty - Decimal(request.POST.getlist('item_quantity')[index]) 
                        
                            if(store_item_instance.on_hand_qty<0):
                                raise ValueError(f"out quantity is more than available quantity")
                            # Set other fields for the new transaction
                            store_item_instance.store_transaction_id = storeTransactionVhead.id
                            store_item_instance.transaction_date = given_date
                            store_item_instance.item_id = elem
                            store_item_instance.store_id = store.id

                            # Save the instance to the database
                            store_item_instance.save()
                           
                        else:
                            message = "canot possible Item not present in store"
                            raise ValueError(message)
                        store_item_curreEdit(store.id,elem,given_date,'mout', request.POST.getlist('item_quantity')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)


                    jobOrderDetailsExist  =  models.Job_Order_Detail.objects.filter(job_order_header_id =  request.POST['job_order_header_id'] ,direction='incoming').exists()
                    # # # #print(jobOrderDetailsExist)
                    if jobOrderDetailsExist:
                        jobOrderDetails =  models.Job_Order_Detail.objects.get(job_order_header_id =  request.POST['job_order_header_id'] ,direction='incoming')
                        # # # #print(jobOrderDetails.quantity_result)
                        jobOrderDetails.quantity_result -=  Decimal(request.POST.getlist('item_quantity')[index])
                        # # # #print(4708)
                        jobOrderDetails.updated_at = datetime.now()
                        jobOrderDetails.save()
                    # # # #print(4709)
                
                
                jobOrderHeader = models.Job_Order.objects.get(pk = request.POST['job_order_header_id'])
                jobOrderDetails = models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['job_order_header_id'],direction ='outgoing')
                
                #outgoing material utilised by vendor to make the exported material
                for detail in jobOrderDetails:
                    store = models.Store.objects.get(vendor_id = storeTranasctionHeaderOld.vendor_id)
                    bomDetails = models.Bill_Of_Material_Detail.objects.filter(bill_of_material_header_id = jobOrderHeader.bom_type_head_id , item_id = detail.item_id).first()
                    jobOrderDetailsNew = models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['job_order_header_id'],direction ='outgoing',item_id = detail.item_id).first()
                    jobOrderDetailsNew.quantity_result -=  (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))
                    order_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionVhead.id,
                                item_id=detail.item_id,
                                store_id=store.id,
                                quantity=(bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0])),
                                rate=detail.item.price,
                                amount=((bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0])) * detail.item.price),
                                direction ='outgoing'
                            )
                        )
                    amount_total += Decimal(((bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0])) * detail.item.price))
                    storeItemExists = models.Store_Item.objects.filter(store_id=store.id,item_id = detail.item_id).exists()
                    if storeItemExists:
                        storeItem = models.Store_Item.objects.get(store_id=store.id,item_id = detail.item_id)
                        storeItem.on_hand_qty -= (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))
                        storeItem.closing_qty -=  (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))
                        storeItem.updated_at = datetime.now()
                        storeItem.save()  
                    # change in storeItemCurrent mout
                    # Fetch the last transaction_date less than the given_date if mout
                    given_date = request.POST['transaction_date']
                    
                    
                    # Check for the last record on the given_date
                    record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=detail.item_id, store_id=store.id ,status=1, deleted=0).last()

                    if not record:
                        # If no record is found for the given_date, look for the last record before that date
                        last_transaction_date = models.Store_Item_Current.objects.filter(
                            transaction_date__lt=given_date ,item_id=detail.item_id, store_id=store.id,status=1, deleted=0
                        ).aggregate(Max('transaction_date'))['transaction_date__max']
                        

                        if last_transaction_date:
                            # Fetch the record for the last_transaction_date
                            record = models.Store_Item_Current.objects.filter(
                                transaction_date=last_transaction_date ,item_id=detail.item_id, store_id=store.id,status=1, deleted=0
                            ).last()

                    # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                    
                    
                    store_item_instance = models.Store_Item_Current()

                    if record:
                        # Set values based on the last record found
                        store_item_instance.opening_qty = record.closing_qty
                        store_item_instance.on_hand_qty = record.closing_qty - (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))
                        store_item_instance.closing_qty = record.closing_qty - (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))
                        if(store_item_instance.on_hand_qty<0):
                            raise ValueError(f"out quantity is more than available quantity")

                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionVhead.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = detail.item_id
                        store_item_instance.store_id = store.id

                        # Save the instance to the database
                        store_item_instance.save()
                        
                    else:
                        message = "canot possible Item not present in store"
                        raise ValueError(message)
                    store_item_curreEdit(store.id,detail.item_id,given_date,'mout', (bomDetails.quantity * Decimal(request.POST.getlist('item_quantity')[0]))) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

                    jobOrderDetailsNew.updated_at = datetime.now()
                    jobOrderDetailsNew.save()
                models.Store_Transaction_Detail.objects.bulk_create(order_details)
                storeTransactionVhead.total_amount = amount_total
                storeTransactionVhead.save()
                # job order status change
                jobOrderHeader = models.Job_Order.objects.get(pk = request.POST['job_order_header_id'])
                all_material_recieved =  models.Job_Order_Detail.objects.filter(
                                            job_order_header_id=request.POST['job_order_header_id'],
                                            direction='incoming',
                                            quantity_result=0.00
                                        ).exists()

                # all material  recieved
                if  all_material_recieved:
                    jobOrderHeader.material_reciept = 1 
                    current_time = datetime.now(timezone.utc)
                    time_difference = current_time - jobOrderHeader.created_at
                    total_hours = time_difference.days * 24 + time_difference.seconds / 3600
                    jobOrderHeader.actual_time_take = str(round(total_hours, 3)) + 'hr'
                    jobOrderHeader.job_status = 2
                    jobOrderHeader.updated_at = datetime.now()
                    jobOrderHeader.save()
        
        transaction.commit()    
        
       
        text = f'Store Transaction Edit old Transaction no = {storeTranasctionHeaderOld.transaction_number} and  new transaction no = {new_transaction_no}'
        user_log_details_add(userId,text)
        # # # #print(4725)
        context.update({
            'status': 200,
            'message': "Store Transaction updated Successfully."
        }) 

    except Exception as e :
        #print(f" error is {e}")
        context.update({
            'status': 591.1,
            'message': "internal error",
        })
        transaction.rollback()     

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeTransactionDelete(request):
    context = {}
    storeTransaction = models.Store_Transaction.objects.prefetch_related('store_transaction_detail_set').get(
        pk=request.POST['id'])
    storeTranasctionHeaderOld =  models.Store_Transaction.objects.get(pk = request.POST['id'], status=1, deleted=0)
    try:
        with transaction.atomic():
            if storeTransaction.purchase_order_header_id is not None:
                purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related('purchase_order_detail_set').get(
                    pk=storeTransaction.purchase_order_header_id)
                for purchaseOrderDetail in purchaseOrderHeader.purchase_order_detail_set.all():
                    storeTransactionDetail = models.Store_Transaction_Detail.objects.filter(store_transaction_header_id=storeTransaction.id, item_id=purchaseOrderDetail.item_id).first()
                    if storeTransactionDetail is not None:
                        purchaseOrderDetail.delivered_quantity -= storeTransactionDetail.quantity
                        purchaseOrderDetail.delivered_amount -= storeTransactionDetail.quantity * purchaseOrderDetail.delivered_rate
                        purchaseOrderDetail.delivered_gst_percentage = purchaseOrderDetail.delivered_gst_percentage
                        purchaseOrderDetail.delivered_amount_with_gst -= (storeTransactionDetail.quantity * purchaseOrderDetail.delivered_rate) + (storeTransactionDetail.quantity * purchaseOrderDetail.delivered_rate * purchaseOrderDetail.delivered_gst_percentage / 100)
                        purchaseOrderDetail.updated_at = datetime.now()
                        purchaseOrderDetail.save()
                for storeTransactionDetail in storeTransaction.store_transaction_detail_set.all():
                    data_revertive_from_transaction(storeTranasctionHeaderOld.id, storeTransactionDetail.item_id, storeTransactionDetail.store_id, storeTransactionDetail.quantity, 'out')
                    store_item_current = models.Store_Item_Current.objects.filter(store_transaction_id = storeTranasctionHeaderOld.id,status=1, deleted=0)
                    if store_item_current.exists():
                        store_item_current = store_item_current.first()
                        store_item_current.on_hand_qty -= Decimal(storeTransactionDetail.quantity)
                        store_item_current.closing_qty -= Decimal(storeTransactionDetail.quantity)
                        store_item_current.status = 0
                        store_item_current.deleted = 1
                        store_item_current.updated_at = datetime.now()
                        store_item_current.save()
                    storeItem = models.Store_Item.objects.filter(item_id=storeTransactionDetail.item_id,
                                                                 store_id=storeTransactionDetail.store_id).first()
                    if storeItem is not None:
                        storeItem.on_hand_qty -= Decimal(
                            storeTransactionDetail.quantity)
                        storeItem.closing_qty -= Decimal(
                            storeTransactionDetail.quantity)
                        storeItem.updated_at = datetime.now()
                        storeItem.save()
                purchaseOrderHeader.delivery_status = 1
                purchaseOrderHeader.save()
            if (storeTransaction.job_order_id is  None) and (storeTransaction.purchase_order_header_id is  None) :
                for storeTransactionDetail in storeTransaction.store_transaction_detail_set.all():
                    storeItem = models.Store_Item.objects.filter(item_id=storeTransactionDetail.item_id,
                                                                 store_id=storeTransactionDetail.store_id).first()
                    
                    data_revertive_from_transaction(storeTranasctionHeaderOld.id, storeTransactionDetail.item_id, storeTransactionDetail.store_id, storeTransactionDetail.quantity, 'out')
                    store_item_current = models.Store_Item_Current.objects.filter(store_transaction_id = storeTranasctionHeaderOld.id,status=1, deleted=0)
                    if store_item_current.exists():
                        store_item_current = store_item_current.first()
                        store_item_current.on_hand_qty -= Decimal(storeTransactionDetail.quantity)
                        store_item_current.closing_qty -= Decimal(storeTransactionDetail.quantity)
                        store_item_current.status = 0
                        store_item_current.deleted = 1
                        store_item_current.updated_at = datetime.now()
                        store_item_current.save()
                    if storeItem is not None:
                        storeItem.on_hand_qty -= Decimal(
                            storeTransactionDetail.quantity)
                        storeItem.closing_qty -= Decimal(
                            storeTransactionDetail.quantity)
                        storeItem.updated_at = datetime.now()
                        storeItem.save()
                
                
            if storeTransaction.job_order_id is not None:
                jobOrderStoreTranasctionRetriveVendor(storeTranasctionHeaderOld)
                jobOrderStoreTranasctionRetriveInHouse(storeTranasctionHeaderOld)
               
                
                jobOrderHeader = models.Job_Order.objects.get(pk = storeTranasctionHeaderOld.job_order.id)
                all_material_recieved =  models.Job_Order_Detail.objects.filter(
                                            job_order_header_id=storeTranasctionHeaderOld.job_order.id,
                                            direction='incoming',
                                            quantity_result=0.00
                                        ).exists()
                
                if not all_material_recieved:
                    jobOrderHeader.material_reciept = 0
                    jobOrderHeader.actual_time_take = ''
                    jobOrderHeader.job_status = 1
                    jobOrderHeader.updated_at = datetime.now()
                    jobOrderHeader.save()

            storeTransaction.status = 0
            storeTransaction.deleted = 1
            storeTransaction.save()
            # storeTransaction.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Store Transaction Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Store Transaction Deleted Successfully."
        })
    except Exception as e:
        #print(f"error{e}")
        context.update({
            'status': 592,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def storeTransactionDetails(request):
    context = {}
    header_id = request.GET.get('header_id', None)
    if header_id is not None and header_id != "":
        header_detail = list(models.Store_Transaction.objects.filter(id=header_id).values('pk', 'transaction_number', 'transaction_date', 'total_amount',
                             'purchase_order_header_id', 'purchase_order_header__order_number', 'vendor__name', 'transaction_type_id', 'transaction_type__name'))
        orderDetails = list(models.Store_Transaction_Detail.objects.filter(store_transaction_header_id=header_id).values('pk', 'quantity', 'rate', 'amount', 'gst_percentage', 'amount_with_gst',
                            'item_id', 'item__name', 'store_id', 'store__name', 'store_transaction_header_id', 'store_transaction_header__transaction_number', 'store_transaction_header__transaction_date'))
        context.update({
            'status': 200,
            'message': "Store Transaction Details Fetched Successfully.",
            'header_detail': header_detail,
            'page_items': orderDetails,
        })
    else:
        context.update({
            'status': 588,
            'message': "Please Provide Header Id.",
        })
    return JsonResponse(context)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def storeTransactionLogicalGrnEdit(request):
    context = {}
    #print(request.POST)
    storeTransactionHeader = models.Store_Transaction.objects.get(pk=request.POST['pk'])
    try:
        with transaction.atomic():
            storeTransactionHeader.dispatch_no = request.POST['dispatch_no']
            storeTransactionHeader.vehicle = request.POST['vehicle']
            storeTransactionHeader.dispatch_through = request.POST['dispatch_through']
            storeTransactionHeader.notes = request.POST['notes']
            storeTransactionHeader.destination = request.POST['destination']
            storeTransactionHeader.eway_bill = request.POST['eway_bill']
            storeTransactionHeader.updated_at = datetime.now()
            storeTransactionHeader.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'logical grn vechical details updated Successfully.')   
            context.update({
                'status': 200,
                'message': "logical grn vechical details updated Successfully."
            })


        transaction.commit()
    except Exception as e:
        context.update({
                'status': 592.1,
                'message': f"logical grn vechical details not successfully.{e}"
            })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def jobOrderList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    vendor = request.GET.get('vendor_id', None)
    with_item = request.GET.get('with_item', None)
    material_reciept = request.GET.get('material_reciept', None)
    material_issue= request.GET.get('material_issue', None)
    manufacturing_type = request.GET.get('manufacturing_type', None)
    if id is not None and id != "":
        jobOrder = list(models.Job_Order.objects.filter(pk=id)[:1].values('pk', 'order_number', 'order_date', 'manufacturing_type', 'vendor_id', 'vendor__name', 'with_item', 'notes','material_issue','job_status','estimated_time_day','material_reciept','manufacturing_material_type'))
        context.update({
            'status': 200,
            'message': "Job Order Fetched Successfully.",
            'page_items': jobOrder,
        })
    else:
        jobOrders = models.Job_Order.objects.filter(status=1, deleted=0)
        if keyword is not None and keyword != "":
            
            job_status = 0 if keyword.lower() == 'open' else ( 1 if keyword.lower() == 'wip' else ( 2 if keyword.lower() == 'closed' else -1 ) )
            jobOrders = jobOrders.filter(Q(order_number__icontains=keyword) | Q(job_status__icontains = job_status) ).filter(status=1, deleted=0)
        if vendor is not None and vendor != "":
            jobOrders = jobOrders.filter(vendor_id=vendor).filter(status=1, deleted=0)
            if with_item is not None and with_item != "":
                jobOrders = jobOrders.filter(vendor_id=vendor ,with_item=with_item ).filter(status=1, deleted=0)
                if material_reciept is not None and material_reciept != "" and manufacturing_type is not None and manufacturing_type !="":
                    jobOrders = jobOrders.filter(material_reciept=material_reciept , manufacturing_type=manufacturing_type).filter(status=1, deleted=0)
        if material_issue is not None and material_issue != "":
            material_issue_list = [int(x) for x in material_issue.split(',')]
            jobOrders = jobOrders.filter(material_issue__in=material_issue_list ).filter(status=1, deleted=0)
        jobOrders = list(jobOrders.values('pk', 'order_number', 'order_date', 'manufacturing_type', 'vendor_id', 'vendor__name', 'with_item', 'notes','material_issue','job_status','estimated_time_day','material_reciept','manufacturing_material_type'))
        if find_all is not None and int(find_all) == 1:
            # #print(jobOrders)
            context.update({
                'status': 200,
                'message': "Job Orders Fetched Successfully.",
                'page_items': jobOrders,
            })
            return JsonResponse(context)

        per_page = int(env("PER_PAGE_DATA"))
        button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
        current_page = request.GET.get('current_page', 1)

        paginator = CustomPaginator(jobOrders, per_page)
        page_items = paginator.get_page(current_page)
        total_pages = paginator.get_total_pages()

        context.update({
            'status': 200,
            'message': "Job Orders Fetched Successfully.",
            'page_items': page_items,
            'total_pages': total_pages,
            'per_page': per_page,
            'current_page': int(current_page),
            'button_to_show': int(button_to_show),
        })
    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def jobOrderNo(request):
    context ={}
    manufacturing_type= request.GET.get('keyword',None)
    jobOrderlast = models.Job_Order.objects.filter(manufacturing_type=manufacturing_type).last()
    jobOrderCount_match = re.search(r'SEC/(?:TPM|SLF)/(\d{3})/', jobOrderlast.order_number)  # Assuming order_number is the field name
    jobOrderCount = int(jobOrderCount_match.group(1)) if jobOrderCount_match else 0
    vendorShort = 'SLF' if manufacturing_type == 'Self' else 'TPM'
    jobOrderNumber =  env("JOB_ORDER_NUMBER_SEQ").replace("${VENDOR_SHORT}", vendorShort).replace(
                "${AI_DIGIT_3}", str(jobOrderCount + 1).zfill(3)).replace("${FINANCE_YEAR}", datetime.today().strftime('%y') + "-" + (datetime(datetime.today().year + 1, 1, 1).strftime('%y')))
    # # # # #print(jobOrderNumber)
    context.update({
        'status':200,
        'joborderNo': jobOrderNumber
    })
    return JsonResponse(context)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def jobOrderAdd(request):
    context = {}
    # #print(request.POST)
    bomNeeded = request.POST.get('bomNeeded',None)
    if not request.POST['order_number'] or not request.POST['order_date'] or not request.POST['manufacturing_type'] or not request.POST['notes']:
        context.update({
            'status': 589,
            'message': "Order Number/Order Date/Manufacturing Type/Notes has not been provided."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            # #print('5055')
            jobOrderHeader = models.Job_Order()
            jobOrderHeader.order_number = request.POST['order_number']
            jobOrderHeader.order_date = request.POST['order_date']
            jobOrderHeader.manufacturing_type = request.POST['manufacturing_type']
            jobOrderHeader.manufacturing_material_type = request.POST['manufacturing_material_type']
            # #print(5757)
            if 'vendor_id' in request.POST:
                jobOrderHeader.vendor_id = request.POST['vendor_id']
            if 'with_item' in request.POST:
                jobOrderHeader.with_item = eval(request.POST['with_item'])
            # #print(5762)
            jobOrderHeader.notes = request.POST['notes']
            if request.POST.get('bom_type_id',None):
                # #print(request.POST['bom_type_id'],'aaaaaaaaa')
                jobOrderHeader.bom_type_head_id = int(request.POST['bom_type_id'])
            jobOrderHeader.estimated_time_day =  request.POST['hr_day'] + request.POST['timeSpan']
            # #print(5767)
            jobOrderHeader.save()
            # #print(5769)
            job_order_details = []
            bom_material_details = []
            bom_head = None 
            bom_head_exit = None
            # #print('5065')
            if (request.POST.getlist('incoming_item_id')) and (request.POST.getlist('outgoing_item_id')) and ('with_item' in request.POST):
                # #print('5050')
                outgoingIncommingratioHeadCount = models.Outgoing_Incoming_Ratio.objects.all().count() 
                outgoingIncommingratioHead = models.Outgoing_Incoming_Ratio()
                outgoingIncommingratioHead.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", datetime.today().strftime('%Y')
                ).replace(
                    "${AI_DIGIT_5}",str(outgoingIncommingratioHeadCount + 1).zfill(5)
                )
                outgoingIncommingratioHead.transaction_date = request.POST['order_date']
                if 'vendor_id' in request.POST:
                    outgoingIncommingratioHead.vendor_id = request.POST['vendor_id']
                outgoingIncommingratioHead.job_order = jobOrderHeader
                outgoingIncommingratioHead.save()
                # # # # # #print('5061')
                outInDetailRatio = []
                for item_id,quantity in zip( request.POST.getlist('outgoing_item_id'),request.POST.getlist('outgoing_quantity') ):
                    
                    # # # # # #print(int(float(request.POST.getlist('incoming_quantity')[0])))
                    ratio = Fraction(int(float((quantity))) , int(float(request.POST.getlist('incoming_quantity')[0]))) #ratio = outgoing : income
                    # # # # # #print(ratio)
                    # # # # # #print('id',outgoingIncommingratioHead.id)
                    # # # # # #print('item',int(request.POST.getlist('incoming_item_id')[0]))
                    # # # # # #print('neu', ratio.numerator)
                    # # # # # #print('deno', ratio.denominator)
                    # # # # # #print('eato',str(ratio) )
                    outInDetailRatio.append(
                        models.Outgoing_Incoming_Ratio_Details(
                            outgoing_incoming_ratio_header_id = outgoingIncommingratioHead.id,
                            item_outgoing_id = int(item_id),
                            item_incomming_id = int(request.POST.getlist('incoming_item_id')[0]),
                            numerator = ratio.numerator,
                            denominator = ratio.denominator,
                            ratio = str(ratio)  

                        )
                    )
                # # # # # #print('4995')
                models.Outgoing_Incoming_Ratio_Details.objects.bulk_create(outInDetailRatio)
            
            #for incomming quantity 
            incomming_item_id = 0
            for item_id,quantity in zip( request.POST.getlist('incoming_item_id'),request.POST.getlist('incoming_quantity') ):
                job_order_details.append(
                    models.Job_Order_Detail(
                        job_order_header_id=jobOrderHeader.id,
                        item_id=int(item_id),
                        quantity=float(quantity),
                        required_quantity=float(quantity),
                        quantity_result = float(quantity),
                        direction="incoming"
                    )
                )
                incomming_item_id = int(item_id)
                incomming_item_quantity = float(quantity)

            # bill of material add
            if (request.POST.get('bomNeeded',None) and int(bomNeeded) == 1 ) : 
                if  (models.Bill_Of_Material.objects.filter(bom_item_id = incomming_item_id, status=1, deleted=0).exists()) :
                    bom_head_exit = models.Bill_Of_Material.objects.filter(bom_item_id = incomming_item_id, status=1, deleted=0).first()
                    jobOrderHeader.bom_type_head_id = bom_head_exit.id
                else: 
                    # # # # #print(5145)
                    billOfMaterialMasterHeaderExist = models.Bill_Of_Material_Master.objects.filter(item_id = incomming_item_id).exists()
                    if not billOfMaterialMasterHeaderExist:
                        
                        billOfMaterialHeaderMaster = models.Bill_Of_Material_Master()
                        
                        billOfMaterialHeaderMaster.item_id = incomming_item_id
                        billOfMaterialHeaderMaster.save()
                       
                    else:
                        # # #print(request.POST['bom_item_id'])
                        billOfMaterialHeaderMaster = models.Bill_Of_Material_Master.objects.filter(item_id = request.POST['bom_item_id']).first()
                    bom_head = models.Bill_Of_Material()
                    bom_head.bom_item_id = incomming_item_id
                    bom_head.quantity = 1
                    bom_head.price =  models.Item.objects.get(pk=incomming_item_id).price
                    bom_head.uom_id =models.Item.objects.get(pk=incomming_item_id).uom_id
                    bom_head.bom_master_id = billOfMaterialHeaderMaster.id
                    bom_head.bom_type = 1
                    bom_head.save()
                    jobOrderHeader.bom_type_head_id = bom_head.id
                    jobOrderHeader.save()
            bomLevel = 1
            #for Outgoing Quantity
            for item_id, quantity in zip(request.POST.getlist('outgoing_item_id'),request.POST.getlist('outgoing_quantity')):
                # # #print(5154)
                job_order_details.append(
                    models.Job_Order_Detail(
                        job_order_header_id=jobOrderHeader.id,
                        item_id=int(item_id),
                        quantity=float(quantity),
                        required_quantity=float(quantity),
                        quantity_result = float(quantity),
                        direction="outgoing"
                    )
                )
              
                if (request.POST.get('bomNeeded',None) and int(bomNeeded) == 1) and  bom_head  :
                    bomExist = models.Bill_Of_Material_Master.objects.filter(item_id = item_id).exists()
                    # #print(bomExist,'sssssssss',item_id)
                    if bomExist :
                        bomHead =  models.Bill_Of_Material.objects.filter(bom_item_id = item_id).first()
                        x = bomHead.level
                        bomHeadId = bomHead.id
                        # #print(type(bomLevel) == type(bomHead.level))
                        bomLevel += x
                        # #print(bom_head.id)
                    bom_material_details.append(
                        models.Bill_Of_Material_Detail(
                            bill_of_material_header_id=bom_head.id,
                            item_id = int(item_id) if not bomExist else None,
                            bom_level_id  = bomHeadId if bomExist else None,
                            quantity=float(quantity) / incomming_item_quantity,
                            price = models.Item.objects.get(pk=item_id).price
                        )
                    )
          
                # # #print(5822)
                if(request.POST.get('bomNeeded',None) and int(bomNeeded) == 1) and bom_head_exit :
                    models.Bill_Of_Material_Detail.objects.filter(bill_of_material_header_id = bom_head_exit.id).delete()
                    bomExist = models.models.Bill_Of_Material_Master.objects.filter(item_id = item_id).exists()
                    if bomExist :
                        bomHead =  models.Bill_Of_Material.objects.filter(bom_item_id = item_id).first()
                        bomHeadId = bomHead.id
                        x = bomHead.level
                        bomLevel += x
                    bom_material_details.append(
                        models.Bill_Of_Material_Detail(
                            bill_of_material_header_id=bom_head_exit.id,
                            item_id = int(item_id) if not bomExist else None,
                            bom_level_id  = bomHeadId if bomExist else None,
                            quantity=float(quantity)/incomming_item_quantity,
                            price = models.Item.objects.get(pk=item_id).price
                        )
                    )
                
            # # #print(5843)
            if bom_material_details:
                models.Bill_Of_Material_Detail.objects.bulk_create(bom_material_details)
                bom_head.level = int(bomLevel)
                bom_head.save() 
            models.Job_Order_Detail.objects.bulk_create(job_order_details)
            # # # # # #print('4271')
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Job Order Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Job Order Created Successfully."
        })
    except Exception as e:
        #print(f"An error occurred: {e}")
        context.update({
            'status': 590,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)

def jobOrderTimeComplete(pk):
    id = pk 
    jobOrderHeader = models.Job_Order.objects.prefetch_related('job_order_detail_set').get(pk=id)
    current_time = datetime.now(timezone.utc)
    time_difference = current_time - jobOrderHeader.created_at
    total_hours = time_difference.days * 24 + time_difference.seconds / 3600
    jobOrderHeader.actual_time_take = str(round(total_hours, 3)) + 'hr'
    jobOrderHeader.job_status = 2 # 0 for active 1 for in pogress 2 for complete
    jobOrderHeader.material_reciept = 2 # full mateial recieved
    jobOrderHeader.updated_at = datetime.now()
    jobOrderHeader.save()

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def selfJobOrderReciept(request):
    context = {}
    
    id = request.POST['pk']
    userId = request.COOKIES.get('userId', None)
    try:
        jobOrderHeader = models.Job_Order.objects.prefetch_related('job_order_detail_set').get(pk=id)
        jobOrderDetails = list(models.Job_Order_Detail.objects.filter(job_order_header_id= id , direction='incoming'))

        with transaction.atomic():
        # check whether current recieved quantity excceded from requireed resultant material recieved quantity
            for index in range(0,len(jobOrderDetails)):
                jobOrderDetEdit = models.Job_Order_Detail.objects.filter(job_order_header_id=id,item_id = jobOrderDetails[index].item_id, direction='incoming').first()
                if(float(jobOrderDetEdit.quantity_result)<float( request.POST['incoming_quantity'])):
                    context.update({
                    'status': 500,
                    'message': f"current recieved quantity cannot be more than {int(float(jobOrderDetEdit.quantity_result))} "
                    })
                    return JsonResponse(context)
            # creation of store transaction for job order for material recieved
            store_transaction_count = models.Store_Transaction.objects.all().count()
            storeTransactionHeader=models.Store_Transaction()
            # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'SP')
            transaction_type = models.Transaction_Type.objects.get(name='SP')
            storeTransactionHeader.transaction_type = transaction_type
            storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                "${CURRENT_YEAR}", current_year).replace(
                "${AI_DIGIT_5}", ai_digit_5()).replace(
                "${transaction_type_id}", str(transaction_type.id).zfill(2))

            storeTransactionHeader.transaction_date= request.POST['transaction_date']
            storeTransactionHeader.job_order_id = id
            storeTransactionHeader.creator_id = userId
            storeTransactionHeader.save()
           
            orderDetails =[]
            material_reciept_all = False
            for index in range(0,len(jobOrderDetails)):
                # changes resultant required material reciept quantity from job order details
                jobOrderDetEdit = models.Job_Order_Detail.objects.filter(job_order_header_id=id,item_id = jobOrderDetails[index].item_id, direction='incoming').first()
                jobOrderDetEdit.quantity_result = float(jobOrderDetEdit.quantity_result) - float( request.POST['incoming_quantity'])
                jobOrderDetEdit.updated_at = datetime.now()
                jobOrderDetEdit.save()

                # store transaction details save 
                orderDetails.append(
                    models.Store_Transaction_Detail(
                        store_transaction_header = storeTransactionHeader,
                        item_id = jobOrderDetails[index].item_id,
                        store_id = request.POST['store_id'],
                        quantity = request.POST['incoming_quantity'],
                        rate = jobOrderDetails[index].item.price,
                        amount = float(jobOrderDetails[index].item.price) * float(jobOrderDetails[index].quantity)
                    )
                )

               
                resultant_quantity_result =  models.Job_Order_Detail.objects.filter(job_order_header_id= id,direction='outgoing')
                
                for detail in resultant_quantity_result:
                    boMHeadDetailsExist = models.Bill_Of_Material_Detail.objects.filter(item_id = detail.item_id, bill_of_material_header_id = jobOrderHeader.bom_type_head_id).exists()
                    if boMHeadDetailsExist:
                        bomDetailsFirst = models.Bill_Of_Material_Detail.objects.filter(item_id= detail.item_id , bill_of_material_header_id =jobOrderHeader.bom_type_head).first()
                        BomQuantity  = float(bomDetailsFirst.quantity)
                    resultant_quantity_result_first = models.Job_Order_Detail.objects.filter(item_id = detail.item_id, job_order_header_id= id,direction='outgoing').first()
                    resultant_quantity_result_first.quantity_result = 0.0  if not boMHeadDetailsExist else (resultant_quantity_result_first.quantity_result - Decimal(BomQuantity* float(request.POST['incoming_quantity'])))
                    
                    resultant_quantity_result_first.updated_at = datetime.now()
                    resultant_quantity_result_first.save()
                    
                # checking whether material fully recieved or not
                material_reciept_all = True  if (float(jobOrderDetEdit.quantity_result) == 0.00) else False
                storeItem = models.Store_Item.objects.filter(item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id']).exists()
                
                if storeItem:
                    storeItem = models.Store_Item.objects.filter(item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id']).first()
                    storeItem.on_hand_qty += Decimal(request.POST['incoming_quantity'])
                    storeItem.closing_qty += Decimal(request.POST['incoming_quantity'])
                    storeItem.updated_at = datetime.now()
                    storeItem.save()
                else:
                    storeItem = models.Store_Item()
                    storeItem.item_id = jobOrderDetails[index].item_id
                    storeItem.store_id = request.POST['store_id']
                    storeItem.opening_qty = Decimal(request.POST['incoming_quantity'])
                    storeItem.on_hand_qty = Decimal(request.POST['incoming_quantity'])
                    storeItem.closing_qty =Decimal(request.POST['incoming_quantity'])
                    storeItem.save()
                   
                    # change in storeItemCurrent min
                    # Fetch the last transaction_date less than the given_date
                given_date = request.POST['transaction_date']
                
                # Check for the last record on the given_date
                record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id'],status=1, deleted=0).last()

                if not record:
                    # If no record is found for the given_date, look for the last record before that date
                    last_transaction_date = models.Store_Item_Current.objects.filter(
                        transaction_date__lt=given_date,item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id'],status=1, deleted=0
                    ).aggregate(Max('transaction_date'))['transaction_date__max']

                    

                    if last_transaction_date:
                        # Fetch the record for the last_transaction_date
                        record = models.Store_Item_Current.objects.filter(
                            transaction_date=last_transaction_date,item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id'],status=1, deleted=0
                        ).last()

                # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                
                
                store_item_instance = models.Store_Item_Current()

                if record:
                    # Set values based on the last record found
                    store_item_instance.opening_qty = record.closing_qty
                    store_item_instance.on_hand_qty = record.closing_qty + Decimal(
                        request.POST['incoming_quantity']
                    )
                    store_item_instance.closing_qty = record.closing_qty + Decimal(
                        request.POST['incoming_quantity']
                    )
                else:
                    # Set values based on the current transaction if no prior record exists
                    store_item_instance.opening_qty = Decimal(
                        0.00
                    )
                    store_item_instance.on_hand_qty = Decimal(
                        request.POST['incoming_quantity']
                    )
                    store_item_instance.closing_qty = Decimal(
                        request.POST['incoming_quantity']
                    )
                if(store_item_instance.on_hand_qty<0):
                    raise ValueError(f"ohnand quantity is less than 0")
                # Set other fields for the new transaction
                store_item_instance.store_transaction_id = storeTransactionHeader.id
                store_item_instance.transaction_date = given_date
                store_item_instance.item_id = jobOrderDetails[index].item_id
                store_item_instance.store_id = request.POST['store_id']
                # Save the instance to the database
                store_item_instance.save()
               
                storeInstanceexist = models.Store_Item_Current.objects.filter(item_id = jobOrderDetails[index].item_id,store_id = request.POST['store_id'],transaction_date = given_date,store_transaction_id= storeTransactionHeader.id,status=1, deleted=0)
                
                if(not storeInstanceexist.exists()):
                    
                    raise ValueError('This is a test exception')

                store_item_curreEdit(request.POST['store_id'],jobOrderDetails[index].item_id,given_date,'min',request.POST['incoming_quantity']) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

            models.Store_Transaction_Detail.objects.bulk_create(orderDetails)
            # full material recived closing job order task 
           
            if(material_reciept_all == True):
                jobOrderTimeComplete(id)
                text = f'Job Order NO: {jobOrderHeader.order_number} closed succesfully and store transaction created'
                user_log_details_add(userId,text)
            text = 'store transaction created'
        transaction.commit()
        user_log_details_add(userId,text)
        text = text + ' ' + 'Sucessfully'
        context.update({
                'status': 200,
                'message': text
            })
    except Exception as e:
        #print(f'issue: {e}')
        context.update({
            'status': 593,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()    
    return JsonResponse(context)    
   

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def jobOrderEdit(request):
    context = {}
    bomNeeded = request.POST.get('bomNeeded',None)
    if not request.POST['order_number'] or not request.POST['order_date']  or not request.POST['notes']:
        context.update({
            'status': 589,
            'message': "Order Number/Order Date/Manufacturing Type/Notes has not been provided."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            jobOrderHeader = models.Job_Order.objects.prefetch_related('job_order_detail_set').get(pk=request.POST['id'])
            jobOrderHeader.order_number = request.POST['order_number']
            # #print(request.POST)
            jobOrderHeader.order_date = request.POST['order_date']
            # jobOrderHeader.manufacturing_type = request.POST['manufacturing_type']
            if 'vendor_id' in request.POST:
                jobOrderHeader.vendor_id = request.POST['vendor_id']
            if 'with_item' in request.POST:
                jobOrderHeader. with_item = eval(request.POST['with_item'])
            else:
                jobOrderHeader.with_item = False
            jobOrderHeader.estimated_time_day =  request.POST['hr_day'] + request.POST['timeSpan']
            jobOrderHeader.manufacturing_material_type = request.POST['manufacturing_material_type']
            jobOrderHeader.notes = request.POST['notes']
            if request.POST.get('bom_type_id',None):
                jobOrderHeader.bom_type_head_id = request.POST['bom_type_id']
            # #print(models.Bill_Of_Material.objects.filter(pk=request.POST['bom_type_id']).exists())
            jobOrderHeader.updated_at = now()
            # #print(6068)
            jobOrderHeader.save()
            if jobOrderHeader.material_issue == 0:
                jobOrderHeader.job_order_detail_set.all().delete()
            job_order_details = []
            outInDetailRatio =[]
            bom_material_details = []
            bom_head = None 
            bom_head_exit = None
            # #print(6076)
            # out going incomming ratio table updation
            if (request.POST.getlist('incoming_item_id')) and (request.POST.getlist('outgoing_item_id')) and ('with_item' in request.POST):
                
                outgoingIncommingratioHead = models.Outgoing_Incoming_Ratio.objects.prefetch_related('outgoing_incoming_ratio_details_set').get(job_order_id = request.POST['id'])
                
                outgoingIncommingratioHead.updated_at = now()
                outgoingIncommingratioHead.save()
                if jobOrderHeader.material_issue == 0:
                    outgoingIncommingratioHead.outgoing_incoming_ratio_details_set.all().delete()
                
                # ratio table
                for item_id, quantity in zip(request.POST.getlist('outgoing_item_id'), request.POST.getlist('outgoing_quantity')):
                    ratio = Fraction(int(float(quantity)) , int(float(request.POST.getlist('incoming_quantity')[0]))) #ratio = outgoing : incomeG
                    
                    outInDetailRatio.append(
                        models.Outgoing_Incoming_Ratio_Details(
                            outgoing_incoming_ratio_header_id = outgoingIncommingratioHead.id,
                            item_outgoing_id = int(item_id),
                            item_incomming_id = int(request.POST.getlist('incoming_item_id')[0]),
                            numerator = ratio.numerator,
                            denominator = ratio.denominator,
                            ratio = str(ratio)  

                        )
                    )
                
                models.Outgoing_Incoming_Ratio_Details.objects.bulk_create(outInDetailRatio)
            #incomming details     
            for item_id, quantity in zip(request.POST.getlist('incoming_item_id'), request.POST.getlist('incoming_quantity')):
                
                if  jobOrderHeader.material_issue == 0:
                    job_order_details.append(
                        models.Job_Order_Detail(
                            job_order_header_id=jobOrderHeader.id,
                            item_id=int(item_id),
                            quantity=float(quantity),
                            required_quantity=float(quantity),
                            quantity_result = float(quantity),
                            direction="incoming"
                        )
                    )
                incomming_item_id = int(item_id)
                incomming_item_quantity = float(quantity)
            # # #print(5299,'aaaaaaaaaaa')
            # bill of material add

            #outgoing
            for item_id, quantity in zip(request.POST.getlist('outgoing_item_id'), request.POST.getlist('outgoing_quantity')):
                if  jobOrderHeader.material_issue == 0:
                    job_order_details.append(
                        models.Job_Order_Detail(
                            job_order_header_id=jobOrderHeader.id,
                            item_id=int(item_id),
                            quantity=float(quantity),
                            required_quantity=float(quantity),
                            quantity_result = float(quantity),
                            direction="outgoing"
                        )
                    )
            if (request.POST.get('bomNeeded',None) and int(bomNeeded) == 1) : 
        
                if  (models.Bill_Of_Material.objects.filter(bom_item_id = incomming_item_id, status=1, deleted=0).exists()) :
                    bom_head_exit = models.Bill_Of_Material.objects.filter(bom_item_id = incomming_item_id, status=1, deleted=0).first()
                    jobOrderHeader.bom_type_head_id = bom_head_exit.id
                    jobOrderHeader.save()
                else: 
                    billOfMaterialMasterHeaderExist = models.Bill_Of_Material_Master.objects.filter(item_id = incomming_item_id).exists()
                    if not billOfMaterialMasterHeaderExist:
                        # # #print(6011,'aaaaaaaaaaa')
                        billOfMaterialHeaderMaster = models.Bill_Of_Material_Master()
                        # # #print('11111111')
                        billOfMaterialHeaderMaster.item_id = incomming_item_id
                        # # #print(6040)
                        billOfMaterialHeaderMaster.save()
                    else:
                        # # #print(request.POST['bom_item_id'])
                        billOfMaterialHeaderMaster = models.Bill_Of_Material_Master.objects.filter(item_id = request.POST['bom_item_id']).first()
  
                    bom_head = models.Bill_Of_Material()
                    bom_head.bom_item_id = incomming_item_id
                    bom_head.quantity = 1 
                    bom_head.bom_master_id = billOfMaterialHeaderMaster.id
                    bom_head.bom_type = 1
                    
                    bom_head.price =  models.Item.objects.get(pk=incomming_item_id).price
                    bom_head.uom_id =models.Item.objects.get(pk=incomming_item_id).uom_id

                    bom_head.save()
                    jobOrderHeader.bom_type_head_id = bom_head.id
                    jobOrderHeader.save()
            # outgoing details
            bomLevel = 1
            for item_id, quantity in zip(request.POST.getlist('outgoing_item_id'), request.POST.getlist('outgoing_quantity')):
                jobOrderDetailsExist = models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['id'], item_id = int(item_id), quantity__gt =F('required_quantity')).exists()
                if not jobOrderDetailsExist:
                    jobOrderDetailsExist = models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['id'], item_id = int(item_id)).exists()
                    if jobOrderDetailsExist:
                        models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['id'], item_id = int(item_id)).delete()
                        job_order_details.append(
                            models.Job_Order_Detail(
                                job_order_header_id=jobOrderHeader.id,
                                item_id=int(item_id),
                                quantity=float(quantity) /incomming_item_quantity,
                                required_quantity=float(quantity),
                                quantity_result = float(quantity),
                                direction="outgoing"
                            )
                        )
                if (request.POST.get('bomNeeded',None) and int(bomNeeded) == 1) and  bom_head :
                    # # # # #print(5165)
                    bomExist = models.Bill_Of_Material_Master.objects.filter(item_id = item_id).exists()
                    if bomExist :
                        bomHead =   models.Bill_Of_Material.objects.filter(bom_item_id = item_id).first()
                        bomHeadId = bomHead.id
                        x = bomHeadId.level
                        bomLevel += x
                        
                    bom_material_details.append(
                        models.Bill_Of_Material_Detail(
                            bill_of_material_header_id=bom_head.id,
                            item_id = int(item_id) if not bomExist else None,
                            bom_level_id  = bomHeadId if bomExist else None,
                            quantity=float(quantity)/incomming_item_quantity,
                            price = models.Item.objects.get(pk=item_id).price
                        )
                    )
          

                if(request.POST.get('bomNeeded',None) and int(bomNeeded) == 1) and bom_head_exit :

                    models.Bill_Of_Material_Detail.objects.filter(bill_of_material_header_id = bom_head_exit.id).delete()
                    bomExist = models.models.Bill_Of_Material_Master.objects.filter(item_id = item_id).exists()
                    if bomExist :
                        bomHead = models.Bill_Of_Material.objects.filter(bom_item_id = item_id).first()
                        bomHeadId = bomHead.id
                        x = bomHeadId.level
                        bomLevel += x
                    bom_material_details.append(
                        models.Bill_Of_Material_Detail(
                            bill_of_material_header_id=bom_head_exit.id,
                            item_id = int(item_id) if not bomExist else None,
                            bom_level_id  = bomHeadId if bomExist else None,
                            quantity=float(quantity),
                            price = models.Item.objects.get(pk=item_id).price
                        )
                    )
            
            if bom_material_details:
                models.Bill_Of_Material_Detail.objects.bulk_create(bom_material_details)
                bom_head.level = bomLevel
                bom_head.save()
            models.Job_Order_Detail.objects.bulk_create(job_order_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Job Order Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Job Order Updated Successfully."
        })
    except Exception as e :
        # #print(type(e))
        #print(f"error is {e}")
        context.update({
            'status': 592,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def jobOrderDelete(request):
    context = {}
    jobOrder = models.Job_Order.objects.get(pk=request.POST['id'])
    try:
        with transaction.atomic():
            jobOrder.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Job Order Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Job Order Deleted Successfully."
        })
    except Exception:
        context.update({
            'status': 593,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def jobOrderDetails(request):
    context = {}
    # # # #print(5493)
    header_id = request.GET.get('header_id', None)
    direction = request.GET.get('direction', None)
    materialReciept = request.GET.get('materialReciept', None)
    if header_id is not None and header_id != "":
        header_detail = list(models.Job_Order.objects.filter(pk=header_id)[:1].values('pk', 'order_number', 'order_date', 'manufacturing_type', 'vendor_id', 'vendor__name', 'with_item', 'notes'))
        if direction is not None and direction != "":
            if direction == 'incoming' :
                orderDetails = list(models.Job_Order_Detail.objects.filter(job_order_header_id=header_id, direction=direction,quantity_result__gt=0).values('pk', 'job_order_header_id', 'job_order_header__order_number','item_id', 'quantity', 'required_quantity', 'item__name','item__price', 'direction','item__item_type__gst_percentage','quantity_result'))
            else:
                orderDetails = list(models.Job_Order_Detail.objects.filter(job_order_header_id=header_id, direction=direction,required_quantity__gt=0).values('pk', 'job_order_header_id', 'job_order_header__order_number','item_id', 'quantity', 'required_quantity', 'item__name','item__price', 'direction','item__item_type__gst_percentage','quantity_result'))
            
        else:
            orderDetails = list(models.Job_Order_Detail.objects.filter(job_order_header_id=header_id).values('pk', 'job_order_header_id', 'job_order_header__order_number','item_id', 'quantity', 'required_quantity','item__name','item__price', 'direction','item__item_type__gst_percentage','quantity_result'))
        # # # # #print(orderDetails)
        context.update({
            'status': 200,
            'message': "Job Order Details Fetched Successfully.",
            'job_order_header_detail': header_detail,
            'job_order_details': orderDetails,
        })
    else:
        context.update({
            'status': 594,
            'message': "Please Provide Header Id.",
        })
    return JsonResponse(context)

#for material issue --- developed by saswata


# @api_view(['GET','POST'])
# @permission_classes([IsAuthenticated])
# def materialIssueDetails(request):
#     context = {}
#     id = request.GET.get('id', None)
#     find_all = request.GET.get('find_all', None)
#     keyword = request.GET.get('keyword', None)
#     job_Order_header_id = request.GET.get('job_Order_id',None)
#     store_id = request.GET.get('store_id',None)
#     header_detail_res = list(models.Job_Order_Detail.objects.filter(job_order_header=job_Order_header_id).values('pk','item_id','item__name','job_order_header__vendor__name','job_order_header__vendor_id'))
#     job_order_head = list(models.Job_Order.objects.filter(pk=job_Order_header_id).values('pk','vendor_id','vendor__name'))
#     context.update({
#         'status': 200,
#         'job_order_head': job_order_head,
#         'page_items': header_detail_res
#
#     })
#
#     return JsonResponse(context)


@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def getActualQuantity(request):
    context = {}
    item_id = request.GET.get('item_id',None)
    store_id = request.GET.get('store_id',None)
    transaction_date =  request.GET.get('transaction_date',None)
    #print(item_id,store_id,transaction_date)
    item =  models.Item.objects.get( pk=int(item_id))
    
    try:
        actual_quantity = models.Store_Item.objects.filter(store_id=int(store_id), item_id=int(item_id))
        #print( actual_quantity.first().on_hand_qty,6628)
        actual_quantity_On_that_date = models.Store_Item_Current.objects.filter(store_id=int(store_id), item_id=int(item_id),transaction_date=transaction_date,status=1, deleted=0).order_by('transaction_date','created_at').last()
        if not actual_quantity_On_that_date :
            
            actual_quantity_On_that_date = (
                    models.Store_Item_Current.objects.filter(
                    store_id=int(store_id),
                    item_id=int(item_id),
                    transaction_date__lt=transaction_date ,status=1, deleted=0 # Only dates less than the given date
                )
                .order_by('transaction_date','created_at')  # Order by transaction_date in descending order
                .last()  # Fetch the first (most recent) entry
            )
       
        context.update({
            'status': 200,
            'actual_quantity': actual_quantity.first().on_hand_qty if actual_quantity else 0,
            'actual_quantity_On_that_date' : actual_quantity_On_that_date.closing_qty if actual_quantity_On_that_date else 0,
            'item_price' : item.price
        })
        #print(context)
    except:
        context.update({
            'status': 200,
            'on_hand_qty_res': '0.00',
            'item_price' : item.price
        })
    
    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def materialIssueDetails(request):
    context = {}
    job_order_id = request.GET.get('job_order_id', None)
    if job_order_id is not None and job_order_id != "" :
        if models.Store_Transaction.objects.filter(job_order_id=job_order_id, transaction_type__name="MIS").exists():
            material_issue = list(
                models.Store_Transaction.objects.filter(job_order_id=job_order_id, transaction_type__name="MIS")[:1].values(
                'id', 'vendor_id',
                'vendor__name',
                'transaction_type_id',
                'transaction_type__name',
                'transaction_number',
                'transaction_date',
                'job_order_id',
                'job_order__order_number'
                )
            )
            material_issue_details = list(
                models.Store_Transaction_Detail.objects.filter(
                    store_transaction_header_id=models.Store_Transaction.objects.get(job_order_id=job_order_id, transaction_type__name="MIS").id
                ).values('pk',
                         'store_transaction_header_id',
                         'store_id',
                         'store__name',
                         'item_id',
                         'item__name',
                         'quantity',
                         ))
        else:
            material_issue=material_issue_details=None

        context.update({
            'status': 200,
            'message': "Material Issue Details Fetched Successfully.",
            'material_issue': material_issue,
            'material_issue_details': material_issue_details,
        })
    else:
        context.update({
            'status': 594,
            'message': "Please Provide an Id.",
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialIssueAdd(request):
    context = {}
    message = 'Something Went Wrong. Please Try Again.'
    userId = request.COOKIES.get('userId', None)
    if not request.POST['job_order_id'] or not request.POST['issue_date'] or not request.POST['store_id']:
        context.update({
            'status': 594,
            'message': "Job Order/Issue Date/Store has not been provided."
        })
        return JsonResponse(context)

    try:
        with transaction.atomic():
            job_order_income_detalis = list(models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['job_order_id'] , direction = 'incoming' ))
            #vendor store exist  for third party stock add 
            # # # # #print(5620)
            if request.POST['vendor_id'] :
                venStoreExist = models.Store.objects.filter(vendor_id =request.POST['vendor_id']).exists()
                if((len(job_order_income_detalis) > 0) and (venStoreExist is False)):
                    message = 'no store present for third party please create a store for third party'
                    context.update({
                    'status': 595,
                    'message': message

                    })
                    transaction.rollback()
                    return JsonResponse(context)
            #for out going
            vendor_store = ''
            store_transaction_count = models.Store_Transaction.objects.all().count()
            storeTransactionHeader=models.Store_Transaction()
            if request.POST['vendor_id']:
                vendor_store=models.Store.objects.get(vendor_id=request.POST['vendor_id'])
                storeTransactionHeader.vendor_id = request.POST['vendor_id']
            
            # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'MIS')
            transaction_type = models.Transaction_Type.objects.get(name='MIS')
            storeTransactionHeader.transaction_type = transaction_type
            storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                "${CURRENT_YEAR}", current_year).replace(
                "${AI_DIGIT_5}", ai_digit_5()).replace(
                "${transaction_type_id}", str(transaction_type.id).zfill(2))
            
            storeTransactionHeader.transaction_date=request.POST['issue_date']
            storeTransactionHeader.job_order_id = request.POST['job_order_id']
            storeTransactionHeader.total_amount = request.POST['total_amount']
            storeTransactionHeader.creator_id = userId
            if request.POST['vehicle']!="" and request.POST.get('vehicle',None):
                storeTransactionHeader.vehicle = request.POST['vehicle']
            storeTransactionHeader.save()
            
            # job status from job order
            jobOrderEdits = models.Job_Order.objects.get(pk = request.POST['job_order_id'])

            # for incoming material virtual transaction
            thirdPartyInQuantity = 0.00
            itemInThrdParty = ''
            # # #print(5662)
            if request.POST['vendor_id'] and len(job_order_income_detalis) > 0 :

                # store transaction of virtual incomming material on thrid party stock

                store_transaction_count = models.Store_Transaction.objects.all().count()
                storeTransactionHeaderIn=models.Store_Transaction()
                
                storeTransactionHeaderIn.vendor_id = request.POST['vendor_id']
                
                # # # # # #print(models.Transaction_Type.objects.get(name = 'MIS'))
                # storeTransactionHeaderIn.transaction_type = models.Transaction_Type.objects.get(name = 'GRNT')
                transaction_type = models.Transaction_Type.objects.get(name='GRNT')
                storeTransactionHeaderIn.transaction_type = transaction_type
                storeTransactionHeaderIn.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                
                storeTransactionHeaderIn.transaction_date=request.POST['issue_date']
                storeTransactionHeaderIn.job_order_id = request.POST['job_order_id']
                
                storeTransactionHeaderIn.save()

                store_transaction_details = []
                store_items_add=[]
                # #print(job_order_income_detalis)
               # incomming material will be assigned only one time no parial
                for index in range(0, len(job_order_income_detalis)):
                    # # # # #print(jobOrderEdits.job_status )
                    if job_order_income_detalis[index].required_quantity!=0 and (jobOrderEdits.job_status == 0):
                        # # # # #print(5517)
                        thirdPartyInQuantity = float(job_order_income_detalis[index].quantity)
                        itemInThrdParty = job_order_income_detalis[index].item_id
                        # # # # # #print(job_order_income_detalis[index].item.price)
                        job_order_income_detalis[index].required_quantity=float(job_order_income_detalis[index].required_quantity)-thirdPartyInQuantity
                        try:
                            job_order_income_det = models.Job_Order_Detail.objects.filter(job_order_header_id=request.POST['job_order_id'],item_id = itemInThrdParty,direction ='incoming').first() 
                            job_order_income_det.required_quantity = job_order_income_detalis[index].required_quantity
                            job_order_income_det.save()
                        except:
                            pass
                        store_transaction_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header=storeTransactionHeaderIn,
                                item_id=itemInThrdParty,
                                store=vendor_store,
                                quantity=Decimal(thirdPartyInQuantity),
                                rate = float(job_order_income_detalis[index].item.price),
                                amount =thirdPartyInQuantity * float(job_order_income_detalis[index].item.price)
                            )
                        )

                   
                        # virual added incomming material in thrid party stock
                        if models.Store_Item.objects.filter(store=vendor_store, item_id=job_order_income_detalis[index].item_id).exists():
                            # # # # #print(5538)
                            store_item=models.Store_Item.objects.get(store=vendor_store, item_id=job_order_income_detalis[index].item_id)
                            store_item.on_hand_qty+=Decimal(thirdPartyInQuantity)
                            store_item.closing_qty+= Decimal(thirdPartyInQuantity)
                            store_item.updated_at = datetime.now()
                            store_item.save()

                        # If the item does not exist in vendor store so new store item is being created
                        else:
                            # # # # #print(5547)
                            store_items_add.append(
                                models.Store_Item(
                                    store=vendor_store,
                                    item_id=int(job_order_income_detalis[index].item_id),
                                    opening_qty=float(thirdPartyInQuantity),
                                    on_hand_qty=float(thirdPartyInQuantity),
                                    closing_qty=float(thirdPartyInQuantity),
                                )
                            )
                        # change in storeItemCurrent min
                        # Fetch the last transaction_date less than the given_date
                        given_date =request.POST['issue_date']
                        
                        
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=itemInThrdParty, store_id=vendor_store.id,status=1, deleted=0).last()

                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date,item_id=itemInThrdParty, store_id=vendor_store.id,status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date,item_id=itemInThrdParty, store_id=vendor_store.id,status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                        
                        store_item_instance = models.Store_Item_Current()

                        if record:
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty + Decimal(thirdPartyInQuantity)
                            store_item_instance.closing_qty = record.closing_qty + Decimal(thirdPartyInQuantity)
                        else:
                            # Set values based on the current transaction if no prior record exists
                            store_item_instance.opening_qty = Decimal(0.00)
                            store_item_instance.on_hand_qty = Decimal(thirdPartyInQuantity)
                            store_item_instance.closing_qty = Decimal(thirdPartyInQuantity)
                        if(store_item_instance.on_hand_qty<0):
                             raise ValueError(f"ohnand quantity is less than 0")
                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionHeaderIn.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = itemInThrdParty
                        store_item_instance.store_id = vendor_store.id

                        # Save the instance to the database
                        store_item_instance.save()
                        
                        store_item_curreEdit(vendor_store.id,itemInThrdParty,given_date,'min', thirdPartyInQuantity) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                        print(itemInThrdParty,vendor_store.name,thirdPartyInQuantity)
                        
                   
                 #material issue issued for job order
                
                if (store_transaction_details or store_items_add) and (job_order_income_detalis[index].required_quantity!=0 and (jobOrderEdits.job_status == 0)) :
                    models.Store_Transaction_Detail.objects.bulk_create(store_transaction_details)
                    models.Store_Item.objects.bulk_create(store_items_add)
                else :
                    raise ValueError('error comes on 7082')
            

            #job satatus change
            jobOrderDetailsExist = models.Job_Order_Detail.objects.filter(job_order_header_id = request.POST['job_order_id'],direction='outgoing',required_quantity = 0.00).exists()
            # # #print(6223)
            jobOrderEdits.job_status = 1  # 0 for 'active' 1 for in 'pogress' 2 for in 'complete'
            jobOrderEdits.updated_at = datetime.now()
            jobOrderEdits.save()

            store_transaction_details = []
            store_items_add=[]
            outgoing_incomming_details = []
            # material issue for godown
            all_material_issued=True
            # # #print(5755)
            for index, elem in enumerate(request.POST.getlist('item_id')):
                sentQuantity = float(request.POST.getlist('quantity_sent')[index])
                requiredQuantity = float(request.POST.getlist('required_quantity')[index])
                if sentQuantity!=requiredQuantity:
                    all_material_issued=False
                try:
                    joDetail = models.Job_Order_Detail.objects.get(job_order_header_id=request.POST['job_order_id'], item_id=elem, direction="outgoing")
                    joDetail.required_quantity=requiredQuantity-sentQuantity
                    joDetail.save()
                except:
                    pass
                # # #print(5767)
                store_transaction_details.append(
                    models.Store_Transaction_Detail(
                        store_transaction_header=storeTransactionHeader,
                        item_id=elem,
                        store_id=request.POST['store_id'],
                        quantity=sentQuantity,
                        rate = float(request.POST.getlist('rate')[index]),
                        amount = float(request.POST.getlist('amount')[index])
                    )
                )
                # #print(5778)
                if request.POST['vendor_id']:

                    # third party grn transaction for out going material
                    outgoing_incomming_details.append(
                        models.Store_Transaction_Detail(
                            store_transaction_header=storeTransactionHeaderIn,
                            item_id=elem,
                            store=vendor_store,
                            quantity=sentQuantity,
                            rate = float(request.POST.getlist('rate')[index]),
                            amount = float(request.POST.getlist('amount')[index])
                        )
                    )
                    # If the item exists in vendor store
                    # # # # #print(5793)
                    if models.Store_Item.objects.filter(store=vendor_store, item_id=elem).exists():
                        store_item=models.Store_Item.objects.get(store=vendor_store, item_id=elem)
                        store_item.on_hand_qty+=Decimal(request.POST.getlist('quantity_sent')[index])
                        store_item.closing_qty+= Decimal(request.POST.getlist('quantity_sent')[index])
                        store_item.updated_at = datetime.now()
                        store_item.save()
                    # If the item does not exist in vendor store so new store item is being created
                    else:
                        store_items_add.append(
                            models.Store_Item(
                                store=vendor_store,
                                item_id=int(elem),
                                opening_qty=float(request.POST.getlist('quantity_sent')[index]),
                                on_hand_qty=float(request.POST.getlist('quantity_sent')[index]),
                                closing_qty=float(request.POST.getlist('quantity_sent')[index]),
                            )
                        )

                    # change in storeItemCurrent min
                    # Fetch the last transaction_date less than the given_date
                    given_date =request.POST['issue_date']
                   
                    # Check for the last record on the given_date
                    record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=elem, store_id=vendor_store.id,status=1, deleted=0).last()

                    if not record:
                        # If no record is found for the given_date, look for the last record before that date
                        last_transaction_date = models.Store_Item_Current.objects.filter(
                            transaction_date__lt=given_date,item_id=elem, store_id=vendor_store.id,status=1, deleted=0
                        ).aggregate(Max('transaction_date'))['transaction_date__max']

                        

                        if last_transaction_date:
                            # Fetch the record for the last_transaction_date
                            record = models.Store_Item_Current.objects.filter(
                                transaction_date=last_transaction_date,item_id=elem, store_id=vendor_store.id,status=1, deleted=0
                            ).last()

                    # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                    
                    
                    store_item_instance = models.Store_Item_Current()

                    if record:
                        # Set values based on the last record found
                        store_item_instance.opening_qty = record.closing_qty
                        store_item_instance.on_hand_qty = record.closing_qty + Decimal(
                            request.POST.getlist('quantity_sent')[index]
                        )
                        store_item_instance.closing_qty = record.closing_qty + Decimal(
                           request.POST.getlist('quantity_sent')[index]
                        )
                    else:
                        # Set values based on the current transaction if no prior record exists
                        store_item_instance.opening_qty = Decimal(
                            0.00
                        )
                        store_item_instance.on_hand_qty = Decimal(
                           request.POST.getlist('quantity_sent')[index]
                        )
                        store_item_instance.closing_qty = Decimal(
                            request.POST.getlist('quantity_sent')[index]
                        )
                    if(store_item_instance.on_hand_qty<0):
                        raise ValueError(f"onhand quantity is less than 0")
                    # Set other fields for the new transaction
                    store_item_instance.store_transaction_id = storeTransactionHeaderIn.id
                    store_item_instance.transaction_date = given_date
                    store_item_instance.item_id = elem
                    store_item_instance.store_id = vendor_store.id

                    # Save the instance to the database
                    store_item_instance.save()
                    
                    store_item_curreEdit(vendor_store.id,elem,given_date,'min', request.POST.getlist('quantity_sent')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

                # # # # #print(5811)    
                # In house store items being reduced
                # # # # # #print(models.Store.objects.filter(id=request.POST['store_id']).exists())
                if models.Store.objects.filter(id=request.POST['store_id']).exists():
                    in_house_store=models.Store.objects.get(id=request.POST['store_id'])
                    # # # # #print('AAAAA',in_house_store.id,elem)
                    # # # # #print(models.Store_Item.objects.filter(store_id=in_house_store.id, item_id=elem).exists())
                    if  models.Store_Item.objects.filter(store_id=in_house_store.id,item_id=elem).exists():
                        # # # # #print(5818)
                        store_item = models.Store_Item.objects.get(store_id=in_house_store.id, item_id=elem)
                        # # # # #print(5815)
                        store_item.on_hand_qty -= Decimal(request.POST.getlist('quantity_sent')[index])
                        store_item.closing_qty -= Decimal(request.POST.getlist('quantity_sent')[index])
                        store_item.updated_at = datetime.now()
                        store_item.save()
                    # change in storeItemCurrent mout
                    # Fetch the last transaction_date less than the given_date if mout
                    given_date =request.POST['issue_date']
                    
                    
                    # Check for the last record on the given_date
                    record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=elem, store=in_house_store.id,status=1, deleted=0 ).last()

                    if not record:
                        # If no record is found for the given_date, look for the last record before that date
                        last_transaction_date = models.Store_Item_Current.objects.filter(
                            transaction_date__lt=given_date ,item_id=elem, store_id=in_house_store.id,status=1, deleted=0
                        ).aggregate(Max('transaction_date'))['transaction_date__max']

                        

                        if last_transaction_date:
                            # Fetch the record for the last_transaction_date
                            record = models.Store_Item_Current.objects.filter(
                                transaction_date=last_transaction_date ,item_id=elem, store_id=in_house_store.id,status=1, deleted=0
                            ).last()

                    # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                    
                    
                    store_item_instance = models.Store_Item_Current()

                    if record:
                        # Set values based on the last record found
                        store_item_instance.opening_qty = record.closing_qty
                        store_item_instance.on_hand_qty = record.closing_qty - Decimal(request.POST.getlist('quantity_sent')[index])
                        store_item_instance.closing_qty = record.closing_qty -Decimal(request.POST.getlist('quantity_sent')[index])
                    
                        if(store_item_instance.on_hand_qty<0):
                            raise ValueError(f"out quantity is more than available quantity")
                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionHeader.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = elem
                        store_item_instance.store_id = in_house_store.id

                        # Save the instance to the database
                        store_item_instance.save()
                        
                    else:
                        message = "canot possible item found in this strore"
                        raise ValueError(message)
                    store_item_curreEdit(in_house_store.id,elem,given_date,'mout', request.POST.getlist('quantity_sent')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
                    # # # # #print(5825)
            # # #print(5818)
            jobOrderEditsdetails = not models.Job_Order_Detail.objects.filter(direction='outgoing',
            job_order_header_id=jobOrderEdits.id).exclude(required_quantity=0.00).exists()
            # # #print(6308,'AAAAAAAa')
            all_material_issued = True if jobOrderEditsdetails else False
            
            if all_material_issued:
                jobOrderEdits.material_issue = 3
            else:
                # #print(6314,"bbbbbbbb")
                jobOrderEditsdetails = models.Job_Order_Detail.objects.filter(
                    direction='outgoing',
                    job_order_header_id=jobOrderEdits.id,
                    quantity=F('required_quantity')
                ).exists()

                # #print(jobOrderEditsdetails)
                jobOrderEdits.material_issue = 1 if jobOrderEditsdetails else 2
            jobOrderEdits.save()
              
            models.Store_Transaction_Detail.objects.bulk_create(store_transaction_details)
            models.Store_Transaction_Detail.objects.bulk_create(outgoing_incomming_details)
            models.Store_Item.objects.bulk_create(store_items_add)
            
            user_log_details_add(userId,'Material Issue add')
        transaction.commit()

        context.update({
            'status': 200,
            'message': "Material Issue Created Successfully."
        })

    except Exception as e:
        print(e)
        context.update({
            'status': 595,
            'message': message
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialIssueEdit(request):
    context = {}
    userId = request.COOKIES.get('userId', None)
    if not request.POST['issue_date']:
        context.update({
            'status': 596,
            'message': "Issue Date has not been provided."
        })
        return JsonResponse(context)
    try:
        with transaction.atomic():
            vendor_id = request.POST.get('vendor_id',None)
            item_id = request.POST.getlist('item_id')
            issue_date=request.POST['issue_date']
            store_transaction_id = request.POST['id']
            storeTransactionHeader=models.Store_Transaction.objects.get(pk = store_transaction_id)
            storeTransactionHeader.total_amount = request.POST['total_amount']
            if request.POST['vehicle']!="" and request.POST.get('vehicle',None):
                storeTransactionHeader.vehicle = request.POST['vehicle']
            storeTransactionHeader.creator_id = userId
            storeTransactionHeader.save()

            for index in range(0,len(item_id)):
                if(vendor_id):
                    store_item_vendor_update = models.Store_Item.objects.get(store_id = request.POST['store_id'] , item_id=item_id[index])
                    store_item_vendor_update.on_hand_qty = (float(store_item_vendor_update.on_hand_qty)- float(request.POST.getlist('quantity_sent_og')[0])) +float(request.POST.getlist('quantity_sent')[0])
                    store_item_vendor_update.closing_qty =(float(store_item_vendor_update.closing_qty)- float(request.POST.getlist('quantity_sent_og')[0])) + float(request.POST.getlist('quantity_sent')[0])
                    store_item_vendor_update.updated_at =  datetime.now()
                    store_item_vendor_update.save()

                store_transaction_deat_update = models.Store_Transaction_Detail.objects.get(store_transaction_header_id=store_transaction_id,item_id= item_id[index])
                store_transaction_deat_update.quantity = request.POST.getlist('quantity_sent')[index]
                store_transaction_deat_update.rate = request.POST.getlist('rate')[index]
                store_transaction_deat_update.amount = request.POST.getlist('amount')[index]
                store_transaction_deat_update.updated_at = datetime.now()
                store_transaction_deat_update.save()

                store_item_update = models.Store_Item.objects.get(store_id = request.POST['store_id'] , item_id=item_id[index])
                store_item_update.on_hand_qty = (float(store_item_update.on_hand_qty)+ float(request.POST.getlist('quantity_sent_og')[0])) - float(request.POST.getlist('quantity_sent')[0])
                store_item_update.closing_qty =(float(store_item_update.closing_qty)+ float(request.POST.getlist('quantity_sent_og')[0])) - float(request.POST.getlist('quantity_sent')[0])
                store_item_update.updated_at =  datetime.now()
                store_item_update.save()
            
            user_log_details_add(userId,'Material Issue Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material Issue Edited Successfully."
        })

    except Exception:
        context.update({
            'status': 597,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)

#for grn inspection--- developed by saswata

#grnInspection Header List ----developed by saswata
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def grnInspectionHeaderList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    ins_completed = request.GET.get('ins_completed',None)

    try:
        if id is not None and id != "":
            grnInspection = list(models.Grn_Inspection_Transaction.objects.filter(pk=id ,ins_done = 1)[:1].values(
                'pk', 'vendor_id', 'vendor__name', 'transaction_number'))
            context.update({
                'status': 200,
                'message': "grnInspection header Fetched Successfully.",
                'page_items': grnInspection,
            })
        else:
            if keyword is not None and keyword != "":
                # # # # # #print("4244",request.GET)
                grnInspection = list(
                    models.Grn_Inspection_Transaction.objects.filter(
                        Q(vendor__name__icontains=keyword) | Q(transaction_number__icontains=keyword) 
                    ).filter(
                        status=1, deleted=0 ,ins_done = 1).values('pk', 'vendor_id', 'vendor__name', 'transaction_number')
                )
            elif ins_completed is not None and ins_completed != "":
                grnInspection = list(models.Grn_Inspection_Transaction.objects.filter(status=1, deleted=0 ,ins_completed = 0).values(
                    'pk', 'vendor_id', 'vendor__name', 'transaction_number'))
                context.update({
                    'status': 200,
                    'message': "Store Items Fetched Successfully.",
                    'page_items': grnInspection,
                })
                return JsonResponse(context)

            else:
                grnInspection = list(models.Grn_Inspection_Transaction.objects.filter(status=1, deleted=0 ,ins_done = 1).values(
                    'pk', 'vendor_id', 'vendor__name', 'transaction_number'))
            if find_all is not None and int(find_all) == 1:
                context.update({
                    'status': 200,
                    'message': "Store Items Fetched Successfully.",
                    'page_items': grnInspection,
                })
                return JsonResponse(context)

            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)

            paginator = CustomPaginator(grnInspection, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()
            context.update({
                'status': 200,
                'message': "grn Inspection Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })
       
    except:
        context ={
            'status':597,
            'message':'server error'
        }
    return JsonResponse(context)


#GRN Inspection Transaction Details List ----developed by saswata
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getGrnInspectionTransactionDetail(request):

    try:
        # Getting grn inspection transaction details whose inspection is not done
        if int(request.GET.get('ins_done'))==0:
            grn_Ins_Det = list(models.Grn_Inspection_Transaction_Detail.objects.filter(
                    grn_inspection_transaction_header_id = int(request.GET.get('insId')), ins_done = 0
                ).values(
                    'pk',
                    'grn_inspection_transaction_header_id',
                    'grn_inspection_transaction_header__vendor_id',
                    'grn_inspection_transaction_header__vendor__name',
                    'grn_inspection_transaction_header__purchase_order_header_id',
                    'grn_inspection_transaction_header__purchase_order_header__order_number',
                    'grn_inspection_transaction_header__job_order_id',
                    'grn_inspection_transaction_header__job_order__order_number',
                    'grn_inspection_transaction_header__invoice_challan',
                    'item_id',
                    'item__name',
                    'store_id',
                    'store__name',
                    'quantity',
                    'rate',
                    'amount',
                    'gst_percentage'
                ))
        # Getting grn inspection transaction details whose inspection is done
        if int(request.GET.get('ins_done')) == 1:
            grn_Ins_Det = list(models.Grn_Inspection_Transaction_Detail.objects.filter(
                grn_inspection_transaction_header_id=int(request.GET.get('insId')), ins_done=1,
            ).exclude(reject_quantity=0).values(
                'pk',
                'grn_inspection_transaction_header_id',
                'grn_inspection_transaction_header__vendor_id',
                'grn_inspection_transaction_header__vendor__name',
                'grn_inspection_transaction_header__purchase_order_header_id',
                'grn_inspection_transaction_header__purchase_order_header__order_number',
                'grn_inspection_transaction_header__job_order_id',
                'grn_inspection_transaction_header__job_order__order_number',
                'grn_inspection_transaction_header__invoice_challan',
                'item_id',
                'item__name',
                'store_id',
                'store__name',
                'quantity',
                'rate',
                'amount',
                'reject_quantity',
                'quantity',
                'gst_percentage'
            ))
        context ={
            'status':200,
            'page_items': grn_Ins_Det
        }
    except:
        context ={
        'status':598,
        'message':'server error1'
        }


    return JsonResponse(context)

#grnInspection add and update  ----developed by saswata

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def addGrnDetailisInsTransaction(request):
    context = {}
    userId = request.COOKIES.get('userId', None)
    ins_completed = 0 if not all(request.POST.getlist('accp_quantity')) else 1 #if ins_completed is 0 means all item is not inspected may be inspection happened paritally
    try:
        if any(request.POST.getlist('accp_quantity')):
            with transaction.atomic():
                grn_ins_head = models.Grn_Inspection_Transaction.objects.get(pk = request.POST['insTraId'])
                grn_ins_head.invoice_challan = request.POST['invoice_challan']
                grn_ins_head.ins_done = 1
                grn_ins_head.ins_completed = ins_completed
                grn_ins_head.updated_at = datetime.now()
                grn_ins_head.save()
                # # # # #print('5179')
                storeTranscHeadPresent = models.Store_Transaction.objects.filter(grn_inspection_id = request.POST['insTraId']).first()
                if storeTranscHeadPresent is not None:
                    storeTransactionHeader = storeTranscHeadPresent
                    storeTransactionHeader.total_amount = float(storeTransactionHeader.total_amount) + float (request.POST['totalPrice'])
                    storeTransactionHeader.notes=grn_ins_head.notes
                    storeTransactionHeader.updated_at = datetime.now()
                    storeTransactionHeader.save()
                else:
                    # # # # #print(6073)
                    store_transaction_count = models.Store_Transaction.objects.all().count()
                    storeTransactionHeader = models.Store_Transaction()
                    storeTransactionHeader.vendor_id = request.POST['vendor_id']
                    storeTransactionHeader.creator_id = userId
                    # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'GRN')
                    transaction_type = models.Transaction_Type.objects.get(name='GRN')
                    storeTransactionHeader.transaction_type = transaction_type
                    storeTransactionHeader.invoice_challan = request.POST['invoice_challan']
                    if (request.POST.get('purchase_order_header_id',None) and request.POST['purchase_order_header_id']!=""):
                        storeTransactionHeader.purchase_order_header_id = request.POST[
                            'purchase_order_header_id']
                    if(request.POST.get('job_order_header_id',None) and request.POST['job_order_header_id']!=""):
                        storeTransactionHeader.job_order_id = request.POST[
                            'job_order_header_id']
                    storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                        "${CURRENT_YEAR}", current_year).replace(
                        "${AI_DIGIT_5}", ai_digit_5()).replace(
                        "${transaction_type_id}", str(transaction_type.id).zfill(2))
                    storeTransactionHeader.transaction_date = request.POST['issue_date']
                    storeTransactionHeader.total_amount = request.POST['totalPrice']
                    storeTransactionHeader.grn_inspection_id = request.POST['insTraId']
                    storeTransactionHeader.notes = grn_ins_head.notes
                    storeTransactionHeader.save()
                # # # # #print(6091)
                order_details =[]
                material_reciept_all = 0
                for index in  range(0,len(request.POST.getlist('accp_quantity'))):
                    if request.POST.getlist('accp_quantity')[index] != '':
                        grn_ins_det = models.Grn_Inspection_Transaction_Detail.objects.get(pk =request.POST.getlist('ins_det_id')[index])
                        grn_ins_det.ins_done = 1
                        grn_ins_det.accepted_quantity  = request.POST.getlist('accp_quantity')[index] 
                        grn_ins_det.reject_quantity = request.POST.getlist('rej_quantity')[index] 
                        grn_ins_det.inspection_date = request.POST['issue_date']
                        grn_ins_det.amount = request.POST.getlist('amount')[index]
                        grn_ins_det.amount_with_gst = request.POST.getlist(
                                    'actualPrice')[index]
                        grn_ins_det.updated_at = datetime.now()
                        grn_ins_det.save()
                        # # # # # #print('5219')
                        order_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header_id=storeTransactionHeader.id,
                                item_id=request.POST.getlist('item_id')[index],
                                store_id=request.POST.getlist('store_id')[index],
                                quantity=request.POST.getlist('accp_quantity')[index],
                                rate=request.POST.getlist('rate')[index],
                                amount=request.POST.getlist('amount')[index],
                                gst_percentage=request.POST.getlist('gst_percentage')[index],
                                amount_with_gst=request.POST.getlist('actualPrice')[index]
                                
                            )
                        )
                        
                        storeItem = models.Store_Item.objects.filter(
                            item_id=request.POST.getlist('item_id')[index], store_id=request.POST.getlist('store_id')[index]).first()
                        # # # # # #print('5235')
                        if storeItem is None:
                            storeItem = models.Store_Item()
                            storeItem.opening_qty = Decimal(request.POST.getlist('accp_quantity')[index])
                            storeItem.on_hand_qty = Decimal(request.POST.getlist('accp_quantity')[index])
                            storeItem.closing_qty = Decimal(request.POST.getlist('accp_quantity')[index])
                            storeItem.item_id = request.POST.getlist('item_id')[index]
                            storeItem.store_id = request.POST.getlist('store_id')[index]
                            storeItem.save()
                        else:
                            storeItem.on_hand_qty += Decimal(request.POST.getlist('accp_quantity')[index])
                            storeItem.closing_qty += Decimal(request.POST.getlist('accp_quantity')[index])
                            storeItem.updated_at = datetime.now()
                            storeItem.save()

                        # change in storeItemCurrent min
                        # Fetch the last transaction_date less than the given_date
                        given_date =  request.POST['issue_date']
                        
                        
                        # Check for the last record on the given_date
                        record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=request.POST.getlist('item_id')[index], store_id=request.POST.getlist('store_id')[index],status=1, deleted=0).last()

                        if not record:
                            # If no record is found for the given_date, look for the last record before that date
                            last_transaction_date = models.Store_Item_Current.objects.filter(
                                transaction_date__lt=given_date,item_id=request.POST.getlist('item_id')[index], store_id=request.POST.getlist('store_id')[index],status=1, deleted=0
                            ).aggregate(Max('transaction_date'))['transaction_date__max']

                            

                            if last_transaction_date:
                                # Fetch the record for the last_transaction_date
                                record = models.Store_Item_Current.objects.filter(
                                    transaction_date=last_transaction_date,item_id=request.POST.getlist('item_id')[index], store_id=request.POST.getlist('store_id')[index],status=1, deleted=0
                                ).last()

                        # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                        
                        
                        store_item_instance = models.Store_Item_Current()

                        if record:
                            # Set values based on the last record found
                            store_item_instance.opening_qty = record.closing_qty
                            store_item_instance.on_hand_qty = record.closing_qty + Decimal(
                                request.POST.getlist('accp_quantity')[index]
                            )
                            store_item_instance.closing_qty = record.closing_qty + Decimal(
                               request.POST.getlist('accp_quantity')[index]
                            )
                        else:
                            # Set values based on the current transaction if no prior record exists
                            store_item_instance.opening_qty = Decimal(0.00)
                            store_item_instance.on_hand_qty = Decimal(
                                request.POST.getlist('accp_quantity')[index]
                            )
                            store_item_instance.closing_qty = Decimal(
                                request.POST.getlist('accp_quantity')[index]
                            )
                        if(store_item_instance.on_hand_qty<0):
                            raise ValueError(f"onhand quantity is less than 0")
                        # Set other fields for the new transaction
                        store_item_instance.store_transaction_id = storeTransactionHeader.id
                        store_item_instance.transaction_date = given_date
                        store_item_instance.item_id = request.POST.getlist('item_id')[index]
                        store_item_instance.store_id = request.POST.getlist('store_id')[index]

                        # Save the instance to the database
                        store_item_instance.save()
                        
                        store_item_curreEdit(request.POST.getlist('store_id')[index], request.POST.getlist('item_id')[index],given_date,'min', request.POST.getlist('accp_quantity')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)


                        # # # # # #print('5249')
                        if (request.POST.get('job_order_header_id',None) and request.POST['job_order_header_id']!=""):
                            # # # # # #print('5251')
                            job_order_details= models.Job_Order_Detail.objects.filter(item_id=request.POST.getlist('item_id')[index], 
                                                job_order_header_id=request.POST['job_order_header_id']).first()
                            # # # # # #print(job_order_details.quantity_result)
                            job_order_details.quantity_result += Decimal(request.POST.getlist('rej_quantity')[index] )
                            job_order_details.updated_at = datetime.now()
                            job_order_details.save()
                            store = models.Store.objects.get(vendor_id=request.POST['vendor_id'])
                            storeItem = models.Store_Item.objects.filter(
                            item_id=request.POST.getlist('item_id')[index], store =store ).first()
                            storeItem.on_hand_qty += Decimal(request.POST.getlist('rej_quantity')[index])
                            storeItem.closing_qty += Decimal(request.POST.getlist('rej_quantity')[index])
                            storeItem.updated_at = datetime.now()
                            storeItem.save()
                            # # # # # #print('5257')
                            material_reciept_all = 0 if float(job_order_details.quantity_result)>0.00 else 1 

                            #closed or re open job order
                            if (material_reciept_all == 1):
                                # # # # # #print('5871')
                                job_order = models.Job_Order.objects.filter(pk=request.POST['job_order_header_id']).get() 
                                # # # # # #print(job_order)
                                job_order.material_reciept = 1
                                job_order.job_status = 2
                                current_time = datetime.now(timezone.utc)
                                time_difference = current_time - job_order.created_at
                                total_hours = time_difference.days * 24 + time_difference.seconds / 3600
                                job_order.actual_time_take = str(round(total_hours, 3)) + 'hr'
                                job_order.updated_at = datetime.now()  
                            else :
                                # # # # # #print('5877')
                                job_order = models.Job_Order.objects.filter(pk=request.POST['job_order_header_id']).get() 
                                job_order.material_reciept = 0
                                job_order.job_status = 1
                                job_order.actual_time_take =""
                                job_order.updated_at = datetime.now()
                            job_order.save()
                        # # # # # #print('5257')
                models.Store_Transaction_Detail.objects.bulk_create(order_details)
                # # # # # #print('5879')
                if request.POST['purchase_order_header_id'] != "" and (request.POST['purchase_order_header_id']) != 0:
                    for index, elem in enumerate(request.POST.getlist('item_id')):
                        if request.POST.getlist('accp_quantity')[index] != '':
                            purchaseOrderItem = models.Purchase_Order_Detail.objects.get(
                                purchase_order_header_id= request.POST['purchase_order_header_id'] ,item_id = elem)
                            purchaseOrderItem.delivered_quantity += Decimal(request.POST.getlist('accp_quantity')[index])
                            purchaseOrderItem.delivered_rate = Decimal(request.POST.getlist('rate')[index])
                            purchaseOrderItem.delivered_amount += Decimal(request.POST.getlist('amount')[index])
                            purchaseOrderItem.delivered_gst_percentage = Decimal(request.POST.getlist('gst_percentage')[index])
                            purchaseOrderItem.delivered_amount_with_gst += Decimal(request.POST.getlist('actualPrice')[index])
                            purchaseOrderItem.updated_at = datetime.now()
                            purchaseOrderItem.save()
                    purchaseOrderHeader = models.Purchase_Order.objects.prefetch_related(
                        'purchase_order_detail_set').get(pk=request.POST['purchase_order_header_id'])
                    flag = True
                    for purchaseOrderDetail in purchaseOrderHeader.purchase_order_detail_set.all():
                        if Decimal(purchaseOrderDetail.quantity) > Decimal(purchaseOrderDetail.delivered_quantity):
                            flag = False
                            break
                    if flag == True:
                        purchaseOrderHeader.delivery_status = 3
                    else:
                        purchaseOrderHeader.delivery_status = 2
                    purchaseOrderHeader.updated_at = datetime.now()
                    purchaseOrderHeader.save()  
               
                user_log_details_add(userId,'GRN INspection Add')
            transaction.commit()
            context ={
                'status':200,
                'message':'inspection details updated succesfully and store transaction upadated sucesfully'
            }
    except:
        context ={
            'status':599,
            'message':'server error'
        }
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialReturnAdd(request):
    context = {}
    if not request.POST['reason'] or not request.POST['return_date'] :
        context.update({
            'status': 531,
            'message': "Reason or return date has not been provided."
        })
        return JsonResponse(context)
    total_amount = 0.00
    try:
        with transaction.atomic():
            # Reason = On excess issue of items against a job
            if int(request.POST['reason'])==1:
                store_transaction_count = models.Store_Transaction.objects.all().count()
                material_issue=models.Store_Transaction.objects.get(transaction_type__name="MIS", job_order_id=request.POST['job_order_id'])
                material_return=models.Store_Transaction()
                if material_issue.vendor: material_return.vendor=material_issue.vendor
                # material_return.transaction_type=models.Transaction_Type.objects.get(name="MR")
                transaction_type = models.Transaction_Type.objects.get(name='MR')
                material_return.transaction_type = transaction_type
                material_return.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                material_return.transaction_date = request.POST['return_date']
                material_return.job_order=material_issue.job_order

                material_return.save()

                in_house_store=models.Store_Transaction_Detail.objects.filter(store_transaction_header=material_issue).first().store
                vendor_store=models.Store.objects.filter(vendor=material_issue.vendor).first()
                store_transaction_details = []

                for (item_id, previous_quantity, updated_quantity) in zip(request.POST.getlist('item_id'),request.POST.getlist('previous_quantity'),request.POST.getlist('updated_quantity')):
                    store_transaction_details.append(
                        models.Store_Transaction_Detail(
                            store_transaction_header=material_return,
                            item_id=item_id,
                            store=in_house_store,
                            quantity=Decimal(updated_quantity)-Decimal(previous_quantity)
                        )
                    )

                    vendor_store_item = models.Store_Item.objects.get(store=vendor_store, item_id=item_id)
                    vendor_store_item.on_hand_qty -= Decimal(previous_quantity)-Decimal(updated_quantity)
                    vendor_store_item.closing_qty -= Decimal(previous_quantity)-Decimal(updated_quantity)
                    vendor_store_item.updated_at = datetime.now()
                    vendor_store_item.save()

                    in_house_store_item = models.Store_Item.objects.get(store=in_house_store, item_id=item_id)
                    in_house_store_item.on_hand_qty += Decimal(previous_quantity)-Decimal(updated_quantity)
                    in_house_store_item.closing_qty += Decimal(previous_quantity)-Decimal(updated_quantity)
                    in_house_store_item.updated_at = datetime.now()
                    in_house_store_item.save()

                models.Store_Transaction_Detail.objects.bulk_create(store_transaction_details)

            # Reason = For rejected material during inspection process
            elif int(request.POST['reason'])==2:
                store_transaction_count = models.Store_Transaction.objects.all().count()
                grn_inspection_transaction_header=models.Grn_Inspection_Transaction.objects.get(id=request.POST['grn_inspection_id'])

                material_return = models.Store_Transaction()
                material_return.vendor = grn_inspection_transaction_header.vendor
                if grn_inspection_transaction_header.purchase_order_header: material_return.purchase_order_header = grn_inspection_transaction_header.purchase_order_header
                # material_return.transaction_type = models.Transaction_Type.objects.get(name="MR")
                transaction_type = models.Transaction_Type.objects.get(name='MR')
                material_return.transaction_type = transaction_type
                material_return.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                    "${CURRENT_YEAR}", current_year).replace(
                    "${AI_DIGIT_5}", ai_digit_5()).replace(
                    "${transaction_type_id}", str(transaction_type.id).zfill(2))
                material_return.transaction_date = request.POST['return_date']
                material_return.destination = request.POST['destination']
                material_return.vehicle = request.POST['Vehicle_no']
                material_return.dispatch_no = request.POST['dispatch_no']
                if request.POST.get('invoice_challan_no',None):
                    material_return.invoice_challan = request.POST['invoice_challan_no']
                if request.POST.get('note',None):
                    material_return.notes = request.POST['note']
                if request.POST.get('grn_inspection_no',None):
                    material_return.grn_inspection_id = request.POST['grn_inspection_no']
                if request.POST.get('dispatch_no',None):
                    material_return.dispatch_no = request.POST['dispatch_no']    
                material_return.save()

                store_transaction_details = []
                for i in range(len(request.POST.getlist('item_id'))):
                    if request.POST.getlist('return_item')[i]=="1":
                        store_transaction_details.append(
                            models.Store_Transaction_Detail(
                                store_transaction_header=material_return,
                                item_id=request.POST.getlist('item_id')[i],
                                store_id=request.POST.getlist('store_id')[i],
                                quantity=request.POST.getlist('reject_quantity')[i]
                            )
                        )
                        total_amount += float(request.POST.getlist('reject_quantity')[i])
                material_return.total_amount = total_amount
                material_return.save()
                models.Store_Transaction_Detail.objects.bulk_create(store_transaction_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Material Return Add')

        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material Return Created Successfully."
        })
    except Exception as e:
        context.update({
            'status': 533,
            'message': f"Something Went Wrong. Please Try Again.{e}"
        })
        transaction.rollback()
    return JsonResponse(context)

# on transit transaction ---deveoped by saswata

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getOnTransitTransactionHeadersList(request):
    context = {}
    # # # # # #print("4679",request.GET)
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    flag = request.GET.get('flag', None)
    try:
        if id is not None and id != "":
            onTransitTransactionHeader = list(models.On_Transit_Transaction.objects.filter(pk=id)[:1].values(
                'pk', 'transaction_number', 'transaction_date','transaction_in_date', 'source_store_id','source_store__name' ,'destination_store_id','destination_store__name'))
            context.update({
                'status': 200,
                'message': "onTransitTransaction header Fetched Successfully.",
                'page_items': onTransitTransactionHeader,
            })
        elif flag is not None and flag != "":
            if keyword is not None and keyword != "":
                # # # # # #print("4244",request.GET)
                onTransitTransactionHeader = list(
                    models.On_Transit_Transaction.objects.filter(
                        Q(transaction_number__icontains=keyword) 
                    ).filter(
                        status=1, deleted=0,flag=1).values('pk', 'transaction_number', 'transaction_date','transaction_in_date', 'source_store_id','source_store__name' ,'destination_store_id','destination_store__name')
                )

            else:
                # # # # # #print("4263",request.GET)   
                onTransitTransactionHeader = list(models.On_Transit_Transaction.objects.filter(status=1, deleted=0,flag=1).values(
                    'pk', 'transaction_number', 'transaction_date','transaction_in_date', 'source_store_id','source_store__name' ,'destination_store_id','destination_store__name'))
                # # # # # #print(onTransitTransactionHeader)
            if find_all is not None and int(find_all) == 1:
                context.update({
                    'status': 200,
                    'message': "Store Items Fetched Successfully.",
                    'page_items': onTransitTransactionHeader,
                })
                return JsonResponse(context)

            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)

            paginator = CustomPaginator(onTransitTransactionHeader, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()

            context.update({
                'status': 200,
                'message': "onTransitTransaction Header Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })
        else:
            if keyword is not None and keyword != "":
                onTransitTransactionHeader = list(
                    models.On_Transit_Transaction.objects.filter(
                        Q(transaction_number__icontains=keyword) 
                    ).filter(
                        status=1, deleted=0,flag=0).values('pk', 'transaction_number', 'transaction_date','transaction_in_date', 'source_store_id','source_store__name' ,'destination_store_id','destination_store__name')
                )

            else:  
                onTransitTransactionHeader = list(models.On_Transit_Transaction.objects.filter(status=1, deleted=0,flag=0).values(
                    'pk', 'transaction_number', 'transaction_date','transaction_in_date', 'source_store_id','source_store__name' ,'destination_store_id','destination_store__name'))

            if find_all is not None and int(find_all) == 1:
                context.update({
                    'status': 200,
                    'message': "Store Items Fetched Successfully.",
                    'page_items': onTransitTransactionHeader,
                })
                return JsonResponse(context)

            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)

            paginator = CustomPaginator(onTransitTransactionHeader, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()

            context.update({
                'status': 200,
                'message': "onTransitTransaction Header Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })
    except Exception:
       context.update({
            'status': 534,
            'message': "Something Went Wrong. Please Try Again."
        }) 


    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getOnTransitTransactionDetalisList(request):
    context = {}
    # # # # # #print(request.GET)
    id = request.GET.get('transactionNumber',None)
    try:
        if id is not None and id !="":
            onTransitTransactionDetails = list(models.On_Transit_Transaction_Details.objects.filter(on_transit_transaction_header_id = id).values('pk',
            'item_id',
            'item__name',
            'on_transit_transaction_header__source_store_id',
            'on_transit_transaction_header__source_store__name',
            'on_transit_transaction_header__destination_store_id',
            'on_transit_transaction_header__destination_store__name',
            'quantity',
            'amount',
            'rate'
            ))
           
            context.update({
                'onTransitTransactionDetails':onTransitTransactionDetails,
                'status': 200,
                'message': 'onTransitTransaction Details fetched sucesfully'
            })
    except Exception:
        context.update({
            'status': 535,
            'message': "Something Went Wrong. Please Try Again."
        }) 

    return JsonResponse(context)


# material out
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialOutDetailsAdd(request):
    context = {}
    # # # # # #print(request.POST['sourceStore'])
    try:
        with transaction.atomic():

            # on transit transaction for material out header save

            on_transit_transaction_count = models.On_Transit_Transaction.objects.all().count()            
            on_transit_transaction_header =  models.On_Transit_Transaction()
            on_transit_transaction_header.transaction_number =  env("MAT_TRANSFER_OUT_SEQ").replace(
                            "${CURRENT_YEAR}", datetime.today().strftime('%Y')).replace(
                            "${AI_DIGIT_5}", str(on_transit_transaction_count + 1).zfill(5))           
            on_transit_transaction_header.transaction_date =request.POST['issue_date']          
            on_transit_transaction_header.source_store_id = request.POST['sourceStore']
            on_transit_transaction_header.destination_store_id = request.POST['desStore']
            on_transit_transaction_header.vechical_no = request.POST['vehicle_no']
            on_transit_transaction_header.save()
            # store transaction header  for material out save

            store_transaction_count = models.Store_Transaction.objects.all().count()
            storeTransactionHeader = models.Store_Transaction()
            # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'MOUT')
            transaction_type = models.Transaction_Type.objects.get(name='MOUT')
            storeTransactionHeader.transaction_type = transaction_type
            storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                "${CURRENT_YEAR}", current_year).replace(
                "${AI_DIGIT_5}", ai_digit_5()).replace(
                "${transaction_type_id}", str(transaction_type.id).zfill(2))
            storeTransactionHeader.transaction_date = request.POST['issue_date']
            storeTransactionHeader.reference_id =  int(on_transit_transaction_header.id)
            storeTransactionHeader.save()

            on_transit_transaction_details = []
            order_details = []

            for index in range(0,len(request.POST.getlist('item_id'))):

                # on transit transaction detalis for material out save
                on_transit_transaction_details.append(
                    models.On_Transit_Transaction_Details(
                         on_transit_transaction_header_id = on_transit_transaction_header.id,
                         item_id = request.POST.getlist('item_id')[index],
                         quantity = request.POST.getlist('quantity_sent')[index],
                         rate = request.POST.getlist('rate')[index],
                         amount = float(request.POST.getlist('quantity_sent')[index]) * float(request.POST.getlist('rate')[index])
                    )
                )

                # on store transaction detalis for material out save
                order_details.append(
                        models.Store_Transaction_Detail(
                            store_transaction_header_id=storeTransactionHeader.id,
                            item_id=request.POST.getlist('item_id')[index],
                            store_id= request.POST['sourceStore'],
                            quantity=request.POST.getlist('quantity_sent')[index],
                            rate=request.POST.getlist('rate')[index],
                            amount=float(request.POST.getlist('quantity_sent')[index]) * float(request.POST.getlist('rate')[index]),
                            
                        )
                    )
                
                # item deducted from source store

                storeItem = models.Store_Item.objects.filter(
                    item_id=request.POST.getlist('item_id')[index], store_id=request.POST['sourceStore']).first()
                storeItem.on_hand_qty -= Decimal(request.POST.getlist('quantity_sent')[index])
                storeItem.closing_qty -= Decimal(request.POST.getlist('quantity_sent')[index])
                storeItem.updated_at = datetime.now()
                storeItem.save()
                # change in storeItemCurrent mout
                # Fetch the last transaction_date less than the given_date if mout
                given_date = request.POST['issue_date']
                
                
                # Check for the last record on the given_date
                record = models.Store_Item_Current.objects.filter(transaction_date=given_date,item_id=request.POST.getlist('item_id')[index], store_id= request.POST['sourceStore'],status=1, deleted=0).last()

                if not record:
                    # If no record is found for the given_date, look for the last record before that date
                    last_transaction_date = models.Store_Item_Current.objects.filter(
                        transaction_date__lt=given_date ,item_id=request.POST.getlist('item_id')[index], store_id= request.POST['sourceStore'],status=1, deleted=0
                    ).aggregate(Max('transaction_date'))['transaction_date__max']

                    

                    if last_transaction_date:
                        # Fetch the record for the last_transaction_date
                        record = models.Store_Item_Current.objects.filter(
                            transaction_date=last_transaction_date ,item_id=request.POST.getlist('item_id')[index], store_id= request.POST['sourceStore'],status=1, deleted=0
                        ).last()

                # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                
                
                store_item_instance = models.Store_Item_Current()

                if record:
                    # Set values based on the last record found
                    store_item_instance.opening_qty = record.closing_qty
                    store_item_instance.on_hand_qty = record.closing_qty - Decimal(request.POST.getlist('quantity_sent')[index])
                    store_item_instance.closing_qty = record.closing_qty - Decimal(request.POST.getlist('quantity_sent')[index])

                    if(store_item_instance.on_hand_qty<0):
                        raise ValueError(f"out quantity is more than available quantity")

                    # Set other fields for the new transaction
                    store_item_instance.store_transaction_id = storeTransactionHeader.id
                    store_item_instance.transaction_date = given_date
                    store_item_instance.item_id = request.POST.getlist('item_id')[index]
                    store_item_instance.store_id = request.POST['sourceStore']

                    # Save the instance to the database
                    store_item_instance.save()
                    
                else:
                    message = "canot possible"
                    raise ValueError('canot possible')
                store_item_curreEdit(request.POST['sourceStore'],request.POST.getlist('item_id')[index],given_date,'mout',request.POST.getlist('quantity_sent')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
    
            models.On_Transit_Transaction_Details.objects.bulk_create(on_transit_transaction_details)
            models.Store_Transaction_Detail.objects.bulk_create(order_details)

            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Material Out Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material out created sucessfully"
        })
    except Exception as e:
        context.update({
            'status': 536,
            'message': f"Something Went Wrong. Please Try Again.{e}"
        })
        transaction.rollback()

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialOutDetailsDelete(request):
    context = {}
    # # # # # #print(request.POST)
    try:
        materialOut = models.On_Transit_Transaction.objects.get(pk=request.POST['id'])
        materialOutDetails = list(models.On_Transit_Transaction_Details.objects.filter(on_transit_transaction_header_id = request.POST['id']).values('pk','item_id','quantity'))
        store_id = materialOut.source_store_id
        storeTransaction = models.Store_Transaction.objects.get(reference_id =request.POST['id'],transaction_type_id= 6 )
        
        with transaction.atomic():
            materialOut.delete()
            storeTransaction.delete()

            # item added to  source store
            for index in materialOutDetails:
                # # # # # #print(index['item_id'])
                storeItem = models.Store_Item.objects.filter(
                    item_id=index['item_id'], store_id=store_id).first()

                storeItem.on_hand_qty += Decimal(index['quantity'])
                storeItem.closing_qty += Decimal(index['quantity'])
                storeItem.updated_at = datetime.now()
                
                storeItem.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Material Out Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material Out transaction Deleted Sucessfully"
        })

    except Exception:
        context.update({
            'status': 537,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialOutDetailsEdit(request):
    context = {}
    # # # # # #print(request.POST)
    try:
        # # # # # #print(request.POST)
        with transaction.atomic():

            # on transit transaction for material out header save

            on_transit_transaction_header = models.On_Transit_Transaction.objects.get(pk=request.POST['pk'])
           
            on_transit_transaction_header.vechical_no = request.POST['vehicle_no']
            on_transit_transaction_header.save()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Material Out Edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material Out Edited Sucessfully"
        })
    except Exception:
        pass
        context.update({
            'status': 538,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


# @api_view(['GET'])
# def storeTransferReportExport(request):
#     # # # # # #print(request.GET)
#     from_date = request.GET.get('form_date')
#     to_date = request.GET.get('to_date')
#     file_type=request.GET.get('file_type')
#     # # # # # #print(from_date)
#     # # # # # #print("3314")
#     # return JsonResponse({})
#     if from_date is not None and from_date != "" and to_date is not None and to_date != "" :
#         # # # # # #print("3316")
#         page_items = models.On_Transit_Transaction_Details.objects.filter(on_transit_transaction_header__transaction_date__range=(from_date,to_date),status =1 , deleted=0).order_by('on_transit_transaction_header__transaction_number')
        
#     else:
#         page_items = models.On_Transit_Transaction_Details.objects.filter(status=1 , deleted = 0 ).order_by('on_transit_transaction_header__transaction_number')
#         # page_items = models.Store_Item.objects.raw("SELECT * FROM store_items GROUP BY store_id ,item_id ")    
#     # for p in page_items:
#     #     # # # # #print(p.item_id)
#     # return JsonResponse({})
#     # # # # # #print(page_items)
#     # return JsonResponse({})
#     if file_type=="xlsx":
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)

#         for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
#             if not f.endswith(".xlsx"):
#                 continue
#             os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

#         # tmpname = str(datetime.now().microsecond) + ".xlsx"
#         tmpname = "Stock Transfer Report" + ".xlsx"
#         wb = Workbook()

#         # grab the active worksheet
#         ws = wb.active

#         # Data can be assigned directly to cells
#         ws['A1'] = "Transaction number"
#         ws['B1'] = "Item"
#         ws['C1'] = "Source"
#         ws['D1'] = "Destination"
#         ws['E1'] = "Material Out Date"
#         ws['F1'] = "Material Out Quantity"
#         ws['G1'] = "Material IN Date"
#         ws['H1'] = "Material In Quantity"


#         # Rows can also be appended
#         for each in page_items:
#             # # # # # #print(each.on_transit_transaction_header.transaction_date)
            
#             mat_in_date= each.on_transit_transaction_header.transaction_in_date if each.on_transit_transaction_header.transaction_in_date != None else ""
#             mat_in_quantity = str(each.recieved_quntity) if float(each.recieved_quntity)!= 0.00 else ""
#             ws.append([
#             each.on_transit_transaction_header.transaction_number,
#             each.item.name, 
#             each.on_transit_transaction_header.source_store.name,
#             each.on_transit_transaction_header.destination_store.name,
#             str(each.on_transit_transaction_header.transaction_date),
#             str(each.quantity),
#             str(mat_in_date),
#             mat_in_quantity ])
#         # return JsonResponse({})
#         # Save the file
#         wb.save(settings.MEDIA_ROOT + '/reports/' + tmpname)
#         os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name':  tmpname
#         })
#     elif file_type == "csv":
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)

#         # Clean up any existing CSV files in the directory
#         for f in os.listdir(settings.MEDIA_ROOT + '/reports/'):
#             if not f.endswith(".csv"):
#                 continue
#             os.remove(os.path.join(settings.MEDIA_ROOT + '/reports/', f))

#         tmpname = "Stock Transfer Report" + ".csv"

#         with open(os.path.join(directory_path, tmpname), 'w', newline='') as csvfile:
#             writer = csv.writer(csvfile)
#             writer.writerow(["Transaction number", 
#             "Item","Source", 
#             "Destination", 
#             "Material Out Date",
#             "Material Out Quantity",
#             "Material In Date",
#             "Material In Quantity"])
#             for each in page_items:
                
#                 mat_in_date= each.on_transit_transaction_header.transaction_in_date if each.on_transit_transaction_header.transaction_in_date != None else ""
#                 mat_in_quantity = str(each.recieved_quntity) if float(each.recieved_quntity)!= 0.00 else ""
#                 writer.writerow(
#                     [
#                         each.on_transit_transaction_header.transaction_number,
#                         each.item.name, 
#                         each.on_transit_transaction_header.source_store.name,
#                         each.on_transit_transaction_header.destination_store.name,
#                         str(each.on_transit_transaction_header.transaction_date), 
#                         str(each.quantity),
#                         str(mat_in_date),
#                         mat_in_quantity
#                     ])

#         os.chmod(settings.MEDIA_ROOT + '/reports/' + tmpname, 0o777)
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name': tmpname
#         })

#     elif file_type == "pdf":
#         # # # # # #print("3395")
#         # Create a new PDF document with smaller margins
#         pdf = FPDF(orientation='P', unit='mm', format='A3')  # Adjust unit and format if needed
#         pdf.set_left_margin(5)
#         pdf.set_top_margin(5)

#         pdf.add_page()
#         pdf.set_font("Arial", size=9)

#         # Add a header row with bold text
#         pdf.set_font("Arial", size=9, style='B')  # Set font style to bold
#         pdf.cell(9, 10, txt="S.No.", border=1, align='C')  # Add S.No. column
#         pdf.cell(50, 10, txt="Item", border=1, align='C')
#         pdf.cell(50, 10, txt="Source", border=1, align='C')
#         pdf.cell(50, 10, txt="Destination", border=1, align='C')
#         pdf.cell(30, 10, txt="Material out Date", border=1, align='C')
#         pdf.cell(30, 10, txt="Material Out", border=1, align='C')
#         pdf.cell(30, 10, txt="Material In date ", border=1, align='C')
#         pdf.cell(20, 10, txt="Material In", border=1, align='C')
#         pdf.set_font("Arial", size=9)  # Reset font style to normal
#         pdf.ln(10)  # Move to the next line

#         # Cell height (adjust as needed)
#         cell_height = 10

#         # Add data rows
#         counter = 1  # Counter for serial numbers
#         for each in page_items:
#             mat_in_date= each.on_transit_transaction_header.transaction_in_date if each.on_transit_transaction_header.transaction_in_date != None else ""
#             mat_in_quantity = str(each.recieved_quntity) if float(each.recieved_quntity)!= 0.00 else ""
#             pdf.cell(9, 10, txt=str(counter), border=1, align='C')  # Add S.No. for each row
#             pdf.cell(50, 10, txt=each.item.name, border=1)
#             pdf.cell(50, 10, txt=each.on_transit_transaction_header.source_store.name, border=1)
#             pdf.cell(50, 10, txt=each.on_transit_transaction_header.destination_store.name, border=1)
#             pdf.cell(30, 10, txt=str(each.on_transit_transaction_header.transaction_date), border=1)
#             pdf.cell(30, 10, txt=str(each.quantity), border=1)
#             pdf.cell(30, 10, txt=str(mat_in_date), border=1)
#             # Right align price for each data row
#             pdf.cell(20, 10, txt= mat_in_quantity, border=1)
#             pdf.ln(10)
#             counter += 1  # Increment counter for next row

#         # Save the PDF file
#         directory_path = settings.MEDIA_ROOT + '/reports/'
#         path = Path(directory_path)
#         path.mkdir(parents=True, exist_ok=True)
#         tmpname = "Stock_Transfer_Report.pdf"
#         pdf.output(os.path.join(directory_path, tmpname))
#         os.chmod(os.path.join(directory_path, tmpname), 0o777)
#         # # # # # #print("3439")
#         return JsonResponse({
#             'code': 200,
#             'filename': settings.MEDIA_URL + 'reports/' + tmpname,
#             'name': tmpname
#         })

#material in 
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def materialInDetailsAdd(request):
    context = {}
    # # # # # #print(request.POST)
    try:
        # pass
        #print(8734)
        with transaction.atomic():
            #material added on on transit transaction heder flag be 1 
            transitTransactionHeader = models.On_Transit_Transaction.objects.get(pk=request.POST['transactionNumber'])
            # # # # # #print(transitTransactionHeader.id)
            transitTransactionHeader.flag = 1
            transitTransactionHeader.transaction_in_date = request.POST['issue_date']
            transitTransactionHeader.updated_at = datetime.now()

            transitTransactionHeader.save()
            #store transaction created for material in
            store_transaction_count = models.Store_Transaction.objects.all().count()
            storeTransactionHeader= models.Store_Transaction()
            # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'MIN')
            transaction_type = models.Transaction_Type.objects.get(name='MIN')
            storeTransactionHeader.transaction_type = transaction_type
            storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                "${CURRENT_YEAR}", current_year).replace(
                "${AI_DIGIT_5}", ai_digit_5()).replace(
                "${transaction_type_id}", str(transaction_type.id).zfill(2))
            storeTransactionHeader.transaction_date = request.POST['issue_date']
            storeTransactionHeader.reference_id =  int(transitTransactionHeader.id)
            storeTransactionHeader.save()
            # # # # # #print("4987")
           
            order_details = []
            for index in range(0,len(request.POST.getlist('item_id'))):
                #on transit transaction details for material in changed
                transitTransactionDetails = models.On_Transit_Transaction_Details.objects.get(pk=request.POST.getlist('details_id')[index])
                # # # # # #print(transitTransactionDetails)
                transitTransactionDetails.recieved_quntity = request.POST.getlist('quantity_recieved')[index]
                transitTransactionDetails.reject_quantity = request.POST.getlist('quantity_reject')[index]
                transitTransactionDetails.amount = request.POST.getlist('amount')[index]
                transitTransactionDetails.notes = request.POST.getlist('notes')[index] if request.POST.getlist('notes')[index] != "" else None
                transitTransactionDetails.updated_at = datetime.now()
                transitTransactionDetails.save()
                #store transaction details created for material in 
                order_details.append(
                    models.Store_Transaction_Detail(
                        store_transaction_header_id=storeTransactionHeader.id,
                        item_id=request.POST.getlist('item_id')[index],
                        store_id= request.POST['destination_store_id'],
                        quantity=request.POST.getlist('quantity_recieved')[index],
                        rate=request.POST.getlist('rate')[index],
                        amount=request.POST.getlist('amount')[index]
                        
                    )
                )

                # item added to destination store
                storeItem = models.Store_Item.objects.filter(
                    item_id=request.POST.getlist('item_id')[index], store_id=request.POST['destination_store_id']).first()
                if storeItem is None:
                    storeItem = models.Store_Item()
                    storeItem.opening_qty = Decimal(request.POST.getlist('quantity_recieved')[index])
                    storeItem.on_hand_qty = Decimal(request.POST.getlist('quantity_recieved')[index])
                    storeItem.closing_qty = Decimal(request.POST.getlist('quantity_recieved')[index])
                    storeItem.item_id = request.POST.getlist('item_id')[index]
                    storeItem.store_id = request.POST['destination_store_id']
                    storeItem.save()
                else:
                    storeItem.on_hand_qty += Decimal(request.POST.getlist('quantity_recieved')[index])
                    storeItem.closing_qty += Decimal(request.POST.getlist('quantity_recieved')[index])
                    storeItem.updated_at = datetime.now()
                storeItem.save()
                # change in storeItemCurrent min
                # Fetch the last transaction_date less than the given_date
                given_date = request.POST['issue_date']
                
               
                # Check for the last record on the given_date
                record = models.Store_Item_Current.objects.filter(transaction_date=given_date, item_id=request.POST.getlist('item_id')[index], store_id=request.POST['destination_store_id'],status=1, deleted=0).last()
                
                if not record:
                    # If no record is found for the given_date, look for the last record before that date
                    last_transaction_date = models.Store_Item_Current.objects.filter(
                        transaction_date__lt=given_date, item_id=request.POST.getlist('item_id')[index], store_id=request.POST['destination_store_id'],status=1, deleted=0
                    ).aggregate(Max('transaction_date'))['transaction_date__max']

                    

                    if last_transaction_date:
                        # Fetch the record for the last_transaction_date
                        record = models.Store_Item_Current.objects.filter(
                            transaction_date=last_transaction_date, item_id=request.POST.getlist('item_id')[index], store_id=request.POST['destination_store_id'],status=1, deleted=0
                        ).last()
                #print(record)
                # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                
                
                store_item_instance = models.Store_Item_Current()

                if record:
                    # Set values based on the last record found
                    store_item_instance.opening_qty = record.closing_qty
                    store_item_instance.on_hand_qty = record.closing_qty + Decimal(request.POST.getlist('quantity_recieved')[index])
                    store_item_instance.closing_qty = record.closing_qty +Decimal(request.POST.getlist('quantity_recieved')[index])
                else:
                    # Set values based on the current transaction if no prior record exists
                    store_item_instance.opening_qty = Decimal(
                       0.00
                    )
                    store_item_instance.on_hand_qty =Decimal(request.POST.getlist('quantity_recieved')[index])
                    store_item_instance.closing_qty =Decimal(request.POST.getlist('quantity_recieved')[index])
                if(store_item_instance.on_hand_qty<0):
                    raise ValueError(f"on hand quantity is less than 0")
                # Set other fields for the new transaction
                store_item_instance.store_transaction_id = storeTransactionHeader.id
                store_item_instance.transaction_date = given_date
                store_item_instance.item_id = request.POST.getlist('item_id')[index]
                store_item_instance.store_id = request.POST['destination_store_id']

                # Save the instance to the database
                store_item_instance.save()
                
                store_item_curreEdit(request.POST['destination_store_id'],request.POST.getlist('item_id')[index],given_date,'min',request.POST.getlist('quantity_recieved')[index]) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)

            models.Store_Transaction_Detail.objects.bulk_create(order_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Material In Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Material In Transaction added Sucessfully"
        })

    except Exception as e:
        #print(e)
        context.update({
            'status': 539,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()

    return JsonResponse(context)


# physical Inspection on Store Items --- developed by saswata
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getPhysicalInspectionHeadersList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    # # # # # #print(request.GET)
    try:
        if id is not None and id != "":
            phyInspectHeader = list(models.Physical_Inspection.objects.filter(pk=id)[:1].values('pk','transaction_number',
            'inspection_date',
            'store_id',
            'store__name',
            'item_catagories_id',
            'item_catagories__name',
            'item_type_id',
            'item_type__name'
            ))
            context.update({
            'status': 200,
            'message': "phycial Inspect  Header Fetched Successfully.",
            'page_items': phyInspectHeader,
        })

        else:
            # # # # # #print('5128')
            if keyword is not None and keyword != "":
                phyInspectHeader = list(models.Physical_Inspection.objects.filter(Q(transaction_number__icontains=keyword) | 
                Q(store__name__icontains=keyword) |
                Q(item_catagories__name__icontains=keyword) |
                Q(item_type__name__icontains=keyword)
                    ).values('pk','transaction_number',
                    'inspection_date',
                    'store_id',
                    'store__name',
                    'item_catagories_id',
                    'item_catagories__name',
                    'item_type_id',
                    'item_type__name'
                ))
            else:
                phyInspectHeader = list(models.Physical_Inspection.objects.filter(status=1, deleted=0).values('pk','transaction_number',
                'inspection_date',
                'store_id',
                'store__name',
                'item_catagories_id',
                'item_catagories__name',
                'item_type_id',
                'item_type__name'
                ))

            if find_all is not None and int(find_all) == 1:
                    context.update({
                        'status': 200,
                        'message': "phycial Inspect  Header Fetched Successfully.",
                        'page_items': phyInspectHeader,
                    })
                    return JsonResponse(context)
            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)

            paginator = CustomPaginator(phyInspectHeader, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()

            context.update({
                'status': 200,
                'message': "phycial Inspect  Header Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })



    except Exception:
        context.update({
            'status': 540,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def physicalInspectionDetailsAdd(request):
    context ={}
    # # # # # #print(request.POST)
    try:
        with transaction.atomic():

            physical_inspection_count = models.Physical_Inspection.objects.all().count()
            physicalInspectionHeader = models.Physical_Inspection()
            physicalInspectionHeader.transaction_number = env("PHY_INSP_TRANSACTION_SEQ").replace(
                "${CURRENT_YEAR}", datetime.today().strftime('%Y')).replace("${AI_DIGIT_5}", str(physical_inspection_count + 1).zfill(5))        
            physicalInspectionHeader.inspection_date = request.POST['issue_date']
            physicalInspectionHeader.store_id = request.POST['storeId']
            physicalInspectionHeader.item_catagories_id = request.POST['itemCat']
            physicalInspectionHeader.item_type_id = request.POST['itemType']
            physicalInspectionHeader.save()

             #store transaction created for material in
            store_transaction_count = models.Store_Transaction.objects.all().count()
            storeTransactionHeader= models.Store_Transaction()
            # storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name = 'PHY')
            transaction_type = models.Transaction_Type.objects.get(name='PHY')
            storeTransactionHeader.transaction_type = transaction_type
            storeTransactionHeader.transaction_number = env("STORE_TRANSACTION_NUMBER_SEQ").replace(
                "${CURRENT_YEAR}", current_year).replace(
                "${AI_DIGIT_5}", ai_digit_5()).replace(
                "${transaction_type_id}", str(transaction_type.id).zfill(2))
            storeTransactionHeader.transaction_date = request.POST['issue_date']
            storeTransactionHeader.reference_id =  int(physicalInspectionHeader.id)
            storeTransactionHeader.save()

            phyInsDet = []
            order_details = []
            for index in range(0,len(request.POST.getlist('item_id'))):
                if request.POST.getlist('physical_quantity')[index]!='':

                    adjustQuan = (float( request.POST.getlist('physical_quantity')[index]) - float(request.POST.getlist('book_quantity')[index])) if float(request.POST.getlist('book_quantity')[index]) > 0 else  (float(request.POST.getlist('book_quantity')[index]) + float( request.POST.getlist('physical_quantity')[index]))

                    phyInsDet.append(
                        models.Physical_Inspection_Details(
                            physical_inspection_header_id = physicalInspectionHeader.id,
                            item_id = request.POST.getlist('item_id')[index],
                            booked_quantity = request.POST.getlist('book_quantity')[index],
                            physical_quantity = request.POST.getlist('physical_quantity')[index],
                            adjusted_quantity = adjustQuan,
                            notes = request.POST.getlist('notes')[index]
                        )
                    )

                    order_details.append(
                    models.Store_Transaction_Detail(
                        store_transaction_header_id=storeTransactionHeader.id,
                        item_id=request.POST.getlist('item_id')[index],
                        store_id= request.POST['storeId'],
                        quantity=adjustQuan
                        )
                    )

            models.Physical_Inspection_Details.objects.bulk_create(phyInsDet)
            models.Store_Transaction_Detail.objects.bulk_create(order_details)           
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Physical Inspection Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Physical Inspection Transaction added Sucessfully"
        })
        
    except Exception:
        context.update({
            'status': 542,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


# purchase bill --- developed by saswata
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getPurchaseBillHeadersList(request):
    context = {}
    id = request.GET.get('id', None)
    find_all = request.GET.get('find_all', None)
    keyword = request.GET.get('keyword', None)
    # # # # # #print(request.GET)
    try:
        if id is not None and id != "":
            purchaseBill = list(models.Purchase_Bill.objects.filter(pk=id)[:1].values('pk','transaction_number',
            'transaction_date',
            'e_way_no',
            'e_way_date',
            'vechical_no',
            'total_igst',
            'total_cgst',
            'total_sgst',
            'total_amount',
            'total_gst_amount',
            'invoice_no',
            'vendor__name',
            'notes',
            'flag'
            ))
            context.update({
            'status': 200,
            'message': "purchase bill  Header Fetched Successfully.",
            'page_items': purchaseBill,
        })

        else:
            if keyword is not None and keyword != "":
                purchaseBill = list(models.Purchase_Bill.objects.filter(Q(transaction_number__icontains=keyword) | 
                Q(invoice_no__icontains=keyword) | Q(vendor__name__icontains=keyword)

                    ).values('pk','transaction_number',
                    'transaction_date',
                    'e_way_no',
                    'e_way_date',
                    'vechical_no',
                    'total_igst',
                    'total_cgst',
                    'total_sgst',
                    'total_amount',
                    'total_gst_amount',
                    'invoice_no',
                    'vendor__name',
                    'notes',
                    'flag'
                ))
            else:
                purchaseBill = list(models.Purchase_Bill.objects.filter(status=1, deleted=0).values('pk','transaction_number',
                'transaction_date',
                'e_way_no',
                'e_way_date',
                'vechical_no',
                'total_igst',
                'total_cgst',
                'total_sgst',
                'total_amount',
                'total_gst_amount',
                'invoice_no',
                'vendor__name',
                'notes',
                'flag'
                ))

            if find_all is not None and int(find_all) == 1:
                    context.update({
                        'status': 200,
                        'message': "purchase bill  Header Fetched Successfully.",
                        'page_items': purchaseBill,
                    })
                    return JsonResponse(context)

            per_page = int(env("PER_PAGE_DATA"))
            button_to_show = int(env("PER_PAGE_PAGINATION_BUTTON"))
            current_page = request.GET.get('current_page', 1)

            paginator = CustomPaginator(purchaseBill, per_page)
            page_items = paginator.get_page(current_page)
            total_pages = paginator.get_total_pages()

            context.update({
                'status': 200,
                'message': "purchase bill  Header Fetched Successfully.",
                'page_items': page_items,
                'total_pages': total_pages,
                'per_page': per_page,
                'current_page': int(current_page),
                'button_to_show': int(button_to_show),
            })



    except Exception:
        context.update({
            'status': 540,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseBillDetailsAdd(request):
    context = {}
    # #print(request.POST)
   
    # exit()
    try:
        with transaction.atomic():
            purchase_bill_head_count = models.Purchase_Bill.objects.all().count()
            purcahse_bill_header = models.Purchase_Bill()
            # # # # # #print("5283")
            purcahse_bill_header.transaction_number = env("PURCHASE_BILL_TRANSACTION_SEQ").replace(
            "${CURRENT_YEAR}", datetime.today().strftime('%Y')).replace("${AI_DIGIT_5}", str(purchase_bill_head_count + 1).zfill(5))
            purcahse_bill_header.transaction_date = request.POST['issue_date']
 
            purcahse_bill_header.vendor_id = request.POST['vendor']

            purcahse_bill_header.invoice_no =  request.POST['invoice']
     
            purcahse_bill_header.e_way_no = request.POST['e_way']
            
            purcahse_bill_header.e_way_date = request.POST['e_way_date']

            purcahse_bill_header.vechical_no = request.POST['vehicle_no']
            purcahse_bill_header.total_gst = Decimal(request.POST['total_igst']) if float(request.POST['total_igst'])!= 0.00  else (Decimal(float(request.POST['total_cgst']) + float(request.POST['total_sgst'])))
            purcahse_bill_header.total_igst = Decimal(request.POST['total_igst'])
            purcahse_bill_header.total_cgst = Decimal(request.POST['total_cgst'])
            purcahse_bill_header.total_sgst = Decimal(request.POST['total_sgst'])
            purcahse_bill_header.total_amount = Decimal(request.POST['total']) # total amount with discount
            purcahse_bill_header.notes = request.POST['notes']
            purcahse_bill_header.total_amount_exclude_discount = Decimal(request.POST['total_acc'])
            purcahse_bill_header.total_discount = Decimal(request.POST['total_discount'])
            # #print(8100)
            purcahse_bill_header.total_discount_amount = Decimal(request.POST['total_discount_amount'])
            # #print((request.POST['tds_percentage']))
            if request.POST.get('tds_percentage',None):
                purcahse_bill_header.tds_percentage = Decimal(request.POST['tds_percentage']) 
            purcahse_bill_header.tds_deduction = Decimal(request.POST['tds_deduction'])
            if request.POST.get('tcs_percentage',None):
                purcahse_bill_header.tcs_percentage = Decimal(request.POST['tcs_percentage'])
            purcahse_bill_header.tcs_deduction = Decimal(request.POST['tcs_decuction'])
            purcahse_bill_header.accet_amount_tax_deduc =Decimal(request.POST['accet_tds_tcs_de_amount'])
            purcahse_bill_header.roundof_total_amount =  Decimal(request.POST['roundof_amount'])
            # #print(8106)
            # #print(Decimal(request.POST['accet_tds_tcs_de_amount']))
            # x= 10/0
            purcahse_bill_header.round_off_price = Decimal('0.' + (request.POST['accet_tds_tcs_de_amount'].split('.')[1]))
            purcahse_bill_header.total_gst_amount = Decimal(request.POST['total_amount_with_gst'])
            purcahse_bill_header.notes = request.POST['notes'] 
            purcahse_bill_header.save()
           
            
            if(request.POST.get('igst',None)):
                bill_details = []
                for index in range(0,len(request.POST.getlist('item_id'))):
                    uom_id = models.Uom.objects.get(name = request.POST.getlist('uom')[index])
                    bill_details.append(
                        models.Purchase_Bill_Details(
                            purchase_bill_header_id = purcahse_bill_header.id,
                            item_id = request.POST.getlist('item_id')[index],
                            hsn_code = request.POST.getlist('hsn')[index],
                            quantity = request.POST.getlist('quantity')[index],
                            uom_id = uom_id.id,
                            rate = request.POST.getlist('rate')[index],
                            amount = request.POST.getlist('amount')[index],
                            amount_exclude_discount = request.POST.getlist('amountac')[index],
                            discount_percentage = request.POST.getlist('discount')[index],
                            discount_amount = request.POST.getlist('amountdiscount')[index],
                            igst_percentage = request.POST.getlist('igst')[index],
                            igst_amount =  request.POST.getlist('amount_igst')[index], 
                            gst_percentage = request.POST.getlist('gst')[index],
                            gst_amount = request.POST.getlist('amount_igst')[index],
                            amount_with_gst = request.POST.getlist('amount_gst')[index]

                        )
                    )
                models.Purchase_Bill_Details.objects.bulk_create(bill_details)
                
            else:
                # # # # # #print("5325")
                bill_details = []
                for index in range(0,len(request.POST.getlist('item_id'))):
                    uom_id = models.Uom.objects.get(name = request.POST.getlist('uom')[index])
                    bill_details.append(
                        models.Purchase_Bill_Details(
                            purchase_bill_header_id = purcahse_bill_header.id,
                            item_id = request.POST.getlist('item_id')[index],
                            quantity = request.POST.getlist('quantity')[index],
                            hsn_code = request.POST.getlist('hsn')[index],
                            uom_id = uom_id.id,
                            rate = request.POST.getlist('rate')[index],
                            amount = request.POST.getlist('amount')[index],
                            amount_exclude_discount = request.POST.getlist('amountac')[index],
                            discount_percentage = request.POST.getlist('discount')[index],
                            discount_amount = request.POST.getlist('amountdiscount')[index],
                            cgst_percentage = request.POST.getlist('cgst')[index],
                            cgst_amount = request.POST.getlist('amount_cgst')[index],
                            sgst_percentage = request.POST.getlist('sgst')[index],
                            sgst_amount =  request.POST.getlist('amount_sgst')[index],
                            gst_percentage = request.POST.getlist('gst')[index],
                            gst_amount = Decimal(request.POST.getlist('amount_cgst')[index]) + Decimal(request.POST.getlist('amount_sgst')[index]),
                            amount_with_gst = request.POST.getlist('amount_gst')[index]
                        )

                    )
                models.Purchase_Bill_Details.objects.bulk_create(bill_details)
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'purchase Bill Add')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "purchase bill added succesfully"
        })
    except Exception as e:
        #print(f"error as {e}")
        context.update({
            'status': 544,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseBillDetailsEdit(request):
    context={}
    # # # # # #print(request.POST)
    try:
        with transaction.atomic():

            purcahse_bill_header_update = models.Purchase_Bill.objects.get(pk = request.POST['headerPk'])
            purcahse_bill_header_update.invoice_no =  request.POST['invoice']
            # # # # # #print("5288")
            purcahse_bill_header_update.e_way_no = request.POST['e_way']
            
            purcahse_bill_header_update.e_way_date = request.POST['e_way_date']

            purcahse_bill_header_update.vechical_no = request.POST['vehicle_no']
            if(request.POST.get('igst')):
                purcahse_bill_header_update.total_igst = request.POST['total_igst']
            else:
                purcahse_bill_header_update.total_cgst = request.POST['total_cgst']
                purcahse_bill_header_update.total_sgst = request.POST['total_sgst']
            purcahse_bill_header_update.total_amount = request.POST['total']
            purcahse_bill_header_update.total_amount_exclude_discount = request.POST['total_acc']
            purcahse_bill_header_update.total_discount = request.POST['total_discount']
            purcahse_bill_header_update.total_discount_amount = request.POST['total_discount_amount']
            purcahse_bill_header_update.tds_percentage = request.POST['tds_percentage']
            purcahse_bill_header_update.tds_deduction = request.POST['tds_deduction']
            purcahse_bill_header_update.tcs_percentage = request.POST['tcs_percentage']
            purcahse_bill_header_update.tcs_deduction = request.POST['tcs_decuction']
            purcahse_bill_header_update.accet_amount_tax_deduc =Decimal(request.POST['accet_tds_tcs_de_amount'])
            purcahse_bill_header_update.roundof_total_amount =  Decimal(request.POST['roundof_amount'])
            purcahse_bill_header_update.round_off_price = Decimal('0.'+ request.POST['accet_tds_tcs_de_amount'].split('.')[1])
            purcahse_bill_header_update.total_gst_amount = request.POST['total_amount_with_gst']
            purcahse_bill_header_update.total_gst = request.POST['total_igst'] if float(request.POST['total_igst'])!= 0.00  else (float(request.POST['total_cgst']) + float(request.POST['total_sgst']))
            purcahse_bill_header_update.notes = request.POST['notes']
            purcahse_bill_header_update.updated_at = datetime.now()
            purcahse_bill_header_update.save()

            for index in range(0,len(request.POST.getlist('detailPk'))):
                purcahse_bill_details_update = models.Purchase_Bill_Details.objects.get(pk = request.POST.getlist('detailPk')[index])
                purcahse_bill_details_update.quantity = request.POST.getlist('quantity')[index]
                purcahse_bill_details_update.hsn_code = request.POST.getlist('hsn')[index]
                purcahse_bill_details_update.rate = request.POST.getlist('rate')[index]
                purcahse_bill_details_update.amount = request.POST.getlist('amount')[index]
                if(request.POST.get('igst')):
                    purcahse_bill_details_update.igst_percentage = request.POST.getlist('igst')[index]
                    purcahse_bill_details_update.igst_amount = request.POST.getlist('amount_igst')[index]
                    purcahse_bill_details_update.gst_percentage = request.POST.getlist('gst')[index]
                    purcahse_bill_details_update.gst_amount = request.POST.getlist('amount_igst')[index]
                else:
                    purcahse_bill_details_update.cgst_percentage =  request.POST.getlist('cgst')[index]
                    purcahse_bill_details_update.cgst_amount = request.POST.getlist('amount_cgst')[index]
                    purcahse_bill_details_update.sgst_percentage = request.POST.getlist('sgst')[index]
                    purcahse_bill_details_update.sgst_amount = request.POST.getlist('amount_sgst')[index]
                    purcahse_bill_details_update.amount_with_gst = request.POST.getlist('amount_gst')[index]
                    purcahse_bill_details_update.gst_percentage = request.POST.getlist('gst')[index]
                    purcahse_bill_details_update.gst_amount =Decimal(request.POST.getlist('amount_cgst')[index]) + Decimal(request.POST.getlist('amount_sgst')[index])
                purcahse_bill_details_update.amount_exclude_discount = request.POST.getlist('amountac')[index]
                purcahse_bill_details_update.discount_percentage = request.POST.getlist('discount')[index]
                purcahse_bill_details_update.discount_amount = request.POST.getlist('amountdiscount')[index]             
                
                purcahse_bill_details_update.updated_at = datetime.now()
                purcahse_bill_details_update.save()
                # # # # # #print("5510")
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Purchase bill edit')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "Purchase bill updated sucessfully"
        })
    
    except Exception as e:
        #print(f"error is {e}")
        context.update({
            'status': 545,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def purchaseBillDetailsDelete(request):
    context = {}
    try:
        with transaction.atomic():
            purchas_bill_head = models.Purchase_Bill.objects.get(pk=request.POST['id'])
            purchas_bill_head.delete()
            userId = request.COOKIES.get('userId', None)
            user_log_details_add(userId,'Purchase Bill Delete')
        transaction.commit()
        context.update({
            'status': 200,
            'message': "transaction number deleted succesfully"
        })
    
    except Exception:
        context.update({
            'status': 546,
            'message': "Something Went Wrong. Please Try Again."
        })
        transaction.rollback()

    return JsonResponse(context)
    
@api_view(['GET'])
def purchaseBillDetailsExport(request):
    context = {}
    try:
        # Fetch page items
        page_items = models.Purchase_Bill.objects.filter(status=1, deleted=0, purchase_tally_report=0)
        page_items_exist = page_items.exists()
       
        # If no page items exist, return a response indicating no transactions left
        if not page_items_exist:
            return JsonResponse({
                'status': 404,
                'message': 'Tally report of all transactions already generated. No transactions left.'
            })

        # this is the main sql query -----
        #    sql_query= (SELECT
        #     `purchase_bill_header_id,
        #     -- For IGST percentages
        #     SUM(CASE WHEN igst_percentage = 18 THEN igst_amount ELSE 0 END) AS igst_18,
        #     SUM(CASE WHEN igst_percentage = 28 THEN igst_amount ELSE 0 END) AS igst_28,
            
        #     -- For CGST percentages
        #     SUM(CASE WHEN cgst_percentage = 9 THEN cgst_amount ELSE 0 END) AS cgst_9,
        #     SUM(CASE WHEN cgst_percentage = 14 THEN cgst_amount ELSE 0 END) AS cgst_14,
            
        #     -- For SGST percentages
        #     SUM(CASE WHEN sgst_percentage = 9 THEN sgst_amount ELSE 0 END) AS sgst_9,
        #     SUM(CASE WHEN sgst_percentage = 14 THEN sgst_amount ELSE 0 END) AS sgst_14
        # FROM
        #     purchase_bill_details
        # GROUP BY
        #     purchase_bill_header_id
        # LIMIT 0, 25;
        #         Filtered data with aggregations`)
        #------sql-----
        # filtered_data = models.Purchase_Bill_Details.objects.filter(
        #     purchase_bill_header__purchase_tally_report=0
        # ).values(
        #     'purchase_bill_header_id',
        #     'purchase_bill_header__vendor__name',
        #     'purchase_bill_header__vendor__address',
        #     'purchase_bill_header__vendor__gst_no',
        #     'purchase_bill_header__vendor__state__name',
        #     'purchase_bill_header__invoice_no',
        #     'purchase_bill_header__total_amount',
        #     'purchase_bill_header__total_igst',
        #     'purchase_bill_header__total_cgst',
        #     'purchase_bill_header__total_sgst',
        #     'purchase_bill_header__total_gst_amount',
        #     'purchase_bill_header__total_discount_amount',
        #     'purchase_bill_header__tds_deduction',
        #     'purchase_bill_header__tcs_deduction',
        #     'purchase_bill_header__transaction_number',
        #     'purchase_bill_header__transaction_date',
        #     'purchase_bill_header__round_off_price',
        #     'purchase_bill_header__roundof_total_amount',
        # ).annotate(
        #     gst_0=Sum(Case(
        #         When(gst_percentage=0, then='gst_amount'),
        #         default=0,
        #         output_field=DecimalField()
        #     )),
        #     gst_5=Sum(Case(
        #         When(gst_percentage=5, then='gst_amount'),
        #         default=0,
        #         output_field=DecimalField()
        #     )),
        #     gst_12=Sum(Case(
        #         When(gst_percentage=12, then='gst_amount'),
        #         default=0,
        #         output_field=DecimalField()
        #     )),
        #     gst_28=Sum(Case(
        #         When(gst_percentage=28, then='gst_amount'),
        #         default=0,
        #         output_field=DecimalField()
        #     )),
        #     gst_18=Sum(Case(
        #         When(gst_percentage=18, then='gst_amount'),
        #         default=0,
        #         output_field=DecimalField()
        #     ))
        # ).order_by('purchase_bill_header_id')

        filtered_data = models.Purchase_Bill_Details.objects.filter(
            purchase_bill_header__purchase_tally_report=0
        ).values(
            'purchase_bill_header_id',
            'purchase_bill_header__vendor__name',
            'purchase_bill_header__vendor__address',
            'purchase_bill_header__vendor__gst_no',
            'purchase_bill_header__vendor__state__name',
            'purchase_bill_header__invoice_no',
            'purchase_bill_header__total_amount',
            'purchase_bill_header__total_igst',
            'purchase_bill_header__total_cgst',
            'purchase_bill_header__total_sgst',
            'purchase_bill_header__total_gst_amount',
            'purchase_bill_header__total_discount_amount',
            'purchase_bill_header__tds_deduction',
            'purchase_bill_header__tcs_deduction',
            'purchase_bill_header__transaction_number',
            'purchase_bill_header__transaction_date',
            'purchase_bill_header__round_off_price',
            'purchase_bill_header__roundof_total_amount',
        ).annotate(
            gst_0=Sum(Case(
                When(gst_percentage=0, then='amount'),
                default=0,
                output_field=DecimalField()
            )),
            gst_5=Sum(Case(
                When(gst_percentage=5, then='amount'),
                default=0,
                output_field=DecimalField()
            )),
            gst_12=Sum(Case(
                When(gst_percentage=12, then='amount'),
                default=0,
                output_field=DecimalField()
            )),
            gst_28=Sum(Case(
                When(gst_percentage=28, then='amount'),
                default=0,
                output_field=DecimalField()
            )),
            gst_18=Sum(Case(
                When(gst_percentage=18, then='amount'),
                default=0,
                output_field=DecimalField()
            ))
        ).order_by('purchase_bill_header_id')

        # Create directory if not exists
        directory_path = settings.MEDIA_ROOT + '/purchase_transition_tally/'
        path = Path(directory_path)
        path.mkdir(parents=True, exist_ok=True)

        # Create a new Excel file
        tmpname = "purchasebill_" + datetime.now().strftime("%Y_%m_%d_%H_%M_%S") + ".xlsx"
        wb = Workbook()
        ws = wb.active

       # Add headers
        ws['A1'] = "Tran Type"
        ws['B1'] = "DSDVchNo : $_2"
        ws['C1'] = "Inv Date1"
        ws['D1'] = "Act Code4"
        ws['E1'] = "PartyName : $_3"
        ws['F1'] = "Place6"
        ws['G1'] = "Add2"
        ws['H1'] = "Gst No7"
        ws['I1'] = "TKGimpPartyState 15"
        ws['J1'] = "Reference"
        ws['K1'] = "Ref Dt"
        ws['L1'] = "Accessable Value"
        ws['M1'] = "GST@ 0%"
        ws['N1'] = "GST@ 05%"
        ws['O1'] = "GST@ 12%"
        ws['P1'] = "GST@ 18%"
        ws['Q1'] = "GST@ 28%"
        ws['R1'] = "Output Igst Amt"
        ws['S1'] = "Output Cgst Amt"
        ws['T1'] = "Output Sgst Amt"
        ws['U1'] = "Gst Cess Amt"
        ws['V1'] = "LESS DISCOUNT"
        ws['W1'] = "TCS PAYBLE"
        ws['X1'] = "TDS RECEIVABLE"
        ws['Y1'] = "RoundOffAmt :$_11"
        ws['Z1'] = "Tran Total"
        ws['AA1'] = "VchNarration : $_13"


        # ws['A1'] = "Vendor Name"
        # ws['B1'] = "Vendor Address"
        # ws['C1'] = "Vendor GST No"
        # ws['D1'] = "Invoice No"
        # ws['E1'] = "Total Amount"
        # ws['F1'] = "Total IGST"
        # ws['G1'] = "Total CGST"
        # ws['H1'] = "Total SGST"
        # ws['I1'] = "Total Amount Including GST"
        # ws['J1'] = "IGST 18%"
        # ws['K1'] = "IGST 28%"
        # ws['L1'] = "CGST 9%"
        # ws['M1'] = "CGST 14%"
        # ws['N1'] = "SGST 9%"
        # ws['O1'] = "SGST 14%"
        # ws['P1'] = "vch type"

        # Append data rows
        for each in filtered_data:
            # #print(each)
            ws.append([
                "P",
                each['purchase_bill_header__transaction_number'],
                each['purchase_bill_header__transaction_date'], 
                " ",
                each['purchase_bill_header__vendor__name'],
                each['purchase_bill_header__vendor__address'],
                " ",
                each['purchase_bill_header__vendor__gst_no'],
                each['purchase_bill_header__vendor__state__name'],
                each['purchase_bill_header__invoice_no'],
                " ",
                each['purchase_bill_header__total_amount'],
                each['gst_0'],
                each['gst_5'],
                each['gst_12'],
                each['gst_18'],
                each['gst_28'],
                each['purchase_bill_header__total_igst'],
                each['purchase_bill_header__total_cgst'],
                each['purchase_bill_header__total_sgst'],
                " ",
                each['purchase_bill_header__total_discount_amount'],
                each['purchase_bill_header__tcs_deduction'],
                each['purchase_bill_header__tds_deduction'],
                each['purchase_bill_header__round_off_price'],
                each['purchase_bill_header__roundof_total_amount'],
                " "
            ])
        #print(8058)
        # Save the file
        file_path = os.path.join(directory_path, tmpname)
        
       # Save the file to the server
        wb.save(file_path)

        # os.chmod(settings.MEDIA_ROOT + '/purchase_transition_tally/' + tmpname, 0o777)

        os.chmod(file_path, 0o777)

        # Update page items
        page_items.update(purchase_tally_report=1)

        filename = settings.MEDIA_URL + 'purchase_transition_tally/' + tmpname

        context.update({
            'status': 200,
            'message': 'File generated successfully in server Media :' + filename,
            'file_url': filename
        })
    except Exception as e:
        #print(f'error{e}')
        context.update({
            'status': 546.1,
            'message': "Something went wrong. Please try again."
        })
    return JsonResponse(context)

@api_view(['POST'])
def fgRawDetailsExport(request):
    context = {}
    try:
        page_items = models.Store_Transaction_Detail.objects.filter(store_transaction_header__tally_sync=0,status=1,deleted=0)
        page_items = page_items.filter(store_id =request.POST['store_id']) 
        page_items_exist = page_items.exists()
    
        # If no page items exist, return a response indicating no transactions left
        if not page_items_exist:
            return JsonResponse({
                'status': 404,
                'message': 'Tally report of all transactions already generated. No transactions left.'
            })
        with transaction.atomic():
            filtered_data = models.Store_Transaction_Detail.objects.filter(store_transaction_header__tally_sync=0,status=1,deleted=0)
            filtered_data = filtered_data.filter(store_id =request.POST['store_id'] )
            # Create directory if not exists
            directory_path = settings.MEDIA_ROOT + '/fg_raw_tansition_tally/'
            path = Path(directory_path)
            path.mkdir(parents=True, exist_ok=True)

            # Create a new Excel file
            tmpname = "fgRawreport_" + datetime.now().strftime("%Y_%m_%d_%H_%M_%S") + ".xlsx"
            wb = Workbook()
            ws = wb.active

        # Add headers
            ws['A1'] = "Vch No"
            ws['B1'] = "Date"
            ws['C1'] = "Name of Item"
            ws['D1'] = "Unit"
            ws['E1'] = "Location"
            ws['F1'] = "HSN Code"
            ws['G1'] = "IGSt Rate"
            ws['H1'] = "CGST Rate"
            ws['I1'] = "SGST Rate"
            ws['J1'] = "Cess Rate"
            ws['K1'] = "Quantity"
            ws['L1'] = "Unit"
            ws['M1'] = "Rate"
            ws['N1'] = "Amount"
            ws['O1'] = "Remarks"
            


        

            # Append data rows
            for each in filtered_data:
                # #print(each)
                rate = each.rate if each.rate else each.item.price
                ws.append([
                    each.store_transaction_header.transaction_number,
                    each.store_transaction_header.transaction_date.strftime("%d-%m-%Y"), 
                    each.item.name,
                    each.item.uom.name,
                    each.store.name,
                    each.item.hsn_code,
                    each.item.item_type.gst_percentage,
                    ((each.item.item_type.gst_percentage)/Decimal(2.0)),
                    ((each.item.item_type.gst_percentage)/Decimal(2.0)),
                    '0.00',
                    each.quantity,
                    each.item.uom.name,
                    rate,
                    (each.quantity * rate),
                    ''
                ])
            #print(8058)
            # Save the file
            file_path = os.path.join(directory_path, tmpname)
            
        # Save the file to the server
            wb.save(file_path)

            # os.chmod(settings.MEDIA_ROOT + '/purchase_transition_tally/' + tmpname, 0o777)

            os.chmod(file_path, 0o777)

            # Update page items
            for page_item in page_items:
                storeTranasctionHeader = models.Store_Transaction.objects.get(pk=page_item.store_transaction_header_id)
                storeTranasctionHeader.tally_sync = 1
                storeTranasctionHeader.updated_at = datetime.now()
                storeTranasctionHeader.save()

            filename = settings.MEDIA_URL + 'fg_raw_tansition_tally/' + tmpname

            context.update({
                'status': 200,
                'message': 'File generated successfully in server Media :' + filename,
                'file_url': filename
            })
        transaction
    except Exception as e:
        #print(f'error{e}')
        context.update({
            'status': 546.1,
            'message': "Something went wrong. Please try again."
        })
        transaction.rollback()
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportItemTrackingReport(request):
    context = {}
    item_id = request.POST.get('item_id', None)
    store_id = request.POST.get('store_id', None)
    from_date = request.POST.get('from_date', None)
    to_date = request.POST.get('to_date', None)

   
    # Parse from_date and to_date strings to DateTime objects
    from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

    item = models.Item.objects.filter(pk=item_id)
    IsVendor = models.Store.objects.filter(pk=store_id).first()
    total_quantity = Decimal(0.00)
    total_reciept = Decimal(0.00)
    total_reciept_by_job_order = Decimal(0.00)
    total_out = Decimal(0.00)
    total_out_by_job_order = Decimal(0.00)
    
    data = []
    # if item.exists():
    #     item = item.first()
    #     store = store.first()
    #     # Get all store_transaction_details for the item within the date range
    #     store_transaction_details = models.Store_Transaction_Detail.objects.filter(
    #         item=item,
    #         store=store,
    #         created_at__range=(from_date, to_date + timedelta(days=1)),  # Include to_date
    #     ).filter(store_transaction_header__status=1,store_transaction_header__deleted=0).select_related('store_transaction_header', 'store_transaction_header__transaction_type')

        
    #     # Group store_transaction_details by store
    #     for store_transaction_detail in store_transaction_details:
    #         reciept_quantity = Decimal(0.00)
    #         out_quantity = Decimal(0.00)
    #         if  (store_transaction_detail.store_transaction_header.transaction_type.name == 'MIS' or 
    #         store_transaction_detail.store_transaction_header.transaction_type.name == 'MR' or 
    #         store_transaction_detail.store_transaction_header.transaction_type.name == 'MOUT' or  
    #         store_transaction_detail.store_transaction_header.transaction_type.name == 'MIST' or  
    #         store_transaction_detail.store_transaction_header.transaction_type.name == 'MIV' 
    #         ):
    #             out_quantity = store_transaction_detail.quantity
    #             total_out += out_quantity
    #         else:
    #             reciept_quantity = store_transaction_detail.quantity
    #             total_reciept += reciept_quantity
    #         data.append({
    #             'reciept_quantity': reciept_quantity ,
    #             'out_quantity' : out_quantity,
    #             'rate': store_transaction_detail.rate,
    #             'amount': store_transaction_detail.amount,
    #             'gst_percentage': store_transaction_detail.gst_percentage,
    #             'amount_with_gst': store_transaction_detail.amount_with_gst,
    #             'transaction_number': store_transaction_detail.store_transaction_header.transaction_number,
    #             'transaction_id' : store_transaction_detail.store_transaction_header_id,
    #             'invoice_challan_no' : store_transaction_detail.store_transaction_header.invoice_challan if store_transaction_detail.store_transaction_header.invoice_challan else '---',
    #             'transaction_type': store_transaction_detail.store_transaction_header.transaction_type.name,
    #             'updated_at': store_transaction_detail.updated_at.date(),
    #             'transaction_date' : store_transaction_detail.store_transaction_header.transaction_date
    #         })
    #     data.sort(key=lambda x: (x['transaction_date'], x['transaction_number']))
    if item.exists():
        #print("item exists")
        store_item_currents = models.Store_Item_Current.objects.filter(
            status = 1,
            deleted = 0,
            store_id=store_id,
            item_id=item_id,
            transaction_date__range=(from_date, to_date)
        ).order_by('transaction_date', 'created_at')
        
        for store_item_current in store_item_currents:
            #print(store_item_current)
            rate = item.first().price
            amount = Decimal(0.00)
            reciept_quantity = Decimal(0.00)
            reciept_ByJobOrder = Decimal(0.00)
            utilised_by_jobOrder = Decimal(0.00)
            out_quantity = Decimal(0.00)
            reciept_ByGRN = Decimal(0.00)
            issued_ByJobOrder = Decimal(0.00)
            if store_item_current.store_transaction_id:
                
                

                store_transaction_detail = models.Store_Transaction_Detail.objects.filter(store_id=store_id,
                    item_id=item_id, store_transaction_header_id = store_item_current.store_transaction_id).first()
                # #print(store_transaction_detail,store_item_current.store_transaction_id,store_id,item_id)
                if not  store_transaction_detail  :
                    #print(f"store tr id : {store_item_current.store_transaction_id}") 
                    context.update({
                        'status': 501.1,
                        'types': 'InHouse',
                        'message': f"problem in store tr id : {store_item_current.store_transaction_id}",
                        
                        
                    })
                    return  JsonResponse(context)
                if store_item_current.store_transaction.transaction_type.name == "MIS" or store_item_current.store_transaction.transaction_type.name == "MOUT" or store_item_current.store_transaction.transaction_type.name == "MIST" or store_item_current.store_transaction.transaction_type.name == "MIV":
                    out_quantity = store_transaction_detail.quantity
                    if store_item_current.store_transaction.transaction_type.name == "MIS" or store_item_current.store_transaction.transaction_type.name == "MOUT" :
                        issued_ByJobOrder = store_transaction_detail.quantity
                    if store_item_current.store_transaction.transaction_type.name == "MIST" :
                        utilised_by_jobOrder = store_transaction_detail.quantity

                    total_out += issued_ByJobOrder
                    total_out_by_job_order += utilised_by_jobOrder
                print(store_transaction_detail.store_transaction_header.id)
                if  store_item_current.store_transaction.transaction_type.name == "GRN" or store_item_current.store_transaction.transaction_type.name == "GRNT" or store_item_current.store_transaction.transaction_type.name == "MIN" or store_item_current.store_transaction.transaction_type.name == "SP":
                    if store_item_current.store_transaction.transaction_type.name == "GRN" or store_item_current.store_transaction.transaction_type.name == "MIN" or store_item_current.store_transaction.transaction_type.name == "SP":
                        reciept_ByGRN = store_transaction_detail.quantity
                    if store_item_current.store_transaction.transaction_type.name == "GRNT":
                        reciept_ByJobOrder = store_transaction_detail.quantity

                    reciept_quantity = store_transaction_detail.quantity
                    total_reciept += reciept_ByGRN
                    total_reciept_by_job_order += reciept_ByJobOrder
                rate = store_transaction_detail.rate
                amount = store_transaction_detail.amount
                data.append({
                    'reciept_quantity': reciept_quantity ,
                    'out_quantity' : out_quantity,
                    'rate':rate,
                    'job_order_no': store_transaction_detail.store_transaction_header.job_order.order_number if store_transaction_detail.store_transaction_header.job_order_id else '---',
                    'purchase_order_no': store_transaction_detail.store_transaction_header.purchase_order_header.order_number if store_transaction_detail.store_transaction_header.purchase_order_header_id else '---',
                    'amount': store_transaction_detail.amount,
                    'gst_percentage': store_transaction_detail.gst_percentage,
                    'amount_with_gst': store_transaction_detail.amount_with_gst,
                    'transaction_number': store_transaction_detail.store_transaction_header.transaction_number,
                    'transaction_id' : store_transaction_detail.store_transaction_header_id,
                    'invoice_challan_no' : store_transaction_detail.store_transaction_header.invoice_challan if store_transaction_detail.store_transaction_header.invoice_challan else '---',
                    'transaction_type': store_transaction_detail.store_transaction_header.transaction_type.name,
                    'updated_at': store_transaction_detail.updated_at.date(),
                    'transaction_date' : store_transaction_detail.store_transaction_header.transaction_date,
                    'opening_qty': store_item_current.opening_qty,
                    'closing_qty': store_item_current.closing_qty,
                    'notes': store_item_current.quantity_Transfer if store_item_current.quantity_Transfer else '---'
                })
            else:
                data.append({
                    'reciept_quantity':reciept_quantity ,
                    'out_quantity' : out_quantity,
                    'rate':rate,
                    'amount': '--',
                    'job_order_no':  '---',
                    'purchase_order_no': '---',
                    'gst_percentage': '---',
                    'amount_with_gst': '---',
                    'transaction_number':'--',
                    'transaction_id' : '---',
                    'invoice_challan_no' : '---',
                    'transaction_type': '---',
                    'updated_at': store_item_current.updated_at.date(),
                    'transaction_date' :store_item_current.transaction_date,
                    'opening_qty': store_item_current.opening_qty,
                    'closing_qty': store_item_current.closing_qty,
                    'notes': store_item_current.quantity_Transfer if store_item_current.quantity_Transfer else '---'
                })
        if IsVendor and IsVendor.vendor is not None:
            context.update({
            'status': 200,
            'types': 'vendor',
            'message': "Items Fetched Successfully.",
            'total_reciept': total_reciept,
            'total_out' : total_out,
            'total_reciept_by_job_order': total_reciept_by_job_order,
            'total_out_by_job_order': total_out_by_job_order,
            'page_items': data
            
        })
        else:
           context.update({
            'status': 200,
            'types': 'InHouse',
            'message': "Items Fetched Successfully.",
            'total_reciept': total_reciept,
            'total_out' : total_out,
            'page_items': data
            
        })
        
    return JsonResponse(context)


# @api_view(['GET','POST'])
# @permission_classes([IsAuthenticated])
# def reportInventorySummary(request):
#     context = {}
#     # # # # # #print(request.POST)
#     from_date = request.POST.get('from_date', None)
#     to_date = request.POST.get('to_date', None)
#     store_id = request.POST.get('store_id', None)
#     data =[]
#     on_transit_details =[]
#     total_stockOut = 0.00
#     total_stockIn = 0.00
#     total_order = 0.00
     
#     try:
#         if request.method == 'GET':
#             # # # # # #print('4277')
#             store_Item = models.Store_Transaction_Detail.objects.filter(status=1,
#             deleted=0
#             ).filter(Q(store_transaction_header__transaction_type__name='MIS') | Q(store_transaction_header__transaction_type__name='GRN')).order_by('store_transaction_header__transaction_date')
#         else:
#             store_Item = models.Store_Item.objects.filter(store_id=store_id)
#         for each in store_Item:
           
#             store_transactions_MIS = models.Store_Transaction_Detail.objects.filter(store_id=store_id , item_id = each.item_id,store_transaction_header__transaction_type__name = 'MIS').filter(store_transaction_header__transaction_date__range =(from_date,to_date)).order_by('item_id')
#             store_transactions_GRN = models.Store_Transaction_Detail.objects.filter(store_id=store_id , item_id = each.item_id,store_transaction_header__transaction_type__name = 'GRN').filter(store_transaction_header__transaction_date__range =(from_date,to_date)).order_by('item_id')
#            # stock out
#             if store_transactions_MIS :
#                 for store_transaction in store_transactions_MIS:
#                     total_stockOut += float(store_transaction.quantity)
#                     orderQuantity = models.Job_Order_Detail.objects.filter(item_id = each.item_id ,direction='outgoing' , job_order_header_id = store_transaction.store_transaction_header.job_order_id).first()
#                     # # # # #print(7416)
#                     data.append({
#                                 'item':each.item.name,
#                                 'item_category': each.item.item_type.item_category.name,
#                                 'quantity_order':str(orderQuantity.quantity) if orderQuantity.quantity else '---',
#                                 'date': store_transaction.store_transaction_header.transaction_date,
#                                 'transaction_number': store_transaction.store_transaction_header.transaction_number,
#                                 'vendor':  store_transaction.store_transaction_header.vendor.name,
#                                 'previous_onHand_Quantity': float(each.on_hand_qty) + float(store_transaction.quantity) ,
#                                 'uom': each.item.uom.name,
#                                 'stock_in' : '---',
#                                 'stock_in_upto': '---',
#                                 'stock_out' : store_transaction.quantity,
#                                 'stock_out_upto': total_stockOut,
#                                 'onHand_quantity' : each.on_hand_qty
#                             })
#                     # # # # #print(data)
#             #stock in
#             if store_transactions_GRN :
#                 for store_transaction in store_transactions_GRN:
#                     total_stockIn += float(store_transaction.quantity)
#                     #for purchase Job order
#                     if store_transaction.store_transaction_header.job_order_id:
#                         orderQuantity = models.Job_Order_Detail.objects.filter(item_id = each.item_id ,direction='incoming' , job_order_header_id = store_transaction.store_transaction_header.job_order_id).first()
#                         data.append({
#                                 'item':each.item.name,
#                                 'item_category': each.item.item_type.item_category.name,
#                                 'quantity_order':str(orderQuantity.quantity) if orderQuantity.quantity else '---',
#                                 'date': store_transaction.store_transaction_header.transaction_date,
#                                 'transaction_number': store_transaction.store_transaction_header.transaction_number,
#                                 'vendor':  store_transaction.store_transaction_header.vendor.name,
#                                 'previous_onHand_Quantity': float(each.on_hand_qty) + float(store_transaction.quantity) ,
#                                 'uom': each.item.uom.name,
#                                 'stock_in' : store_transaction.quantity,
#                                 'stock_in_upto': total_stockIn,
#                                 'stock_out' : '---',
#                                 'stock_out_upto': '---',
#                                 'onHand_quantity' : each.on_hand_qty
#                             })
#                     else:
#                         orderQuantity = models.Purchase_Order_Detail.object.filter(item_id = each.item_id , purchase_order_header_id = store_transaction.store_transaction_header.purchase_order_header_id).first()
#                         data.append({
#                                 'item':each.item.name,
#                                 'item_category': each.item.item_type.item_category.name,
#                                 'quantity_order':str(orderQuantity.quantity) if orderQuantity.quantity else '---',
#                                 'date': store_transaction.store_transaction_header.transaction_date,
#                                 'transaction_number': store_transaction.store_transaction_header.transaction_number,
#                                 'vendor':  store_transaction.store_transaction_header.vendor.name,
#                                 'previous_onHand_Quantity': float(each.on_hand_qty) + float(store_transaction.quantity) ,
#                                 'uom': each.item.uom.name,
#                                 'stock_in' : '---',
#                                 'stock_in_upto': '---',
#                                 'stock_out' : store_transaction.quantity,
#                                 'stock_out_upto': total_stockOut,
#                                 'onHand_quantity' : each.on_hand_qty
#                             })
#         # # # # #print(data)
#         sorted_data = sorted(data, key=lambda x: (x['item'], x['date']))               
#         context.update({
#             'status': 200,
#             'message': "Inventory Report Summary  fetch Successfully.",
#             'page_items': sorted_data,
#         })

#     except Exception:
#         context.update({
#             'status': 538.1,
#             'message': "Internal Server Error",
#         })
#     return JsonResponse(context)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def reportInventorySummary(request):
    context = {}
    from_date = request.POST.get('from_date')
    to_date = request.POST.get('to_date')
    store_id = request.POST.get('store_id')
    vendor_id = request.POST.get('vendor_id',None)
    data = []
    total_stockOut = 0.00
    total_stockIn = 0.00
    
    try:
        # Determine the queryset based on the request method
        if request.method == 'GET':
            
            store_items = models.Store_Transaction_Detail.objects.filter(
                status=1,
                deleted=0
            ).filter(
                Q(store_transaction_header__transaction_type__name='MIS') |
                Q(store_transaction_header__transaction_type__name='GRN')
            ).order_by('store_transaction_header__transaction_date')
        else:
            store_items = models.Store_Item.objects.filter(store_id=store_id)

        for each in store_items:
            # Filter transactions by item and date range
            store_transactions_MIS = models.Store_Transaction_Detail.objects.filter(
                store_id=store_id,
                item_id=each.item_id,
                store_transaction_header__status = 1,
                store_transaction_header__deleted = 0,
                store_transaction_header__transaction_type__name='MIS',
                store_transaction_header__transaction_date__range=(from_date, to_date)
            ).order_by('item_id')
            
            store_transactions_GRN = models.Store_Transaction_Detail.objects.filter(
                store_id=store_id,
                item_id=each.item_id,
                store_transaction_header__status = 1,
                store_transaction_header__deleted = 0,
                store_transaction_header__transaction_type__name='GRN',
                store_transaction_header__transaction_date__range=(from_date, to_date)
            ).order_by('item_id')
            # Process stock out transactions
            if(store_transactions_MIS):
                for store_transaction in store_transactions_MIS:
                    total_stockOut += float(store_transaction.quantity)
                    orderQuantity = models.Job_Order_Detail.objects.filter(
                        item_id=each.item_id,
                        direction='outgoing',
                        job_order_header_id=store_transaction.store_transaction_header.job_order_id
                    ).first()
                    data.append({
                        'item': each.item.name,
                        'item_category': each.item.item_type.item_category.name,
                        'quantity_order': str(orderQuantity.quantity) if orderQuantity else '---',
                        'date': store_transaction.store_transaction_header.transaction_date,
                        'transaction_number': store_transaction.store_transaction_header.transaction_number,
                        'vendor': store_transaction.store_transaction_header.vendor.name if store_transaction.store_transaction_header.vendor_id else 'self' ,
                        'previous_onHand_Quantity': float(each.on_hand_qty) + float(store_transaction.quantity),
                        'uom': each.item.uom.name,
                        'stock_in': '---',
                        'stock_in_upto': '---',
                        'stock_out': store_transaction.quantity,
                        'stock_out_upto': total_stockOut,
                        'onHand_quantity': each.on_hand_qty
                    })
            # # # # # #print(data ,"aaaa")
            # Process stock in transactions
            if(store_transactions_GRN):
                for store_transaction in store_transactions_GRN:
                    total_stockIn += float(store_transaction.quantity)
                    orderQuantity = None
                    if store_transaction.store_transaction_header.job_order_id:
                        orderQuantity = models.Job_Order_Detail.objects.filter(
                            item_id=each.item_id,
                            direction='incoming',
                            job_order_header_id=store_transaction.store_transaction_header.job_order_id
                        ).first()
                    else:
                        orderQuantity = models.Purchase_Order_Detail.objects.filter(
                            item_id=each.item_id,
                            purchase_order_header_id=store_transaction.store_transaction_header.purchase_order_header_id
                        ).first()
                    
                    data.append({
                        'item': each.item.name,
                        'item_category': each.item.item_type.item_category.name,
                        'quantity_order': str(orderQuantity.quantity) if orderQuantity else '---',
                        'date': store_transaction.store_transaction_header.transaction_date,
                        'transaction_number': store_transaction.store_transaction_header.transaction_number,
                        'vendor': store_transaction.store_transaction_header.vendor.name if store_transaction.store_transaction_header.vendor_id else 'self',
                        'previous_onHand_Quantity': float(each.on_hand_qty) - float(store_transaction.quantity),
                        'uom': each.item.uom.name,
                        'stock_in': store_transaction.quantity ,
                        'stock_in_upto': total_stockIn ,
                        'stock_out': '---', 
                        'stock_out_upto': '---',
                        'onHand_quantity': each.on_hand_qty
                    })

        # Sort and prepare final context response
        sorted_data = sorted(data, key=lambda x: (x['item'], x['date']))
        context.update({
            'status': 200,
            'message': "Inventory Report Summary fetched successfully.",
            'page_items': sorted_data,
        })

    except Exception as e:
        context.update({
            'status': 500,
            'message': "Internal Server Error: " + str(e),
        })

    return JsonResponse(context)



# @api_view(['GET', 'POST'])
# @permission_classes([IsAuthenticated])
# def reportInventorySummary(request):
#     context = {}
#     from_date = request.POST.get('from_date')
#     to_date = request.POST.get('to_date')
#     store_id = request.POST.get('store_id')
#     vendor_id = request.POST.get('vendor_id', None)

#     try:
#         # Determine the base queryset based on the request method
#         if request.method == 'GET':
#             store_items = models.Store_Transaction_Detail.objects.filter(
#                 status=1,
#                 deleted=0
#             ).filter(
#                 Q(store_transaction_header__transaction_type__name__in=['MIS', 'GRN'])
#             ).select_related(
#                 'store_transaction_header',
#                 'item__item_type__item_category',
#                 'item__uom',
#                 'store_transaction_header__vendor'
#             ).order_by('store_transaction_header__transaction_date')
#         else:
#             store_items = models.Store_Item.objects.filter(store_id=store_id).prefetch_related(
#                 Prefetch(
#                     queryset=models.Store_Transaction_Detail.objects.filter(
#                         store_transaction_header__status=1,
#                         store_transaction_header__deleted=0,
#                         store_transaction_header__transaction_date__range=(from_date, to_date),
#                         store_transaction_header__transaction_type__name__in=['MIS', 'GRN']
#                     ).select_related('store_transaction_header', 'item')
#                 )
#             )

#         # Precompute stock in and stock out using aggregation
#         aggregated_data = models.Store_Transaction_Detail.objects.filter(
#             store_id=store_id,
#             store_transaction_header__transaction_date__range=(from_date, to_date),
#             store_transaction_header__status=1,
#             store_transaction_header__deleted=0,
#         ).values(
#             'item_id',
#             'store_transaction_header__transaction_type__name'
#         ).annotate(
#             total_quantity=Sum('quantity')
#         )

#         # Prepare a dictionary for quick lookups
#         stock_data = {}
#         for entry in aggregated_data:
#             item_id = entry['item_id']
#             transaction_type = entry['store_transaction_header__transaction_type__name']
#             if item_id not in stock_data:
#                 stock_data[item_id] = {'MIS': 0, 'GRN': 0}
#             stock_data[item_id][transaction_type] += entry['total_quantity']

#         # Prepare the final data
#         data = []
#         for each in store_items:
#             item_id = each.item_id
#             item_stock = stock_data.get(item_id, {'MIS': 0, 'GRN': 0})
#             total_stockOut = item_stock['MIS']
#             total_stockIn = item_stock['GRN']

#             # Add stock out data
#             if total_stockOut > 0:
#                 data.append({
#                     'item': each.item.name,
#                     'item_category': each.item.item_type.item_category.name,
#                     'quantity_order': '---',  # Assuming orderQuantity logic is removed for optimization
#                     'date': each.store_transaction_header.transaction_date,
#                     'transaction_number': each.store_transaction_header.transaction_number,
#                     'vendor': each.store_transaction_header.vendor.name if each.store_transaction_header.vendor_id else 'self',
#                     'previous_onHand_Quantity': float(each.on_hand_qty) + float(total_stockOut),
#                     'uom': each.item.uom.name,
#                     'stock_in': '---',
#                     'stock_in_upto': '---',
#                     'stock_out': total_stockOut,
#                     'stock_out_upto': total_stockOut,
#                     'onHand_quantity': each.on_hand_qty,
#                 })

#             # Add stock in data
#             if total_stockIn > 0:
#                 data.append({
#                     'item': each.item.name,
#                     'item_category': each.item.item_type.item_category.name,
#                     'quantity_order': '---',  # Assuming orderQuantity logic is removed for optimization
#                     'date': each.store_transaction_header.transaction_date,
#                     'transaction_number': each.store_transaction_header.transaction_number,
#                     'vendor': each.store_transaction_header.vendor.name if each.store_transaction_header.vendor_id else 'self',
#                     'previous_onHand_Quantity': float(each.on_hand_qty) - float(total_stockIn),
#                     'uom': each.item.uom.name,
#                     'stock_in': total_stockIn,
#                     'stock_in_upto': total_stockIn,
#                     'stock_out': '---',
#                     'stock_out_upto': '---',
#                     'onHand_quantity': each.on_hand_qty,
#                 })

#         # Sort and prepare final context response
#         sorted_data = sorted(data, key=lambda x: (x['item'], x['date']))
#         context.update({
#             'status': 200,
#             'message': "Inventory Report Summary fetched successfully.",
#             'page_items': sorted_data,
#         })

#     except Exception as e:
#         context.update({
#             'status': 500,
#             'message': f"Internal Server Error: {str(e)}",
#         })

#     return JsonResponse(context)


# @api_view(['GET','POST'])
# @permission_classes([IsAuthenticated])
# def reportInventoryStorewise(request):
#     context = {}
#     # #print(request.POST)
#     item_cat_id = request.POST.get('item_cat_id', None)
#     store_type = request.POST.get('store_type', None)
#     data ={}
#     store_item =[]
#     try:
#         if request.method == 'GET':
#             store_item = list(models.Store_Item.objects.filter(status=1 , deleted = 0 ).values('pk','on_hand_qty','item__item_type__item_category__name','item__price','store__name','item__name'))
#         else:
#             if store_type == 'inHouse':
#                 store_item = models.Store_Item.objects.filter(item__item_type__item_category_id=item_cat_id ,store__vendor__isnull=True)
#             else :
#                 store_item = models.Store_Item.objects.filter(item__item_type__item_category_id=item_cat_id ,store__vendor__isnull=False)
#             store_item = list(store_item.values('pk','on_hand_qty','item__item_type__item_category__name','item__price','store__name','item__name','item_id'))    
#         if(len(store_item) == 0):
#             context.update({
#                 'status': 200,
#                 'message': "no item found  ",
#             })
#             return JsonResponse(context)
#         total_material_issue = 0.00
#         total_material_reciept = 0.00

#         for index in range(0,len(store_item)):
#             if store_item[index]['store__name'] not in data:
#                 data[store_item[index]['store__name']] =[]
#             if store_type == 'inHouse':
#                 data[store_item[index]['store__name']].append({
#                     'pk':store_item[index]['pk'],
#                     'on_hand_qty' : store_item[index]['on_hand_qty'],
#                     'item': store_item[index]['item__name'],
#                     'item_category' : store_item[index]['item__item_type__item_category__name'],
#                     'value' : float(store_item[index]['on_hand_qty']) * float(store_item[index]['item__price']),
#                 })
#             else:
#                 # #print(8618)
#                 if models.Job_Order_Detail.objects.filter(item_id = store_item[index]['item_id'],job_order_header__job_status =1).exists() :
#                     jobOrderDetails = models.Job_Order_Detail.objects.filter(item_id = store_item[index]['item_id'],job_order_header__job_status =1)
#                     for jobdet in jobOrderDetails:
#                         # #print(8622)
#                         total_material_issue += float(jobdet.quantity - jobdet.required_quantity) if (jobdet.job_order_header.material_issue == 2 and jobdet.required_quantity > 0) else 0.00
#                         # #print(jobdet.quantity_result if (jobdet.job_order_header.material_issue > 1 and jobdet.quantity_result > 0) else 0.00  )

#                         total_material_receipt += float(jobdet.quantity_result) if (jobdet.job_order_header.material_issue > 1 and jobdet.quantity_result > 0) else 0.00
#                         # #print(8626)
#                 data[store_item[index]['store__name']].append({
#                     'pk':store_item[index]['pk'],
#                     'on_hand_qty' : store_item[index]['on_hand_qty'],
#                     'issue_Quantity': total_material_issue,
#                     'reciept_left': total_material_receipt,
#                     'used_quantity' : total_material_issue + total_material_receipt,
#                     'actual_onhand' : float(store_item[index]['on_hand_qty']) - float(total_material_issue + total_material_receipt),
#                     'item': store_item[index]['item__name'],
#                     'item_category' : store_item[index]['item__item_type__item_category__name'],
#                     'value' : float(store_item[index]['on_hand_qty']) * float(store_item[index]['item__price']),
#                 })
        
#         context.update({
#             'status': 200,
#             'message': "Items Fetched Successfully.",
#             'page_items': data,
#         })

#     except Exception:
#         context.update({
#             'status': 592.1,
#             'message': "Internal Server Error",
#         })

#     return JsonResponse(context)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def reportInventoryStorewise(request):
    context = {}
    data = {}
    store_items = []
    try:
        # Fetch parameters
        item_cat_id = request.POST.get('item_cat_id', None)
        store_type = request.POST.get('store_type', None)
        total_material_issue = Decimal('0.00')
        total_material_receipt = Decimal('0.00')
        blocked_quantity = Decimal('0.00')
        vendor_id = request.POST.get('vendor_id',None)
        # Handle GET request
        if request.method == 'GET':
            store_items = list(
                models.Store_Item.objects.filter(status=1, deleted=0)
                .values(
                    'pk', 'on_hand_qty', 'item__item_type__item_category__name',
                    'item__price', 'store__name', 'item__name'
                )
            )
        else:
            # Handle POST request
            if store_type == 'inHouse':
                store_items = models.Store_Item.objects.filter(
                    item__item_type__item_category_id=item_cat_id,
                    store__vendor__isnull=True
                )
            else:
                if vendor_id is not None and (vendor_id != "" and vendor_id !="           ") :
                    
                    store_items = models.Store_Item.objects.filter(store__vendor_id = vendor_id)

                else:
                    store_items = models.Store_Item.objects.filter(
                    store__vendor__isnull=False
                )
                store_items = store_items.filter(
                    item__item_type__item_category_id=item_cat_id,
                    
                )
       
            store_items = list(
                store_items.values(
                    'pk', 'on_hand_qty', 'item__item_type__item_category__name',
                    'item__price', 'store__name', 'item__name', 'item_id'
                )
            )
           
        # Check if no items were found
        if not store_items:
            context.update({
                'status': 200,
                'message': "No items found.",
            })
            return JsonResponse(context)

        # Process store items
        for store_item in store_items:
            # #print(8896)
            total_material_issue = Decimal('0.00')
            total_material_receipt = Decimal('0.00')
            blocked_quantity = Decimal('0.00')
            store_name = store_item['store__name']
            if store_name not in data:
                data[store_name] = []

            if store_type == 'inHouse':
                # In-house store processing
                data[store_name].append({
                    'pk': store_item['pk'],
                    'on_hand_qty': store_item['on_hand_qty'],
                    'item': store_item['item__name'],
                    'item_category': store_item['item__item_type__item_category__name'],
                    'value': round((float(store_item['on_hand_qty']) * float(store_item['item__price'])),2),
                })
            else:
                # Vendor store processing
                # #print(total_material_issue)
                if models.Job_Order_Detail.objects.filter(
                    item_id=store_item['item_id'], job_order_header__job_status=1,job_order_header__vendor__store__name = store_item['store__name']
                ).exists():
                    jobOrderDetails = models.Job_Order_Detail.objects.filter(
                        item_id=store_item['item_id'], job_order_header__job_status=1,job_order_header__vendor__store__name = store_item['store__name']
                    )

                    for jobdet in jobOrderDetails:
                        # Update material issue and receipt totals
                        total_material_issue = Decimal('0.00')
                        total_material_receipt = Decimal('0.00')
                        blocked_quantity = Decimal('0.00')
                        
                        total_material_issue += (
                            (Decimal(jobdet.quantity - jobdet.required_quantity) )
                            if jobdet.job_order_header.material_issue > 1 and jobdet.direction == 'outgoing'
                            else Decimal('0.00')
                        )
                        blocked_quantity = (
                            (Decimal(jobdet.quantity - jobdet.required_quantity) - (Decimal(jobdet.quantity - jobdet.quantity_result)))
                            if jobdet.job_order_header.material_issue > 1 and jobdet.direction == 'outgoing'
                            else Decimal('0.00')
                        )
                        total_material_receipt += jobdet.quantity_result if jobdet.job_order_header.material_issue > 1 and jobdet.direction == 'incoming' else Decimal('0.00')

                        
                # # #print(total_material_issue,total_material_receipt)
                # Append vendor data
                data[store_name].append({
                    'pk': store_item['pk'],
                    'on_hand_qty': store_item['on_hand_qty'],
                    'issue_Quantity': round(float(total_material_issue),2),
                    'Wip_manufacture_Quantity' : round(float(total_material_receipt),2),
                    'Wip_issued_Quantity' : round(float(blocked_quantity),2),
                    'actual_onhand': round((float(store_item['on_hand_qty']) - float(blocked_quantity + total_material_receipt )),2),
                    'item': store_item['item__name'],
                    'item_category': store_item['item__item_type__item_category__name'],
                    'value': round((float(store_item['on_hand_qty']) * float(store_item['item__price'])),2),
                })
        # # #print(data)

        # Update context with data
        context.update({
            'status': 200,
            'message': "Items fetched successfully.",
            'page_items': data,
        })

    except Exception as e:
        # Log the exception for debugging purposes
        #print(f"Error: {e}")

        # Return internal server error
        context.update({

            'status': 500,
            'message': "Internal Server Error",
        })

    return JsonResponse(context)

@api_view(['GET','POST'])
@permission_classes([IsAuthenticated])
def reportStockTransfer(request):
    context = {}
    # # # # # #print(request.POST)
    from_date = request.POST.get('from_date', None)
    to_date = request.POST.get('to_date', None)
    data ={}
    on_transit_details =[]
    try:
        if request.method == 'GET':
            on_transit_details = list(models.On_Transit_Transaction_Details.objects.filter(status=1 , deleted = 0 ).values(
            'pk','item__name','on_transit_transaction_header__transaction_number',
            'on_transit_transaction_header__transaction_date',
            'on_transit_transaction_header__transaction_in_date',
            'on_transit_transaction_header__source_store__name',
            'on_transit_transaction_header__destination_store__name',
            'quantity',
            'recieved_quntity',
            
            ))
        else:
            # # # # # #print("5041")
            on_transit_details = list(models.On_Transit_Transaction_Details.objects.filter(on_transit_transaction_header__transaction_date__range=(from_date,to_date)).values(
            'pk','item__name','on_transit_transaction_header__transaction_number',
            'on_transit_transaction_header__transaction_date',
            'on_transit_transaction_header__transaction_in_date',
            'on_transit_transaction_header__source_store__name',
            'on_transit_transaction_header__destination_store__name',
            'quantity',
            'recieved_quntity',
            ))    
        # # # # # #print(on_transit_details)
        if(len(on_transit_details) == 0):
            context.update({
                'status': 200,
                'message': "no transaction found ",
            })
            return JsonResponse(context)

        for index in range(0,len(on_transit_details)):

            if on_transit_details[index]['on_transit_transaction_header__transaction_number'] not in data:
                data[on_transit_details[index]['on_transit_transaction_header__transaction_number']] =[]
            data[on_transit_details[index]['on_transit_transaction_header__transaction_number']].append({
                'pk':on_transit_details[index]['pk'],
                'item': on_transit_details[index]['item__name'],
                'source_store' : on_transit_details[index]['on_transit_transaction_header__source_store__name'],
                'destination_store' : on_transit_details[index]['on_transit_transaction_header__destination_store__name'],
                'material_out' : on_transit_details[index]['quantity'],
                'material_in' : on_transit_details[index]['recieved_quntity'] if float(on_transit_details[index]['recieved_quntity']) != 0.00 else "---", 
                'material_out_date' : on_transit_details[index]['on_transit_transaction_header__transaction_date'],
                'material_in_date' : on_transit_details[index]['on_transit_transaction_header__transaction_in_date'] if on_transit_details[index]['on_transit_transaction_header__transaction_in_date'] != None else "---"  ,
            })
        # # # # # #print(data)
        context.update({
            'status': 200,
            'message': "stock Transfer report fetch Successfully.",
            'page_items': data,
        })

    except Exception:
        context.update({
            'status': 538.1,
            'message': "Internal Server Error",
        })
    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportPurchaseOrderByVendor(request):
    context = {}
    item_id = request.POST.get('item_id', None)
    from_date = request.POST.get('from_date', None)
    to_date = request.POST.get('to_date', None)

    # Parse from_date and to_date strings to DateTime objects
    from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
    to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

    item = models.Item.objects.filter(pk=item_id)

    data = {}
    if item.exists():
        item = item.first()
        # Get all store_transaction_details for the item within the date range
        store_transaction_details = models.Store_Transaction_Detail.objects.filter(
            item=item,
            created_at__range=(from_date, to_date + timedelta(days=1)),  # Include to_date
        ).select_related('store_transaction_header', 'store_transaction_header__transaction_type')

        # Group store_transaction_details by store
        for store_transaction_detail in store_transaction_details:
            store = store_transaction_detail.store
            if store.name not in data:
                data[store.name] = []
            data[store.name].append({
                'quantity': store_transaction_detail.quantity,
                'rate': store_transaction_detail.rate,
                'amount': store_transaction_detail.amount,
                'gst_percentage': store_transaction_detail.gst_percentage,
                'amount_with_gst': store_transaction_detail.amount_with_gst,
                'transaction_type': store_transaction_detail.store_transaction_header.transaction_type.name,
                'updated_at': store_transaction_detail.updated_at.date()
            })

    context.update({
        'status': 200,
        'message': "Items Fetched Successfully.",
        'page_items': data,
    })

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportPurchaseOrderByVendor(request):
    context = {}
    vendor_id = request.POST.get('vendor_id', None)
    purchaseOrders = models.Purchase_Order.objects.filter(vendor_id=vendor_id)
    purchaseOrders = list(purchaseOrders.values('pk', 'order_number', 'order_date','total_amount','vendor__name'))

    context.update({
        'status': 200,
        'message': "Items Fetched Successfully.",
        'page_items': purchaseOrders,
    })

    return JsonResponse(context)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportPurchaseOrderByItem(request):
    context = {}
    item_id = request.POST.get('item_id', None)
    purchaseOrderDetails = models.Purchase_Order_Detail.objects.filter(item_id=item_id)
    purchaseOrderDetails = list(purchaseOrderDetails.values(
        'pk',
        'purchase_order_header__order_number',
        'purchase_order_header__order_date',
        'purchase_order_header__vendor__name',
        'quantity',
        'amount_with_gst',
        'delivered_quantity',
        'delivered_amount_with_gst'
    ))

    context.update({
        'status': 200,
        'message': "Items Fetched Successfully.",
        'page_items': purchaseOrderDetails,
    })

    return JsonResponse(context)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reportActivePurchaseOrder(request):
    context = {}
    purchaseOrders = models.Purchase_Order.objects.filter(delivery_status__in=[1, 2])
    purchaseOrders = list(purchaseOrders.values('pk', 'order_number', 'order_date','total_amount','vendor__name'))

    context.update({
        'status': 200,
        'message': "Items Fetched Successfully.",
        'page_items': purchaseOrders,
    })

    return JsonResponse(context)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def cornJobStoreItemQuantityUpdate(request):
    context ={}
    # # # # # #print('inside cornjob')
    # start_date = '2024-01-01' #for testing 
    end_date = datetime.now().date()
    start_date = end_date.replace(day=1)
    next_month_start_date = (start_date.replace(day=1) + timedelta(days=31))
    
    try:
        store_items = models.Store_Item.objects.select_related('store').prefetch_related(
            'store__store_transaction_detail_set'
        )
        for store_item in store_items:
            total_In_quantity = 0.00
            total_Out_quantity = 0.00
            total_Out_Value = 0.00
            total_IN_Value = 0.00
            total_quantity = float(store_item.opening_qty)
            total_value =  total_quantity * float(store_item.item.price)
            item_rate =0.00
            # Access the related store transaction details directly
            store_transaction_dets = store_item.store.store_transaction_detail_set.filter(
                item=store_item.item , store = store_item.store
            ).filter(store_transaction_header__transaction_date__range=(start_date,end_date))

            item_stock_report = models.Item_Stock_Report()
            item_stock_report.item = store_item.item
            item_stock_report.store = store_item.store
            item_stock_report.start_date = start_date
            item_stock_report.end_date = end_date
            item_stock_report.next_month_start_date = next_month_start_date
            item_stock_report.closing_quantity = store_item.closing_qty
            item_stock_report.closing_value = float(store_item.closing_qty) * float(store_item.item.price)
            item_stock_report.rate = store_item.item.price
            item_stock_report.save()
            if store_transaction_dets :
                order_details =[]
                for store_transact_det in store_transaction_dets:
                    transaction_type = store_transact_det.store_transaction_header.transaction_type.name
                    quantity = float(store_transact_det.quantity)
                    if transaction_type == 'GRN' or transaction_type == 'MIN' or transaction_type == 'GRNT' :
                        total_In_quantity += quantity
                        total_IN_Value += (float(store_transact_det.rate) * quantity) if float(store_transact_det.rate) > 0.00 else (float(store_transact_det.item.price) * quantity)
                    elif transaction_type == 'MIS' or transaction_type == 'MOUT':
                        
                        total_Out_quantity += quantity
                        total_Out_Value += (float(store_transact_det.rate) * quantity) if float(store_transact_det.rate) > 0.00 else (float(store_transact_det.item.price) * quantity)
                    else:
                        total_quantity -= quantity
                        total_value -= (float(store_transact_det.rate) * quantity) if float(store_transact_det.rate) > 0.00 else (float(store_transact_det.item.price) * quantity)
                    order_details.append(
                        models.Item_Stock_Report_Details(
                            item_stock_report_header_id = item_stock_report.id,
                            store_transaction_header_id = store_transact_det.store_transaction_header_id,
                            store_transaction_detail_id = store_transact_det.id,
                            transaction_type_id = store_transact_det.store_transaction_header.transaction_type_id,
                            transaction_number = store_transact_det.store_transaction_header.transaction_number,
                            transaction_date = store_transact_det.store_transaction_header.transaction_date,
                            quantity = store_transact_det.quantity,
                            rate = store_transact_det.rate,
                            value = float(store_transact_det.quantity) * float(store_transact_det.rate) 

                        )
                    )
                models.Item_Stock_Report_Details.objects.bulk_create(order_details)
                total_quantity = (total_quantity + total_In_quantity) - total_Out_quantity
                total_value = (total_value + total_IN_Value) - total_Out_Value
                item_rate = total_value/total_quantity if total_quantity >0.00 else 0.00
                item_stock_report.closing_quantity = total_quantity
                item_stock_report.closing_value = total_value
                item_stock_report.rate = item_rate
                item_stock_report.save()
                store_item = models.Store_Item.objects.get(item=store_item.item , store = store_item.store)
                store_item.opening_qty = total_quantity
                store_item.closing_qty  = total_quantity
                store_item.updated_at = datetime.now()
                store_item.save()

        context.update({
            'status': 200,
            'message': "store Item Quantity  updated sucessfully"
        })
    except Exception:

        context.update({
            'status': 539,
            'message': "Somethings went wrong please try again!"
        })
    

    return JsonResponse(context)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportPurchaseMaterailIssue(request):
    context = {}
    try:

        item_id = request.POST['item_id']
        store_id = request.POST['store_id']
        from_date = request.POST['from_date']
        to_date = request.POST['to_date']
        
        storeTransactionDetails = models.Store_Transaction_Detail.objects.filter(Q(store_transaction_header__transaction_type__name="MIS")|
                                    Q(store_transaction_header__transaction_type__name="GRN")).filter(
                                        store_transaction_header__transaction_date__range =(from_date,to_date)
                                    ).filter(
                                        item_id = item_id ,store_id =store_id
                                    ).order_by('store_transaction_header__transaction_date')

        storeTransactionDetails = list(storeTransactionDetails.all())
        data =[]

        item_stock_report_head_bool = models.Item_Stock_Report.objects.filter(store_id = store_id , 
                                        item_id = item_id).filter(next_month_start_date = from_date).exists()

        if(item_stock_report_head_bool):
            item_stock_report_head = models.Item_Stock_Report.objects.filter(store_id = store_id , 
                                    item_id = item_id).filter(next_month_start_date = from_date).first()
            closing_qty = item_stock_report_head.closing_quantity
        else:
            store_Item_head = models.Store_Item.objects.filter(store_id = store_id , 
                                    item_id = item_id).first()
            closing_qty = store_Item_head.opening_qty
        for index in range(0,len(storeTransactionDetails)):
            transaction_type_name = 'RECIEPT' if storeTransactionDetails[index].store_transaction_header.transaction_type.name == 'GRN' else 'ISSUE'
              
            if(transaction_type_name == 'RECIEPT'):
                closing_qty = (closing_qty) + (storeTransactionDetails[index].quantity) 
            else:
                closing_qty = (closing_qty) - (storeTransactionDetails[index].quantity) 
            # # # # # #print(closing_qty)
            data.append({
                'item_name': storeTransactionDetails[index].item.name,
                'transaction_date' :  storeTransactionDetails[index].store_transaction_header.transaction_date,
                'vendor_name': storeTransactionDetails[index].store_transaction_header.vendor.name,
                'transaction_type_name': transaction_type_name,
                'tranQuantity': format(storeTransactionDetails[index].quantity),
                'quantity' : format(closing_qty),
                'rate' :  format(storeTransactionDetails[index].rate),
                'amount' :format(closing_qty * (storeTransactionDetails[index].rate))
            })
        
        context.update({
            'status': 200,
            'message': "Items Fetched Successfully.",
            'page_items' : data
        })
    except Exception:
        context.update({
            'status': 540,
            'message': "Somethings went wrong please try again!",
        })
    return JsonResponse(context)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def reportVendorIssueRecp(request):
    context = {}
    try:

        item_id = request.POST['item_id']
        store_id = request.POST['store_id']
        from_date = request.POST['from_date']
        to_date = request.POST['to_date']
        
        storeTransactionDetails = models.Store_Transaction_Detail.objects.filter(Q(store_transaction_header__transaction_type__name="MIST")|
                                    Q(store_transaction_header__transaction_type__name="GRNT")).filter(
                                        store_transaction_header__transaction_date__range =(from_date,to_date)
                                    ).filter(
                                        item_id = item_id ,store_id =store_id
                                    ).order_by('store_transaction_header__transaction_date')
       
        storeTransactionDetails = list(storeTransactionDetails.all())
        data =[]

        item_stock_report_head_bool = models.Item_Stock_Report.objects.filter(store_id = store_id , 
                                        item_id = item_id).filter(next_month_start_date = from_date).exists()
     
        if(item_stock_report_head_bool):
            item_stock_report_head = models.Item_Stock_Report.objects.filter(store_id = store_id , 
                                    item_id = item_id).filter(next_month_start_date = from_date).first()
            closing_qty = item_stock_report_head.closing_quantity
            rate = item_stock_report_head.item.price
        else:
            store_Item_head = models.Store_Item.objects.filter(store_id = store_id , 
                                    item_id = item_id).first()
            closing_qty = store_Item_head.opening_qty
            rate = store_Item_head.item.price

        data.append({
                'item_name': '',
                'transaction_date' :  str(request.POST['from_date']),
                'jobOrder' : '',
                'transaction_type_name': 'OPENING',
                'tranquantity' : format(closing_qty),
                'totquantity' : format(closing_qty),
                'rate' : format( rate),
                'amount' :format(closing_qty * rate)
            })
        
        for index in range(0,len(storeTransactionDetails)):
            transaction_type_name = 'RECIEPT' if storeTransactionDetails[index].store_transaction_header.transaction_type.name == 'GRNT' else 'ISSUE'
            
            if(transaction_type_name == 'RECIEPT'):
                closing_qty = (closing_qty) + (storeTransactionDetails[index].quantity) 
            else:
                closing_qty = (closing_qty) - (storeTransactionDetails[index].quantity) 
            data.append({
                'item_name': storeTransactionDetails[index].item.name,
                'transaction_date' :  storeTransactionDetails[index].store_transaction_header.transaction_date,
                'jobOrder' : storeTransactionDetails[index].store_transaction_header.job_order.order_number,
                'transaction_type_name': transaction_type_name,
                'tranquantity' : format(storeTransactionDetails[index].quantity),
                'totquantity' : format(closing_qty),
                'rate' :  format(storeTransactionDetails[index].rate),
                'amount' :format(closing_qty * (storeTransactionDetails[index].rate))
            })
        context.update({
            'status': 200,
            'message': "Items Fetched Successfully.",
            'page_items' : data
        })
    except Exception:
        context.update({
            'status': 540,
            'message': "Somethings went wrong please try again!",
        })
    return JsonResponse(context)

def invoice_store_migration(store_id,user_id):
    try:
        with transaction.atomic():
            invoice_details = models.Invoice_Details.objects.filter(
                invoice_header__store_transaction_add=0, 
                invoice_header__status=1, 
                invoice_header__deleted=0, 
                status=1, 
                deleted=0
            ).select_related('invoice_header', 'item')  # Efficiently load related Invoice and Item in a single query
            if not invoice_details.exists():
                return "No modification done in store transaction as no new transaction found"
            old_id = -1
            # Process each invoice detail in a single loop
            for invoice_detail in invoice_details:
                if invoice_detail.invoice_header.id != old_id:
                    storeTransactionHeader = models.Store_Transaction()
                    storeTransactionHeader.transaction_type = models.Transaction_Type.objects.get(name='MIV')
                    storeTransactionHeader.transaction_number = invoice_detail.invoice_header.invoice_no
                    storeTransactionHeader.transaction_date = invoice_detail.invoice_header.date
                    storeTransactionHeader.total_amount = invoice_detail.invoice_header.total_value 
                    storeTransactionHeader.reference_id = invoice_detail.invoice_header.id
                    storeTransactionHeader.save()
                    old_id = invoice_detail.invoice_header.id
                  
                obj = models.Store_Transaction_Detail(
                        store_transaction_header_id=storeTransactionHeader.id,
                        item_id=invoice_detail.item_id,
                        store_id=store_id,
                        quantity=invoice_detail.quantity,
                        rate=invoice_detail.rate,
                        amount=invoice_detail.value,
                    )
                # # # # # #print(obj.__dict__) 
                obj.save()

                storeItem = models.Store_Item.objects.filter(
                    item_id=invoice_detail.item_id, store_id=store_id).first()

                if storeItem is None: 
                    return "some items is mising in this store please check"
                else:

                    storeItem.on_hand_qty -= invoice_detail.quantity
                    storeItem.closing_qty -= invoice_detail.quantity
                    storeItem.updated_at = datetime.now()
                    storeItem.save()
                # change in storeItemCurrent mout
                # Fetch the last transaction_date less than the given_date if mout
                given_date = invoice_detail.invoice_header.date
                
                
                # Check for the last record on the given_date
                record = models.Store_Item_Current.objects.filter(transaction_date=given_date,  item_id=invoice_detail.item_id, store_id=store_id,status=1, deleted=0 ).last()

                if not record:
                    # If no record is found for the given_date, look for the last record before that date
                    last_transaction_date = models.Store_Item_Current.objects.filter(
                        transaction_date__lt=given_date ,item_id=invoice_detail.item_id, store_id=store_id,status=1, deleted=0
                    ).aggregate(Max('transaction_date'))['transaction_date__max']

                    

                    if last_transaction_date:
                        # Fetch the record for the last_transaction_date
                        record = models.Store_Item_Current.objects.filter(
                            transaction_date=last_transaction_date ,item_id=invoice_detail.item_id, store_id=store_id,status=1, deleted=0
                        ).last()

                # Initialize a new instance of Store_Item_Current (Avoid shadowing the model name)
                
                
                store_item_instance = models.Store_Item_Current()

                if record:
                    # Set values based on the last record found
                    store_item_instance.opening_qty = record.closing_qty
                    store_item_instance.on_hand_qty = record.closing_qty - Decimal(invoice_detail.quantity)
                    store_item_instance.closing_qty = record.closing_qty - Decimal(invoice_detail.quantity)
                
                    if(store_item_instance.on_hand_qty<0):
                        raise ValueError(f"out quantity is more than available quantity")
                    # Set other fields for the new transaction
                    store_item_instance.store_transaction_id = storeTransactionHeader.id
                    store_item_instance.transaction_date = given_date
                    store_item_instance.item_id = invoice_detail.item_id
                    store_item_instance.store_id = store_id

                    # Save the instance to the database
                    store_item_instance.save()
                    
                else:
                    message = "canot possible no  item of this store present"
                    raise ValueError(message)
                store_item_curreEdit(store_id,invoice_detail.item_id,given_date,'mout',invoice_detail.quantity) #store_item_curreEdit(store_id, item_id, transaction_date,transact_type,quantity)
            # Update page items
            invoice_header = models.Invoice.objects.all()
            invoice_header.update(store_transaction_add=1)
            userId = user_id
            user_log_details_add(userId,'migration of data from invoice to store transaction done')
        return "Store transaction added successfuly "

    except Exception as e :
        return "Something went wrong!" 


@api_view(['POST'])
def extractDataFromXlsx(request):
    context = {}    
    storeItem = models.Store_Item.objects.filter(store_id=request.POST['store_id']).first()
    if storeItem is None:
        context.update({
            'status': 544,
            'message': "No item present in this particular store please add item on this store"
        })
        return JsonResponse(context)
    if request.FILES.get('file'):
        excel = request.FILES['file']
        flag = 1
        # trying to process files without error
        try:
            invoice = []
            workbook = load_workbook(excel)
            sheet = workbook.active
            row_number = next(cell.row for cell in sheet['A'] if isinstance(cell.value, str) and cell.value.lower() == 'date')
            mapping = {
                "date": None,
                "particulars":None,
                "voucher type":None,
                "voucher no.": None,
                "voucher ref. no.": None,
                "quantity": None,
                "value": None,
                "rate": None,
                "gross total": None,
                "gst sales 18%": None,
                "cgst @ 9%": None,
                "sgst @ 9%": None,
                "round off (+/-)": None,
                "igst sales@ 18%": None,
                "igst 18%": None,
                "rent": None,
                "export sales": None,
            }
            for cell in sheet[row_number]:
                cell_value = cell.value.lower()
                if cell_value in mapping:
                    mapping[cell_value] = cell.column-1
                if all(value is not None for value in mapping.values()):
                    break
            if row_number-1 :
                sheet.delete_rows(1, row_number - 1)
            if sheet.max_row > 0:
                sheet.delete_rows(sheet.max_row)
            i=2
            total_rows = sheet.max_row
            while i<=total_rows:
                header_row=sheet[i]
                if header_row[mapping["voucher type"]].value.lower() == 'sales':
                    try:
                        with transaction.atomic():
                            i+=1  
                            invoice_header=None
                            invoice_no = header_row[mapping["voucher no."]].value
                            if x := models.Invoice.objects.filter(invoice_no=invoice_no).first():
                                invoice_header = x
                            else:
                                invoice_header = models.Invoice()
                                if customers := models.Customer.objects.filter(name=header_row[mapping["particulars"]].value):
                                    invoice_header.customer = customers[0]
                                invoice_header.date = (header_row[mapping["date"]].value).date()
                                invoice_header.invoice_no = invoice_no
                                invoice_header.invoice_ref_no = header_row[mapping["voucher ref. no."]].value
                                invoice_header.total_quantity= Decimal(header_row[mapping["quantity"]].value)
                                invoice_header.total_quantity =  Decimal((header_row[mapping["quantity"]].value))
                                invoice_header.total_value = Decimal(header_row[mapping["value"]].value)
                                invoice_header.gross_total =  Decimal(header_row[mapping["gross total"]].value)     
                                invoice_header.gst_sales = Decimal(x) if (x := header_row[mapping["gst sales 18%"]].value) else 0          
                                invoice_header.cgst = Decimal(x) if (x := header_row[mapping["cgst @ 9%"]].value) else 0                  
                                invoice_header.sgst = Decimal(x) if (x := header_row[mapping["sgst @ 9%"]].value) else 0          
                                invoice_header.round_off = Decimal(x) if (x := header_row[mapping["round off (+/-)"]].value) else 0          
                                invoice_header.igst_sales = Decimal(x) if (x := header_row[mapping["igst sales@ 18%"]].value) else 0          
                                invoice_header.igst = Decimal(x) if (x := header_row[mapping["igst 18%"]].value) else 0          
                                invoice_header.rent = Decimal(x) if (x := header_row[mapping["rent"]].value) else 0          
                                invoice_header.export_sales = Decimal(x) if (x := header_row[mapping["export sales"]].value) else 0          
                                invoice_header.save()
                            invoice_details = []  
                            while i <= total_rows and not sheet[i][0].value:
                                detail_row=sheet[i]
                                i += 1
                                item = models.Item.objects.get(name=detail_row[mapping["particulars"]].value)
                                if not models.Invoice_Details.objects.filter(item=item, invoice_header_id=invoice_header.id).exists():
                                    invoice_details.append(
                                        models.Invoice_Details(
                                            invoice_header_id = invoice_header.id,
                                            item = item,
                                            quantity = handle_empty_cell(detail_row[mapping["quantity"]].value),
                                            rate = handle_empty_cell(detail_row[mapping["rate"]].value),
                                            value = handle_empty_cell(detail_row[mapping["value"]].value)
                                        )
                                    )
                                
                            if len(invoice_details):
                                # pass
                                models.Invoice_Details.objects.bulk_create(invoice_details)
                                flag = 0
                        transaction.commit()
                        context.update({
                            'status': 200,
                            'message': "Invoice added succesfully"
                        })
                        
                    except Exception:
                        # i+=1
                        flag = 1
                        context.update({
                            'status': 544,
                            'message': "Data could not be added to invoice_header table"
                        })
                        transaction.rollback()
                else:
                    i+=1    
                    flag = 0
            context.update({
                'status': 200,
                'message': "Excel read Successfully"+ ((" and " + context['message']) if context  else "")
            })
        except Exception as e:
            flag =1
            context.update({
                'status': 568,
                'message': "Error processing file"
            })
    else:
        flag =1
        context.update({
            'status': 568,
            'message': "File has not been uploaded"
        })
    if flag == 0 :
        message = invoice_store_migration(request.POST['store_id'],request.COOKIES.get('userId', None))
        context.update({
            'status': 200,
            'message': "Process Completed" +  ((", " + context['message']) if context  else "") + " and " + message 
        })
    else :
        context.update({
            'status': 200,
            'message': "Process Not Completed" +  ((", " + context['message']) if context  else "")  
        })
    return JsonResponse(context)


def handle_transaction_detail(detail, transact_type_name):
    """
    Handles the creation of Store_Item_Current records for a given transaction detail.
    """
    # Fetch the last record for the item and store up to the transaction date
    last_record =models.Store_Item_Current.objects.filter(
            transaction_date=detail.store_transaction_header.transaction_date,
            item_id=detail.item.id,
            store_id=detail.store.id,status=1, deleted=0
        ).order_by('transaction_date', 'created_at','id')
    if not last_record:
        last_record =models.Store_Item_Current.objects.filter(
            transaction_date__lt=detail.store_transaction_header.transaction_date,
            item_id=detail.item.id,
            store_id=detail.store.id,status=1, deleted=0
        ).order_by('transaction_date', 'created_at','id')

    # Create a new Store_Item_Current instance
    store_item_instance = models.Store_Item_Current()
    if last_record:
        last_record = last_record.last()
        # Set values based on the last record
        store_item_instance.opening_qty = last_record.closing_qty
        if transact_type_name == 'GRN' or  transact_type_name == 'GRNT' or  transact_type_name == 'MIN' or  transact_type_name == 'SP' :
            store_item_instance.on_hand_qty = last_record.closing_qty + Decimal(detail.quantity)
            store_item_instance.closing_qty = last_record.closing_qty + Decimal(detail.quantity)
        else:  # For other transaction types
            store_item_instance.on_hand_qty = last_record.closing_qty - Decimal(detail.quantity)
            store_item_instance.closing_qty = last_record.closing_qty - Decimal(detail.quantity)
            if (store_item_instance.closing_qty < Decimal(0.00) ) or (last_record.closing_qty < Decimal(0.00)):
                # #print(last_record.closing_qty)
                raise ValueError(f"Negetive stock come on item{detail.item.name} of store {detail.store.name} of date {detail.store_transaction_header.transaction_date} last record {last_record.closing_qty}")
    else:
        # Set default values if no prior record exists
        if transact_type_name == 'GRN' or  transact_type_name == 'GRNT' or  transact_type_name == 'MIN' or  transact_type_name == 'SP':
            store_item_instance.opening_qty = Decimal(0.00)
            store_item_instance.on_hand_qty = Decimal(detail.quantity)
            store_item_instance.closing_qty = Decimal(detail.quantity)
        else:
            raise ValueError(f"Cannot process transaction {detail.store_transaction_header.transaction_number}: No prior record found for item. of Transaction date.{detail.store_transaction_header.transaction_date}, item_name:{detail.item.name} of store {detail.store.name} ")
    if(store_item_instance.on_hand_qty<0):
        raise ValueError(f"onhand quantity is less than 0")
    # Set common fields
    store_item_instance.store_transaction_id = detail.store_transaction_header.id
    store_item_instance.transaction_date = detail.store_transaction_header.transaction_date
    store_item_instance.item_id = detail.item.id
    store_item_instance.store_id = detail.store.id
    store_item_instance.save()

    # Update stock tracking
    if transact_type_name == 'GRN' or  transact_type_name == 'GRNT' or  transact_type_name == 'MIN' or  transact_type_name == 'SP':
        store_item_curreEdit(detail.store.id, detail.item.id, detail.store_transaction_header.transaction_date, 'min', detail.quantity)
    else:
        store_item_curreEdit(detail.store.id, detail.item.id, detail.store_transaction_header.transaction_date, 'mout', detail.quantity)

@api_view(['POST'])
def storeItemCurrentMigrate(request):
    context = {}
    #print("processing ....")
    try:
        # Debugging: Print POST data
        # #print(request.POST)

        # Get the transaction date and convert to datetime with max time
        given_date_str = request.POST['transaction_date']
        given_date = datetime.strptime(given_date_str, '%Y-%m-%d').date()
        given_date = datetime.combine(given_date, datetime.max.time())  # Combine with max time for end of day

        # item_id = request.POST['item_id']

        # store_id = request.POST['store_id']
        # Get the transaction type ID
        transaction_type_id = request.POST['transaction_type_id']

        # Begin transaction
        with transaction.atomic():
            # Fetch transaction type name
            transact_type_name = models.Transaction_Type.objects.filter(pk=transaction_type_id).first()
            if not transact_type_name:
                raise ValueError("Invalid transaction type ID")

            # Fetch store transaction details with filtering and ordering
            store_transaction_details = models.Store_Transaction_Detail.objects.filter(
                store_transaction_header__status=1,
                store_transaction_header__deleted=0,
                store_transaction_header__created_at__lte=given_date,  # Filter by created_at as datetime
                store_transaction_header__transaction_type_id=transaction_type_id,
                item__status=1,
                item__deleted=0
            ).select_related('store_transaction_header', 'item', 'store').order_by(
                'store_transaction_header__transaction_date', 'store_transaction_header__created_at'
            )

            # Debugging: Print the details of fetched store transactions
            # #print(store_transaction_details)

            # Process each transaction detail
            for detail in store_transaction_details:
                #print(f"Working Start on {detail.store_transaction_header.transaction_number}...")
                handle_transaction_detail(detail, transact_type_name.name)
                #print(f"Working Complete on {detail.store_transaction_header.transaction_number}")
        transaction.commit()
        # Return success response
        context.update({
            'status': 200,
            'message': 'All migrations completed. Check stock tracking.'
        })
    except Exception as e:
        # Handle exceptions and return error response
        context.update({
            'status': 500,
            'message': f"An error occurred: {str(e)}"
        })
        transaction.rollback()
    return JsonResponse(context)


